#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parse Chinese purchase contract Excel files (.xls or .xlsx) and output structured JSON.

Template (fixed):
  Row 1 [1,7]  : 合同编号
  Row 2 [2,2]  : 日期 (Excel date serial or string)
  Row 3 [3,2]  : 供方名称           [3,7] : 需方名称
  Row 4 [4,2]  : 供方地址           [4,7] : 需方地址
  Row 5 [5,2]  : 供方联系人         [5,7] : 需方联系人
  Row 6 [6,2]  : 供方电话           [6,7] : 需方电话
  Row 10       : Header row
  Row 11..     : Item rows (variable count) until 小写合计
  Item columns:
    0 name_cn | 1 fba_sku(FNSKU) | 2 spec | 3 unit | 4 quantity
    5 packing_rate | 6 unit_price_with_tax | 7 package_size | 8 net_weight_kg
    9 gross_weight_kg | 10 total_amount
"""

import json
import argparse
import sys
import re
import datetime as _dt
from pathlib import Path

try:
    import xlrd
    from xlrd import xldate as _xldate
except ImportError:
    print("Error: xlrd not found. Install it with: pip install xlrd")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not found. Install it with: pip install openpyxl")
    sys.exit(1)


# Fixed cell positions in the new purchase-contract template
CELL_CONTRACT_NO = (1, 7)
CELL_DATE = (2, 2)
CELL_SUPPLIER_NAME = (3, 2)
CELL_SUPPLIER_ADDRESS = (4, 2)
CELL_SUPPLIER_CONTACT = (5, 2)
CELL_SUPPLIER_PHONE = (6, 2)
CELL_BUYER_NAME = (3, 7)
CELL_BUYER_ADDRESS = (4, 7)
CELL_BUYER_CONTACT = (5, 7)
CELL_BUYER_PHONE = (6, 7)

HEADER_ROW = 10
ITEM_START_ROW = 11

COL_NAME = 0
COL_FBA_SKU = 1
COL_SPEC = 2
COL_UNIT = 3
COL_QTY = 4
COL_PACKING_RATE = 5
COL_UNIT_PRICE = 6
COL_PACKAGE_SIZE = 7
COL_NET_WEIGHT = 8
COL_GROSS_WEIGHT = 9
COL_TOTAL_AMOUNT = 10


def extract_city_from_supplier(supplier_name):
    """Extract city from supplier name (e.g. '义乌市XXX' -> '义乌')."""
    if not supplier_name:
        return ""
    m = re.search(r'(\w+市)', supplier_name)
    if m:
        return m.group(1).replace('市', '')
    m = re.search(r'(\w+县)', supplier_name)
    if m:
        return m.group(1)
    return ""


def parse_package_size(size_str):
    """Parse '122*63.5*4' or '43*45*54cm' -> [122.0, 63.5, 4.0]."""
    if not size_str:
        return None
    s = str(size_str)
    # Strip 'cm'/'CM' (case-insensitive) anywhere in the string
    s = re.sub(r'[cC][mM]', '', s).strip()
    try:
        return [float(x.strip()) for x in s.split('*')]
    except (ValueError, AttributeError):
        return None


def load_excel_sheet(file_path):
    """
    Load Excel sheet from .xls or .xlsx file.
    Returns (sheet, is_openpyxl, datemode). datemode is None for xlsx.
    """
    file_path = Path(file_path)
    if file_path.suffix.lower() == '.xls':
        workbook = xlrd.open_workbook(str(file_path))
        sheet = workbook.sheet_by_index(0)
        return sheet, False, workbook.datemode
    elif file_path.suffix.lower() == '.xlsx':
        workbook = load_workbook(str(file_path))
        sheet = workbook.active
        return sheet, True, None
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}")


def get_cell_value(sheet, row, col, is_openpyxl):
    if is_openpyxl:
        return sheet.cell(row + 1, col + 1).value
    return sheet.cell_value(row, col)


def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _format_date(value, datemode):
    """Convert an Excel date cell (serial number or string or datetime) to 'YYYY-MM-DD'."""
    if value is None or value == "":
        return ""
    if isinstance(value, _dt.datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, _dt.date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        # Excel serial number (xls path)
        if datemode is None:
            datemode = 0
        dt = _xldate.xldate_as_datetime(float(value), datemode)
        return dt.strftime('%Y-%m-%d')
    return str(value).strip()


def _format_phone(value):
    """Phone may be float (xlrd reads numeric cells as float). Convert to plain string."""
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_product_name(name_str):
    """Take only the first line (Chinese name); English name comes from the knowledge base."""
    if not name_str:
        return ""
    return str(name_str).split('\n')[0].strip()


def parse_purchase_contract(input_file, output_file=None):
    """Parse a purchase contract Excel file and return a dict."""
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    sheet, is_openpyxl, datemode = load_excel_sheet(input_file)

    def cv(rc):
        return get_cell_value(sheet, rc[0], rc[1], is_openpyxl)

    # Metadata
    contract_no = _to_str(cv(CELL_CONTRACT_NO))
    date = _format_date(cv(CELL_DATE), datemode)

    supplier_name = _to_str(cv(CELL_SUPPLIER_NAME))
    supplier_address = _to_str(cv(CELL_SUPPLIER_ADDRESS))
    supplier_contact = _to_str(cv(CELL_SUPPLIER_CONTACT))
    supplier_phone = _format_phone(cv(CELL_SUPPLIER_PHONE))
    supplier_city = extract_city_from_supplier(supplier_name)

    buyer_name = _to_str(cv(CELL_BUYER_NAME))
    buyer_address = _to_str(cv(CELL_BUYER_ADDRESS))
    buyer_contact = _to_str(cv(CELL_BUYER_CONTACT))
    buyer_phone = _format_phone(cv(CELL_BUYER_PHONE))

    # Items (variable row count) — read from ITEM_START_ROW until '小写合计'
    nrows = sheet.max_row if is_openpyxl else sheet.nrows
    items = []
    for row_idx in range(ITEM_START_ROW, nrows):
        first_cell = _to_str(get_cell_value(sheet, row_idx, 0, is_openpyxl))
        if '小写合计' in first_cell or '合计' in first_cell:
            break

        name = get_cell_value(sheet, row_idx, COL_NAME, is_openpyxl)
        if not name:
            continue

        spec = get_cell_value(sheet, row_idx, COL_SPEC, is_openpyxl)
        fba_sku = get_cell_value(sheet, row_idx, COL_FBA_SKU, is_openpyxl)
        unit = get_cell_value(sheet, row_idx, COL_UNIT, is_openpyxl)
        quantity = get_cell_value(sheet, row_idx, COL_QTY, is_openpyxl)
        packing_rate = get_cell_value(sheet, row_idx, COL_PACKING_RATE, is_openpyxl)
        unit_price = get_cell_value(sheet, row_idx, COL_UNIT_PRICE, is_openpyxl)
        package_size = get_cell_value(sheet, row_idx, COL_PACKAGE_SIZE, is_openpyxl)
        net_weight = get_cell_value(sheet, row_idx, COL_NET_WEIGHT, is_openpyxl)
        gross_weight = get_cell_value(sheet, row_idx, COL_GROSS_WEIGHT, is_openpyxl)
        total_amount = get_cell_value(sheet, row_idx, COL_TOTAL_AMOUNT, is_openpyxl)

        item = {
            "name_cn": parse_product_name(name),
            "spec": _to_str(spec),
            "fba_sku": _to_str(fba_sku).replace('\n', ''),
            "unit": _to_str(unit),
            "quantity": int(quantity) if isinstance(quantity, (int, float)) else 0,
            "packing_rate": int(packing_rate) if isinstance(packing_rate, (int, float)) else 0,
            "unit_price_with_tax": float(unit_price) if isinstance(unit_price, (int, float)) else 0.0,
            "package_size_cm": parse_package_size(package_size),
            "net_weight_kg": float(net_weight) if isinstance(net_weight, (int, float)) else 0.0,
            "gross_weight_kg": float(gross_weight) if isinstance(gross_weight, (int, float)) else 0.0,
            "total_amount": float(total_amount) if isinstance(total_amount, (int, float)) else 0.0,
        }
        items.append(item)

    # Grand total — find '共计' row, take col 10
    grand_total = 0.0
    for row_idx in range(nrows):
        first_cell = _to_str(get_cell_value(sheet, row_idx, 0, is_openpyxl))
        if '共计' in first_cell:
            v = get_cell_value(sheet, row_idx, COL_TOTAL_AMOUNT, is_openpyxl)
            if isinstance(v, (int, float)):
                grand_total = float(v)
            break

    return {
        "contract_no": contract_no,
        "date": date,
        "supplier": {
            "name": supplier_name,
            "city": supplier_city,
            "address": supplier_address,
            "contact": supplier_contact,
            "phone": supplier_phone,
        },
        "buyer": {
            "name": buyer_name,
            "address": buyer_address,
            "contact": buyer_contact,
            "phone": buyer_phone,
        },
        "items": items,
        "grand_total": grand_total,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse Chinese purchase contract Excel files and output structured JSON"
    )
    parser.add_argument("input_file", help="Path to input Excel file (.xls or .xlsx)")
    parser.add_argument(
        "--output",
        default="/tmp/purchase_contract.json",
        help="Path to output JSON file (default: /tmp/purchase_contract.json)",
    )
    args = parser.parse_args()

    try:
        result = parse_purchase_contract(args.input_file)
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Success! Output written to: {output_path}")
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(f"Fatal error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
