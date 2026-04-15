#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parse a Chinese purchase contract Excel (.xls / .xlsx) via OpenRouter GPT-5.4.

Pipeline:
  1. Read every cell of the first sheet into a 2-D grid (values only, with
     Excel date serials / numeric phone cells normalized to strings).
  2. Expand merged ranges so each logical cell carries the visible value.
  3. Serialize the grid into a markdown-style table (one row per line,
     prefixed with `R<n> |`) — this is the LLM input.
  4. Send once to openai/gpt-5.4 with retry-on-rate-limit, expect a strict
     JSON object matching the schema consumed by scripts/generate_all.py.

Output schema (unchanged from the legacy parser):
{
  "contract_no": str,
  "date": "YYYY-MM-DD",
  "supplier": {"name","city","address","contact","phone"},
  "buyer":    {"name","address","contact","phone"},
  "items": [ {"name_cn","spec","fba_sku","unit","quantity","packing_rate",
              "unit_price_with_tax","package_size_cm","net_weight_kg",
              "gross_weight_kg","total_amount"} ],
  "grand_total": float
}

Dependencies:
    pip install openai xlrd openpyxl --break-system-packages -q
"""

import argparse
import datetime as _dt
import json
import os
import random
import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# --- Config ---------------------------------------------------------------
OPENROUTER_API_KEY = "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"  # placeholder; prefer env var
OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"
MODEL = "openai/gpt-5.4"
MAX_RETRIES = 6
RETRY_BASE_DELAY = 2.0
REQUEST_TIMEOUT = 180


SYSTEM_PROMPT = """你是识别中国境内跨境电商采购合同 Excel 的专家。
用户会把一份采购合同 Excel 按**行号 + 单元格**序列化成文本表格发给你（每一行前面带 `R<n> |` 行号标记）。
你必须只输出一个严格的 JSON 对象，不要任何解释、前后缀或 markdown 代码块。

============================================================
版面说明（当前模板）
============================================================

顶部 metadata 区（大约 R1–R8）：
  - **合同编号**：常见标签"合同编号"、"订单号"、"PO"等；值是形如 `PO2604080482` 的字符串。
  - **日期**：常见标签"日期"、"签订日期"；输出统一为 `YYYY-MM-DD` 字符串。
  - **供方 / 需方区**（供方在左、需方在右，或上下排列）。每方各含：
    · 名称（"供方"/"需方"、"甲方"/"乙方"、"卖方"/"买方"等标签后的公司名）
    · 地址
    · 联系人
    · 电话（纯数字字符串，**不要**科学计数法、不要 `.0`）

  若某字段单元格为空或找不到，字符串字段填空字符串 `""`，不要编造。

**供方城市（supplier.city）**：从供方名称中提取，规则：
  - 如果名称含"XX市"，取"XX"（去掉"市"字）。例："义乌市XXX有限公司" → "义乌"、"上海飞蔻电子商务有限公司" → "上海"。
  - 如果名称含"XX县"，取"XX县"（保留"县"字）。
  - 都没有则填 `""`。

商品表格区：
  - 表头行通常在 metadata 下方（第 9–11 行左右），包含以下列（列名可能略有差异，但含义一致；列顺序以实际表头为准）：
    · **产品名称 / 品名**  → `name_cn`
    · **FNSKU / FBA SKU / SKU**  → `fba_sku`
    · **规格 / 型号 / 颜色/款式**  → `spec`
    · **单位**  → `unit`（如"片"、"件"、"个"、"套"）
    · **数量**  → `quantity`（整数）
    · **箱率 / 装箱率 / 每箱数量**  → `packing_rate`（整数，每箱装几件）
    · **含税单价 / 单价**  → `unit_price_with_tax`（浮点）
    · **外箱尺寸 / 包装尺寸 / 纸箱尺寸**  → `package_size_cm`（见下）
    · **单箱净重 / 净重**  → `net_weight_kg`（浮点）
    · **单箱毛重 / 毛重**  → `gross_weight_kg`（浮点）
    · **金额 / 总金额 / 小计**  → `total_amount`（浮点）

  字段处理规则：
    - `name_cn`：若单元格内含换行，只取**第一行**。
    - `fba_sku`：去掉内部换行，保留原样字符串（区分大小写）。
    - `package_size_cm`：解析形如 `122*63.5*4`、`43*45*54cm`、`122×63.5×4` 的字符串为 `[长, 宽, 高]` 浮点数组（单位 cm，**不含** "cm"、"CM"、"厘米"）。若无法解析或单元格为空 → 输出 `null`。
    - 数值字段若单元格为空或非数字 → 填 `0` 或 `0.0`。

  表格结束标志：
    - 数据行在 `小写合计`、`合计`、`共计` 所在行**之前**结束。这些行本身不是商品，**不要**作为 item。
    - **`grand_total`**：在 `共计` 那一行的金额列读取（即与商品表格同一金额列）。如果只找到 `小写合计`，也可以用它。
    - 空白行、分隔行跳过。

============================================================
输出 JSON Schema（严格遵守字段名和类型）
============================================================

{
  "contract_no": "PO2604080482",
  "date": "2026-04-07",
  "supplier": {
    "name": "上海飞蔻电子商务有限公司",
    "city": "上海",
    "address": "",
    "contact": "赵守彪",
    "phone": "18601710016"
  },
  "buyer": {
    "name": "深圳市艾进贸易有限公司",
    "address": "",
    "contact": "朱妍桥",
    "phone": "13509647412"
  },
  "items": [
    {
      "name_cn": "墙皮",
      "spec": "白色粗布纹",
      "fba_sku": "3D-WP-TEXTURED",
      "unit": "片",
      "quantity": 800,
      "packing_rate": 4,
      "unit_price_with_tax": 40.2475,
      "package_size_cm": [122.0, 63.5, 4.0],
      "net_weight_kg": 13.28,
      "gross_weight_kg": 14.68,
      "total_amount": 32198.0
    }
  ],
  "grand_total": 85394.0
}

注意：`items` 数组的每个元素必须包含以上 11 个字段，缺一不可。
"""


# --- Excel → markdown grid ------------------------------------------------
def _fmt_date(value, datemode) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            from xlrd import xldate
            dt = xldate.xldate_as_datetime(float(value), datemode or 0)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    return str(value).strip()


def _looks_like_date_serial(value) -> bool:
    """Heuristic: Excel date serials for 2000~2099 are roughly 36526~73050."""
    return isinstance(value, (int, float)) and 30000 < float(value) < 80000


def _cell_to_str(value, datemode=None, is_date_hint: bool = False) -> str:
    """Normalize a cell value to a human-readable string, preserving precision."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if is_date_hint and _looks_like_date_serial(value):
            return _fmt_date(value, datemode)
        # Integers stored as float (e.g. phone numbers, quantities)
        if value.is_integer():
            return str(int(value))
        # Avoid scientific notation; strip trailing zeros
        s = f"{value:.10f}".rstrip("0").rstrip(".")
        return s or "0"
    return str(value).strip()


def load_xls_grid(path: Path) -> Tuple[List[List[str]], int]:
    """Load .xls via xlrd. Returns (grid, datemode)."""
    import xlrd  # noqa: F401
    wb = xlrd.open_workbook(str(path), formatting_info=False)
    sheet = wb.sheet_by_index(0)
    datemode = wb.datemode
    nrows, ncols = sheet.nrows, sheet.ncols

    # Build raw grid
    grid: List[List[str]] = []
    for r in range(nrows):
        row: List[str] = []
        for c in range(ncols):
            ctype = sheet.cell_type(r, c)
            val = sheet.cell_value(r, c)
            is_date = ctype == xlrd.XL_CELL_DATE  # type: ignore[attr-defined]
            row.append(_cell_to_str(val, datemode=datemode, is_date_hint=is_date))
        grid.append(row)

    # Expand merged ranges (xlrd gives (rlo, rhi, clo, chi), hi exclusive)
    try:
        merged = sheet.merged_cells  # requires formatting_info but available sometimes
    except Exception:
        merged = []
    for (rlo, rhi, clo, chi) in merged:
        top = grid[rlo][clo]
        if not top:
            continue
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                if r < len(grid) and c < len(grid[r]) and not grid[r][c]:
                    grid[r][c] = top

    return grid, ncols


def load_xlsx_grid(path: Path) -> Tuple[List[List[str]], int]:
    """Load .xlsx via openpyxl. Returns (grid, ncols)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True)
    sheet = wb.active
    nrows, ncols = sheet.max_row, sheet.max_column

    grid: List[List[str]] = []
    for r in range(1, nrows + 1):
        row: List[str] = []
        for c in range(1, ncols + 1):
            val = sheet.cell(r, c).value
            row.append(_cell_to_str(val))
        grid.append(row)

    # Expand merged ranges
    for mr in sheet.merged_cells.ranges:
        rlo, clo, rhi, chi = mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1
        top = grid[rlo][clo] if rlo < len(grid) and clo < len(grid[rlo]) else ""
        if not top:
            continue
        for r in range(rlo, rhi + 1):
            for c in range(clo, chi + 1):
                if r < len(grid) and c < len(grid[r]) and not grid[r][c]:
                    grid[r][c] = top

    return grid, ncols


def load_grid(path: Path) -> Tuple[List[List[str]], int]:
    ext = path.suffix.lower()
    if ext == ".xls":
        return load_xls_grid(path)
    if ext == ".xlsx":
        return load_xlsx_grid(path)
    raise ValueError(f"Unsupported extension: {ext}")


def grid_to_markdown(grid: List[List[str]], ncols: int) -> str:
    """Serialize grid as `R<n> | col1 | col2 | ... |` lines."""
    lines: List[str] = []
    # Column letter header helps LLM reference positions
    col_header = "row | " + " | ".join(f"C{c + 1}" for c in range(ncols)) + " |"
    sep = "----|" + "----|" * ncols
    lines.append(col_header)
    lines.append(sep)
    for i, row in enumerate(grid, 1):
        # Replace pipe and newline characters inside cells so the markdown stays clean
        cells = [c.replace("|", "/").replace("\n", " ⏎ ") for c in row]
        # Pad/truncate to ncols
        if len(cells) < ncols:
            cells = cells + [""] * (ncols - len(cells))
        else:
            cells = cells[:ncols]
        lines.append(f"R{i} | " + " | ".join(cells) + " |")
    return "\n".join(lines)


# --- LLM call -------------------------------------------------------------
def _get_client(api_key: Optional[str]):
    try:
        from openai import OpenAI
    except ImportError:
        print("Error: openai SDK not installed. Run: pip install openai --break-system-packages",
              file=sys.stderr)
        sys.exit(1)
    key = api_key or os.environ.get("OPENROUTER_API_KEY") or OPENROUTER_API_KEY
    if not key or key == "REPLACE_WITH_OPENROUTER_API_KEY":
        print("Error: no OpenRouter API key.", file=sys.stderr)
        sys.exit(1)
    return OpenAI(base_url=OPENROUTER_BASE_URL, api_key=key)


def _is_retryable(err: Exception) -> bool:
    msg = str(err).lower()
    if "rate" in msg and "limit" in msg:
        return True
    for marker in ("429", "too many", "timeout", "timed out", "502", "503", "504", "overload", "temporarily"):
        if marker in msg:
            return True
    return False


def _extract_json(text: str) -> Dict:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{[\s\S]*\}", text)
        if not m:
            raise ValueError(f"No JSON object in model output: {text[:200]!r}")
        return json.loads(m.group(0))


def call_llm(client, filename: str, markdown_grid: str) -> Dict:
    user_msg = (
        f"Excel 文件名：{filename}\n\n"
        f"下面是整张工作表的文本化内容（每行前有 R<n> 行号，列标记为 C<n>）：\n\n"
        f"{markdown_grid}\n\n"
        f"请严格按系统指令输出 JSON。"
    )
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_msg},
    ]

    last_err: Optional[Exception] = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                response_format={"type": "json_object"},
                timeout=REQUEST_TIMEOUT,
            )
            content = resp.choices[0].message.content or ""
            return _extract_json(content)
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt >= MAX_RETRIES or not _is_retryable(e):
                if attempt >= MAX_RETRIES:
                    break
                raise
            delay = RETRY_BASE_DELAY * (2 ** (attempt - 1)) + random.uniform(0, 1.0)
            print(f"  transient error ({e}); retry {attempt}/{MAX_RETRIES} in {delay:.1f}s",
                  file=sys.stderr)
            time.sleep(delay)
    raise RuntimeError(f"LLM call failed after {MAX_RETRIES} retries: {last_err}")


# --- Post-processing ------------------------------------------------------
ITEM_FIELDS = [
    "name_cn", "spec", "fba_sku", "unit", "quantity", "packing_rate",
    "unit_price_with_tax", "package_size_cm", "net_weight_kg",
    "gross_weight_kg", "total_amount",
]


def _as_int(v) -> int:
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        try:
            return int(float(v.strip()))
        except Exception:
            return 0
    return 0


def _as_float(v) -> float:
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except Exception:
            return 0.0
    return 0.0


def _as_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _coerce_package_size(v):
    if v is None:
        return None
    if isinstance(v, list):
        try:
            return [float(x) for x in v]
        except Exception:
            return None
    if isinstance(v, str):
        s = re.sub(r"[cC][mM]|厘米", "", v).strip()
        s = s.replace("×", "*").replace("X", "*").replace("x", "*")
        try:
            return [float(x.strip()) for x in s.split("*") if x.strip()]
        except Exception:
            return None
    return None


def normalize_output(data: Dict) -> Dict:
    """Coerce LLM output to the exact schema and types the downstream expects."""
    out: Dict = {}
    out["contract_no"] = _as_str(data.get("contract_no"))
    out["date"] = _as_str(data.get("date"))

    sup = data.get("supplier") or {}
    out["supplier"] = {
        "name": _as_str(sup.get("name")),
        "city": _as_str(sup.get("city")),
        "address": _as_str(sup.get("address")),
        "contact": _as_str(sup.get("contact")),
        "phone": _as_str(sup.get("phone")),
    }

    buyer = data.get("buyer") or {}
    out["buyer"] = {
        "name": _as_str(buyer.get("name")),
        "address": _as_str(buyer.get("address")),
        "contact": _as_str(buyer.get("contact")),
        "phone": _as_str(buyer.get("phone")),
    }

    items_in = data.get("items") or []
    items_out: List[Dict] = []
    for it in items_in:
        items_out.append({
            "name_cn": _as_str(it.get("name_cn")).split("\n")[0].strip(),
            "spec": _as_str(it.get("spec")),
            "fba_sku": _as_str(it.get("fba_sku")).replace("\n", ""),
            "unit": _as_str(it.get("unit")),
            "quantity": _as_int(it.get("quantity")),
            "packing_rate": _as_int(it.get("packing_rate")),
            "unit_price_with_tax": _as_float(it.get("unit_price_with_tax")),
            "package_size_cm": _coerce_package_size(it.get("package_size_cm")),
            "net_weight_kg": _as_float(it.get("net_weight_kg")),
            "gross_weight_kg": _as_float(it.get("gross_weight_kg")),
            "total_amount": _as_float(it.get("total_amount")),
        })
    out["items"] = items_out
    out["grand_total"] = _as_float(data.get("grand_total"))
    return out


# --- Main -----------------------------------------------------------------
def parse_contract_with_llm(path: str, api_key: Optional[str] = None) -> Dict:
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(path)

    print(f"Loading Excel: {path}", file=sys.stderr)
    grid, ncols = load_grid(p)
    markdown = grid_to_markdown(grid, ncols)
    print(f"  grid: {len(grid)} rows × {ncols} cols; serialized {len(markdown)} chars",
          file=sys.stderr)

    client = _get_client(api_key)
    print(f"  calling {MODEL} ...", file=sys.stderr)
    raw = call_llm(client, p.name, markdown)
    return normalize_output(raw)


def main():
    ap = argparse.ArgumentParser(
        description="Parse a Chinese purchase contract Excel via OpenRouter GPT-5.4."
    )
    ap.add_argument("input_file", help="Path to contract Excel (.xls / .xlsx)")
    ap.add_argument("--output", default="/tmp/purchase_contract.json",
                    help="Output JSON path (default: /tmp/purchase_contract.json)")
    ap.add_argument("--api-key", default=None,
                    help="OpenRouter API key (overrides env var)")
    args = ap.parse_args()

    result = parse_contract_with_llm(args.input_file, api_key=args.api_key)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"Output written to {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
