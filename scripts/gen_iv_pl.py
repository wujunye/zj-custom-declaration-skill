#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Invoice & Packing List (IV&PL) Excel document.
"""

import os
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from helpers import get_lwh, sku_key

# Shared thin border (line_style=1 = 'thin', color index 64 = automatic/black)
_thin = Side(style='thin')
_no_side = Side(style=None)
_border_all = Border(top=_thin, bottom=_thin, left=_thin, right=_thin)
_border_tb = Border(top=_thin, bottom=_thin)  # top+bottom only
_border_tbl = Border(top=_thin, bottom=_thin, left=_thin)
_border_tbr = Border(top=_thin, bottom=_thin, right=_thin)

# Colors: red font (FF0000), yellow fill
_red = 'FF0000'
_yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def gen_iv_pl(
    items: List[dict],
    kb: dict,
    cno: str,
    suffix: str,
    tq: Dict[str, int],
    ta: Dict[str, float],
    ship_alloc: Dict[str, float],
    rate: float,
    out_dir: str,
) -> str:
    """Generate the IV&PL Excel file. Returns the output file path."""

    def _info(sku: str, item: dict) -> dict:
        if sku in kb:
            return kb[sku]
        return {
            'tariff_code': '',
            'english_name': item.get('name_en', ''),
            'declaration_elements': '0|0|塑料|塑料草坪|无品牌|无型号',
            'material': 'plastic',
        }

    wb = Workbook()
    wb.remove(wb.active)

    iv = wb.create_sheet('IV')
    _fill_iv(iv, items, kb, cno, tq, ta, ship_alloc, rate, _info)

    pl = wb.create_sheet('PL')
    _fill_pl(pl, items, kb, cno, tq, _info)

    fn = f'【{cno}】{suffix}IV&PL.xlsx'
    fp = os.path.join(out_dir, fn)
    wb.save(fp)
    return fp


def _fill_iv(ws, items, kb, cno, tq, ta, ship_alloc, rate, info_fn):
    """Fill the Invoice worksheet."""
    # Common styles
    _al_cwc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _al_rwc = Alignment(horizontal='right', vertical='center', wrap_text=True)
    _al_lwt = Alignment(horizontal='left', vertical='top', wrap_text=True)
    _al_lwc = Alignment(horizontal='left', vertical='center', wrap_text=True)
    _al_cc = Alignment(horizontal='center', vertical='center')

    # Set column widths
    col_widths = {0: 4778/256, 1: 6400/256, 2: 6613/256, 3: 2688/256, 4: 3242/256,
                  5: 3840/256, 6: 3072/256, 7: 3754/256, 8: 5674/256, 9: 4437/256}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    # Set row heights
    row_heights = {0: 30, 1: 30, 2: 70, 3: 30, 4: 30, 5: 67, 6: 30, 7: 30, 8: 51, 9: 56}
    for row_idx, height in row_heights.items():
        ws.row_dimensions[row_idx + 1].height = height
    for r in range(10, 13):
        ws.row_dimensions[r + 1].height = 59
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 30

    # Title — no borders on row 1
    ws['A1'] = 'INVOICE'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws.merge_cells('A1:I1')

    # --- Row 2: Shipper ---
    ws['A2'] = 'Shipper:公司名称'
    ws['A2'].font = Font(name='Arial', size=12)
    ws['A2'].alignment = _al_lwt
    ws['A2'].border = _border_all
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Shenzhen Adhoc Trading Co., Ltd.'
    ws['B2'].font = Font(name='Arial', size=10)
    ws['B2'].alignment = _al_cwc
    ws['B2'].border = _border_all
    ws['C2'].border = _border_all
    ws['C2'].font = Font(name='Arial', size=10)
    ws['C2'].alignment = _al_cwc
    ws.merge_cells('D2:E2')
    ws['D2'] = 'INVOICE NO:'
    ws['D2'].font = Font(name='Arial', size=12)
    ws['D2'].alignment = _al_rwc
    ws['D2'].border = _border_all
    ws['E2'].border = _border_all
    ws['E2'].font = Font(name='Arial', size=12)
    ws['E2'].alignment = _al_rwc
    ws.merge_cells('F2:G2')
    ws['F2'] = cno
    ws['F2'].font = Font(name='Arial', size=12)
    ws['F2'].alignment = _al_cwc
    ws['F2'].border = _border_all
    ws['G2'].border = _border_all
    ws['G2'].font = Font(name='Arial', size=12)
    ws['G2'].alignment = _al_cwc
    ws['H2'].font = Font(name='Arial', size=10)
    ws['H2'].alignment = Alignment(vertical='center')
    ws['H2'].border = _border_all
    ws['I2'].font = Font(name='Arial', size=10)
    ws['I2'].alignment = Alignment(vertical='bottom')
    ws['I2'].border = _border_all

    # --- Row 3: ADD ---
    ws['A3'] = 'ADD:地址'
    ws['A3'].font = Font(name='Arial', size=12)
    ws['A3'].alignment = _al_lwt
    ws['A3'].border = _border_all
    ws.merge_cells('B3:C3')
    ws['B3'] = 'Flat 1006, Zhenye International Business Centre,No.3101-90,Qianhai Road, Nanshan District, Shenzhen,China.'
    ws['B3'].font = Font(name='Arial', size=10)
    ws['B3'].alignment = _al_cwc
    ws['B3'].border = _border_all
    ws['C3'].border = _border_all
    ws['C3'].font = Font(name='Arial', size=10)
    ws['C3'].alignment = _al_cwc
    ws.merge_cells('D3:E3')
    ws['D3'] = 'DATE:'
    ws['D3'].font = Font(name='Arial', size=12)
    ws['D3'].alignment = _al_rwc
    ws['D3'].border = _border_all
    ws['E3'].border = _border_all
    ws['E3'].font = Font(name='Arial', size=12)
    ws['E3'].alignment = _al_rwc
    ws.merge_cells('F3:G3')
    ws['F3'] = datetime.now().strftime('%Y/%m/%d')
    ws['F3'].font = Font(name='Arial', size=12)
    ws['F3'].alignment = _al_cwc
    ws['F3'].border = _border_all
    ws['G3'].border = _border_all
    ws['G3'].font = Font(name='Arial', size=12)
    ws['G3'].alignment = _al_cwc
    ws['H3'].font = Font(name='Arial', size=10)
    ws['H3'].alignment = Alignment(vertical='center')
    ws['H3'].border = _border_all
    ws['I3'].font = Font(name='Arial', size=10)
    ws['I3'].alignment = Alignment(vertical='bottom')
    ws['I3'].border = _border_all

    # --- Row 4: ORIGIN COUNTRY (cols D-G only have content, but all cols get borders) ---
    ws.merge_cells('D4:E4')
    ws['D4'] = 'ORIGIN COUNTRY :'
    ws['D4'].font = Font(name='Arial', size=12)
    ws['D4'].alignment = _al_rwc
    ws['D4'].border = Border(top=_thin, bottom=_thin, left=_no_side, right=_thin)
    ws['E4'].border = _border_all
    ws['E4'].font = Font(name='Arial', size=12)
    ws['E4'].alignment = _al_rwc
    ws.merge_cells('F4:G4')
    ws['F4'] = 'CHINA'
    ws['F4'].font = Font(name='Arial', size=12)
    ws['F4'].alignment = _al_cwc
    ws['F4'].border = _border_all
    ws['G4'].border = _border_all
    ws['G4'].font = Font(name='Arial', size=12)
    ws['G4'].alignment = _al_cwc
    ws['H4'].font = Font(name='Arial', size=10)
    ws['H4'].alignment = Alignment(vertical='center')
    ws['H4'].border = _border_all
    ws['I4'].font = Font(name='Arial', size=10)
    ws['I4'].alignment = Alignment(vertical='bottom')
    ws['I4'].border = _border_all

    # --- Row 5: CNEE ---
    ws['A5'] = 'CNEE:公司名称'
    ws['A5'].font = Font(name='Arial', size=12)
    ws['A5'].alignment = _al_lwc
    ws['A5'].border = _border_all
    ws.merge_cells('B5:C5')
    ws['B5'] = 'ZEATALINE INTERNATIONAL TRADING'
    ws['B5'].font = Font(name='Arial', size=10)
    ws['B5'].alignment = _al_cwc
    ws['B5'].border = _border_all
    ws['C5'].border = Border(top=_thin, bottom=_thin, left=_thin, right=_no_side)
    ws['C5'].font = Font(name='Arial', size=10)
    ws['C5'].alignment = _al_cwc
    ws.merge_cells('D5:E5')
    ws['D5'] = 'PRICE TERMS:'
    ws['D5'].font = Font(name='Arial', size=12)
    ws['D5'].alignment = _al_rwc
    ws['D5'].border = _border_all
    ws['E5'].border = _border_all
    ws['E5'].font = Font(name='Arial', size=12)
    ws['E5'].alignment = _al_rwc
    ws['F5'] = 'C&F'
    ws['F5'].font = Font(name='Arial', size=11)
    ws['F5'].alignment = _al_rwc
    ws['F5'].border = _border_all
    ws['G5'] = 'USA'
    ws['G5'].font = Font(name='宋体', size=10, color=_red)
    ws['G5'].alignment = _al_cwc
    ws['G5'].border = _border_all
    ws['H5'].font = Font(name='宋体', size=10, color=_red)
    ws['H5'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['H5'].border = _border_all
    ws['I5'].font = Font(name='宋体', size=10, color=_red)
    ws['I5'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['I5'].border = _border_all

    # --- Row 6: Consignee ADD ---
    ws['A6'] = 'ADD:地址'
    ws['A6'].font = Font(name='Arial', size=12)
    ws['A6'].alignment = _al_lwt
    ws['A6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws.merge_cells('B6:C6')
    ws['B6'] = 'zouzhichang\n890 S Azusa Ave\nCity of Industry, CA 91748\nUS'
    ws['B6'].font = Font(name='Arial', size=10)
    ws['B6'].alignment = _al_cwc
    ws['B6'].border = _border_all
    ws['C6'].border = Border(top=_thin, bottom=_thin, left=_thin, right=_no_side)
    ws['C6'].font = Font(name='Arial', size=10)
    ws['C6'].alignment = _al_cwc
    ws.merge_cells('D6:E6')
    ws['D6'] = 'Reference No.'
    ws['D6'].font = Font(name='Arial', size=12)
    ws['D6'].alignment = _al_rwc
    ws['D6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws['E6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws['E6'].font = Font(name='Arial', size=12)
    ws['E6'].alignment = _al_rwc
    ws['F6'].font = Font(name='宋体', size=12)
    ws['F6'].alignment = _al_cwc
    ws['F6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws['G6'].font = Font(name='Arial', size=12)
    ws['G6'].alignment = _al_cwc
    ws['G6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws['H6'].font = Font(name='宋体', size=10, color=_red)
    ws['H6'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['H6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)
    ws['I6'].font = Font(name='宋体', size=10, color=_red)
    ws['I6'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['I6'].border = Border(top=_thin, bottom=_no_side, left=_thin, right=_thin)

    # --- Row 7: English headers (cols A-C red font) ---
    hdrs = ['No.', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
            'Unit Price', 'USD', 'Total Amount', 'material quality', 'picture']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=7, column=c, value=h)
        if c <= 3:
            cell.font = Font(name='Arial', size=12, bold=False, color=_red)
        else:
            cell.font = Font(name='Arial', size=12, bold=False)
        cell.alignment = _al_cwc
        cell.border = _border_all

    # --- Row 8: Chinese headers (ALL red font + yellow bg) ---
    cn_hdrs = [None, '海关编码（清关用的）', '英文品名', '数量', 'PC(S)',
               '单价', 'USD', '总价', '材质', '产品图片']
    for c, h in enumerate(cn_hdrs, 1):
        cell = ws.cell(row=8, column=c, value=h)
        if c == 1 or c in (5, 7):
            cell.font = Font(name='Arial', size=12, bold=False, color=_red)
        else:
            cell.font = Font(name='宋体', size=12, bold=False, color=_red)
        cell.alignment = _al_cwc
        cell.border = _border_all
        cell.fill = _yellow_fill

    row = 9
    total_usd = 0.0
    total_qty = 0
    idx = 1

    for item in items:
        sku = sku_key(item)
        qty = tq.get(sku, 0)
        if qty == 0:
            continue

        amt = ta.get(sku, 0)
        shipping = ship_alloc.get(sku, 0)
        tax_excl = amt / 1.13
        cnf_total_rmb = tax_excl + shipping
        unit_usd = cnf_total_rmb / qty / rate if qty > 0 else 0
        total_item_usd = cnf_total_rmb / rate

        info = info_fn(sku, item)

        c1 = ws.cell(row=row, column=1, value=idx)
        c1.font = Font(name='Arial', size=10)
        c1.alignment = _al_cwc
        c1.border = _border_all

        c2 = ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000)
        c2.font = Font(name='Arial', size=12)
        c2.alignment = _al_cwc
        c2.border = _border_all

        c3 = ws.cell(row=row, column=3, value=info['english_name'] or item.get('name_en', ''))
        c3.font = Font(name='Arial', size=12)
        c3.alignment = _al_cwc
        c3.border = _border_all

        c4 = ws.cell(row=row, column=4, value=qty)
        c4.font = Font(name='Arial', size=12)
        c4.number_format = '0_);[Red]\\(0\\)'
        c4.alignment = _al_cwc
        c4.border = _border_all

        c5 = ws.cell(row=row, column=5, value='PC(S)')
        c5.font = Font(name='Arial', size=12, color=_red)
        c5.alignment = _al_cwc
        c5.border = _border_all
        c5.fill = _yellow_fill

        c6 = ws.cell(row=row, column=6, value=unit_usd)
        c6.font = Font(name='Arial', size=12)
        c6.number_format = '0.00_);[Red]\\(0.00\\)'
        c6.alignment = _al_cwc
        c6.border = _border_all

        c7 = ws.cell(row=row, column=7, value='USD')
        c7.font = Font(name='Arial', size=12, color=_red)
        c7.alignment = _al_cwc
        c7.border = _border_all
        c7.fill = _yellow_fill

        c8 = ws.cell(row=row, column=8, value=total_item_usd)
        c8.font = Font(name='Arial', size=12)
        c8.number_format = '0.00_);[Red]\\(0.00\\)'
        c8.alignment = _al_cwc
        c8.border = _border_all

        c9 = ws.cell(row=row, column=9, value=info['material'] or 'plastic')
        c9.font = Font(name='Arial', size=14)
        c9.alignment = _al_cwc
        c9.border = _border_all

        # Picture column (col J) — empty with 宋体 14 bold
        c10 = ws.cell(row=row, column=10)
        c10.font = Font(name='宋体', size=14, bold=True)
        c10.alignment = _al_cc
        c10.border = _border_all

        total_usd += total_item_usd
        total_qty += qty
        idx += 1
        row += 1

    # Empty placeholder row
    placeholder_row = row
    ws.cell(row=placeholder_row, column=1, value=idx).font = Font(name='Arial', size=10)
    ws.cell(row=placeholder_row, column=1).alignment = _al_cwc
    ws.cell(row=placeholder_row, column=1).border = _border_all
    # cols B-C: top+bottom only, no left/right between them
    ws.cell(row=placeholder_row, column=2).font = Font(name='Arial', size=12)
    ws.cell(row=placeholder_row, column=2).alignment = _al_cwc
    ws.cell(row=placeholder_row, column=2).border = _border_tb
    ws.cell(row=placeholder_row, column=3).font = Font(name='Arial', size=12)
    ws.cell(row=placeholder_row, column=3).alignment = _al_cwc
    ws.cell(row=placeholder_row, column=3).border = _border_tbr
    for c in range(4, 9):
        cell = ws.cell(row=placeholder_row, column=c)
        cell.font = Font(name='Arial', size=12)
        cell.alignment = _al_cwc
        cell.border = _border_all
    ws.cell(row=placeholder_row, column=4).number_format = '#0'
    ws.cell(row=placeholder_row, column=8).number_format = '0.00_);[Red]\\(0.00\\)'
    # cols I-J
    ws.cell(row=placeholder_row, column=9).font = Font(name='Arial', size=10)
    ws.cell(row=placeholder_row, column=9).alignment = _al_cc
    ws.cell(row=placeholder_row, column=9).border = _border_all
    ws.cell(row=placeholder_row, column=10).font = Font(name='Arial', size=10)
    ws.cell(row=placeholder_row, column=10).alignment = _al_cc
    ws.cell(row=placeholder_row, column=10).border = _border_all
    row += 1

    # Totals row
    totals_row = row
    ws.cell(row=totals_row, column=1).font = Font(name='Arial', size=10)
    ws.cell(row=totals_row, column=1).alignment = _al_cwc
    ws.cell(row=totals_row, column=1).border = _border_all
    ws.cell(row=totals_row, column=2).font = Font(name='Arial', size=12)
    ws.cell(row=totals_row, column=2).alignment = _al_cwc
    ws.cell(row=totals_row, column=2).border = _border_tb
    ws.cell(row=totals_row, column=3).font = Font(name='Arial', size=12)
    ws.cell(row=totals_row, column=3).alignment = _al_cwc
    ws.cell(row=totals_row, column=3).border = _border_tbr

    c4_tot = ws.cell(row=totals_row, column=4, value=total_qty)
    c4_tot.font = Font(name='Arial', size=12)
    c4_tot.number_format = '#0'
    c4_tot.alignment = _al_cwc
    c4_tot.border = _border_all

    c5_tot = ws.cell(row=totals_row, column=5, value='/')
    c5_tot.font = Font(name='Arial', size=12)
    c5_tot.alignment = _al_cwc
    c5_tot.border = _border_all

    c6_tot = ws.cell(row=totals_row, column=6, value='/')
    c6_tot.font = Font(name='Arial', size=12)
    c6_tot.alignment = _al_cwc
    c6_tot.border = _border_all

    c7_tot = ws.cell(row=totals_row, column=7, value='/')
    c7_tot.font = Font(name='Arial', size=12)
    c7_tot.alignment = _al_cwc
    c7_tot.border = _border_all

    c8_tot = ws.cell(row=totals_row, column=8, value=total_usd)
    c8_tot.font = Font(name='Arial', size=12)
    c8_tot.number_format = '0.00_);[Red]\\(0.00\\)'
    c8_tot.alignment = _al_cwc
    c8_tot.border = _border_all

    c9_tot = ws.cell(row=totals_row, column=9, value='/')
    c9_tot.font = Font(name='Arial', size=10)
    c9_tot.alignment = _al_cc
    c9_tot.border = _border_all

    c10_tot = ws.cell(row=totals_row, column=10, value='/')
    c10_tot.font = Font(name='Arial', size=10)
    c10_tot.alignment = _al_cc
    c10_tot.border = _border_all


def _fill_pl(ws, items, kb, cno, tq, info_fn):
    """Fill the Packing List worksheet."""
    # Common styles
    _al_cwc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _al_rwc = Alignment(horizontal='right', vertical='center', wrap_text=True)
    _al_lwt = Alignment(horizontal='left', vertical='top', wrap_text=True)
    _al_lwc = Alignment(horizontal='left', vertical='center', wrap_text=True)
    _al_cc = Alignment(horizontal='center', vertical='center')

    # Set column widths
    col_widths = {0: 4010/256, 1: 7082/256, 2: 6613/256, 3: 2346/256, 4: 3242/256,
                  5: 3840/256, 6: 3072/256, 7: 3114/256, 8: 3157/256, 9: 5034/256}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    # Set row heights
    row_heights = {0: 30, 1: 30, 2: 61, 3: 30, 4: 30, 5: 67, 6: 30, 7: 30, 8: 77}
    for row_idx, height in row_heights.items():
        ws.row_dimensions[row_idx + 1].height = height
    # Data rows: 68, 78, 78, 78 for rows 10-13
    ws.row_dimensions[10].height = 68
    for r in range(11, 14):
        ws.row_dimensions[r].height = 78
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 30
    ws.row_dimensions[16].height = 30

    # --- Title row: PL has borders on merged title cells ---
    ws['A1'] = 'PACKING LIST'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws['A1'].border = _border_all
    ws.merge_cells('A1:I1')
    # Set border on all cells in the merged range
    for c in range(2, 10):  # cols B-I
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        cell.border = _border_all
    # col J row 1 — border
    ws['J1'].font = Font(name='Arial', size=10)
    ws['J1'].alignment = Alignment(vertical='bottom')
    ws['J1'].border = _border_all

    # --- Row 2: Shipper ---
    ws['A2'] = 'Shipper:'
    ws['A2'].font = Font(name='Arial', size=12)
    ws['A2'].alignment = _al_lwt
    ws['A2'].border = _border_all
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Shenzhen Adhoc Trading Co., Ltd.'
    ws['B2'].font = Font(name='Arial', size=10)
    ws['B2'].alignment = _al_cwc
    ws['B2'].border = _border_all
    ws['C2'].border = _border_all
    ws['C2'].font = Font(name='Arial', size=10)
    ws['C2'].alignment = _al_cwc
    ws.merge_cells('D2:E2')
    ws['D2'] = 'INVOICE NO:'
    ws['D2'].font = Font(name='Arial', size=12)
    ws['D2'].alignment = _al_rwc
    ws['D2'].border = _border_all
    ws['E2'].border = _border_all
    ws['E2'].font = Font(name='Arial', size=12)
    ws['E2'].alignment = _al_rwc
    ws.merge_cells('F2:G2')
    ws['F2'] = cno
    ws['F2'].font = Font(name='Arial', size=12)
    ws['F2'].alignment = _al_cwc
    ws['F2'].border = _border_all
    ws['G2'].border = _border_all
    ws['G2'].font = Font(name='Arial', size=12)
    ws['G2'].alignment = _al_cwc
    ws['H2'].font = Font(name='Arial', size=10)
    ws['H2'].alignment = Alignment(vertical='center')
    ws['H2'].border = _border_all
    ws['I2'].font = Font(name='Arial', size=10)
    ws['I2'].alignment = Alignment(vertical='bottom')
    ws['I2'].border = _border_all

    # --- Row 3: ADD ---
    ws['A3'] = 'ADD:'
    ws['A3'].font = Font(name='Arial', size=12)
    ws['A3'].alignment = _al_lwt
    ws['A3'].border = _border_all
    ws.merge_cells('B3:C3')
    ws['B3'] = 'Flat 1006, Zhenye International Business Centre,No.3101-90,Qianhai Road, Nanshan District, Shenzhen,China.'
    ws['B3'].font = Font(name='Arial', size=10)
    ws['B3'].alignment = _al_cwc
    ws['B3'].border = _border_all
    ws['C3'].border = _border_all
    ws['C3'].font = Font(name='Arial', size=10)
    ws['C3'].alignment = _al_cwc
    ws.merge_cells('D3:E3')
    ws['D3'] = 'DATE:'
    ws['D3'].font = Font(name='Arial', size=12)
    ws['D3'].alignment = _al_rwc
    ws['D3'].border = _border_all
    ws['E3'].border = _border_all
    ws['E3'].font = Font(name='Arial', size=12)
    ws['E3'].alignment = _al_rwc
    ws.merge_cells('F3:G3')
    ws['F3'] = datetime.now().strftime('%Y/%m/%d')
    ws['F3'].font = Font(name='宋体', size=12)
    ws['F3'].alignment = _al_cwc
    ws['F3'].border = _border_all
    ws['G3'].border = _border_all
    ws['G3'].font = Font(name='Arial', size=12)
    ws['G3'].alignment = _al_cwc
    ws['H3'].font = Font(name='Arial', size=10)
    ws['H3'].alignment = Alignment(vertical='center')
    ws['H3'].border = _border_all
    ws['I3'].font = Font(name='Arial', size=10)
    ws['I3'].alignment = Alignment(vertical='bottom')
    ws['I3'].border = _border_all

    # --- Row 4: ORIGIN COUNTRY + empty cells with borders ---
    ws['A4'].font = Font(name='Arial', size=12)
    ws['A4'].alignment = _al_lwt
    ws['A4'].border = _border_all
    ws['B4'].font = Font(name='Arial', size=10)
    ws['B4'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['B4'].border = _border_all
    ws['C4'].font = Font(name='Arial', size=10)
    ws['C4'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['C4'].border = _border_all
    ws.merge_cells('D4:E4')
    ws['D4'] = 'ORIGIN COUNTRY :'
    ws['D4'].font = Font(name='Arial', size=12)
    ws['D4'].alignment = _al_rwc
    ws['D4'].border = _border_all
    ws['E4'].border = _border_all
    ws['E4'].font = Font(name='Arial', size=12)
    ws['E4'].alignment = _al_rwc
    ws.merge_cells('F4:G4')
    ws['F4'] = 'CHINA'
    ws['F4'].font = Font(name='Arial', size=12)
    ws['F4'].alignment = _al_cwc
    ws['F4'].border = _border_all
    ws['G4'].border = _border_all
    ws['G4'].font = Font(name='Arial', size=12)
    ws['G4'].alignment = _al_cwc
    ws['H4'].font = Font(name='Arial', size=10)
    ws['H4'].alignment = Alignment(vertical='center')
    ws['H4'].border = _border_all
    ws['I4'].font = Font(name='Arial', size=10)
    ws['I4'].alignment = Alignment(vertical='bottom')
    ws['I4'].border = _border_all

    # --- Row 5: TO ---
    ws['A5'] = 'TO:'
    ws['A5'].font = Font(name='Arial', size=12)
    ws['A5'].alignment = _al_lwc
    ws['A5'].border = _border_all
    ws.merge_cells('B5:C5')
    ws['B5'] = 'ZEATALINE INTERNATIONAL TRADING'
    ws['B5'].font = Font(name='Arial', size=10)
    ws['B5'].alignment = _al_cwc
    ws['B5'].border = _border_all
    ws['C5'].border = _border_all
    ws['C5'].font = Font(name='Arial', size=10)
    ws['C5'].alignment = _al_cwc
    ws.merge_cells('D5:E5')
    ws['D5'] = 'PRICE TERMS:'
    ws['D5'].font = Font(name='Arial', size=12)
    ws['D5'].alignment = _al_rwc
    ws['D5'].border = _border_all
    ws['E5'].border = _border_all
    ws['E5'].font = Font(name='Arial', size=12)
    ws['E5'].alignment = _al_rwc
    ws['F5'] = 'C&F'
    ws['F5'].font = Font(name='Arial', size=11)
    ws['F5'].alignment = _al_rwc
    ws['F5'].border = _border_all
    ws['G5'] = 'USA'
    ws['G5'].font = Font(name='宋体', size=10, color=_red)
    ws['G5'].alignment = _al_cwc
    ws['G5'].border = _border_all
    ws['H5'].font = Font(name='宋体', size=10, color=_red)
    ws['H5'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['H5'].border = _border_all
    ws['I5'].font = Font(name='宋体', size=10, color=_red)
    ws['I5'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['I5'].border = _border_all

    # --- Row 6: Consignee ADD ---
    ws['A6'] = 'ADD:'
    ws['A6'].font = Font(name='Arial', size=12)
    ws['A6'].alignment = _al_lwt
    ws['A6'].border = _border_all
    ws.merge_cells('B6:C6')
    ws['B6'] = 'zouzhichang\n890 S Azusa Ave\nCity of Industry, CA 91748\nUS'
    ws['B6'].font = Font(name='Arial', size=10)
    ws['B6'].alignment = _al_cwc
    ws['B6'].border = _border_all
    ws['C6'].border = _border_all
    ws['C6'].font = Font(name='Arial', size=10)
    ws['C6'].alignment = _al_cwc
    ws.merge_cells('D6:E6')
    ws['D6'] = 'Reference No.'
    ws['D6'].font = Font(name='Arial', size=12)
    ws['D6'].alignment = _al_rwc
    ws['D6'].border = _border_all
    ws['E6'].border = _border_all
    ws['E6'].font = Font(name='Arial', size=12)
    ws['E6'].alignment = _al_rwc
    ws['F6'].font = Font(name='宋体', size=12)
    ws['F6'].alignment = _al_cwc
    ws['F6'].border = _border_all
    ws['G6'].font = Font(name='Arial', size=12)
    ws['G6'].alignment = _al_cwc
    ws['G6'].border = _border_all
    ws['H6'].font = Font(name='宋体', size=10, color=_red)
    ws['H6'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['H6'].border = _border_all
    ws['I6'].font = Font(name='宋体', size=10, color=_red)
    ws['I6'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['I6'].border = _border_all

    # --- Row 7: English headers (cols A-C red font) ---
    hdrs = ['NO', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
            'Box Qty\n(CTNS)', 'N.W.\n(KG)', 'G.W.\n(KG)', 'VOLUME (CBM)', None]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=7, column=c, value=h)
        if c <= 3:
            cell.font = Font(name='Arial', size=12, bold=False, color=_red)
        else:
            cell.font = Font(name='Arial', size=12, bold=False)
        cell.alignment = _al_cwc
        cell.border = _border_all

    # --- Row 8: Chinese headers (cols A-I red font + yellow bg) ---
    cn_hdrs = [None, '海关编码（清关用的）', '英文品名', '数量', 'PC(S)',
               '箱数', '净重', '毛重', '方数', '产品图片']
    for c, h in enumerate(cn_hdrs, 1):
        cell = ws.cell(row=8, column=c, value=h)
        if c == 1 or c == 5:
            cell.font = Font(name='Arial', size=12, bold=False, color=_red)
        else:
            cell.font = Font(name='宋体', size=12, bold=False, color=_red)
        cell.alignment = _al_cwc
        cell.border = _border_all
        if c <= 9:
            cell.fill = _yellow_fill
    # col J row 8: no borders per reference
    ws.cell(row=8, column=10).border = Border()

    row = 9
    tot_qty = 0
    tot_boxes = 0
    tot_nw = 0.0
    tot_gw = 0.0
    tot_vol = 0.0
    idx = 1

    for item in items:
        sku = sku_key(item)
        qty = tq.get(sku, 0)
        if qty == 0:
            continue

        pr = item.get('packing_rate', 1) or 1
        boxes = qty / pr
        l, w, h_ = get_lwh(item)
        nw = item.get('net_weight_kg', 0) * boxes
        gw = item.get('gross_weight_kg', 0) * boxes
        volume = (l * w * h_ / 1_000_000) * boxes

        info = info_fn(sku, item)

        c1 = ws.cell(row=row, column=1, value=idx)
        c1.font = Font(name='Arial', size=10)
        c1.alignment = _al_cwc
        c1.border = _border_all

        c2 = ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000)
        c2.font = Font(name='Arial', size=12)
        c2.alignment = _al_cwc
        c2.border = _border_all

        c3 = ws.cell(row=row, column=3, value=info['english_name'] or item.get('name_en', ''))
        c3.font = Font(name='Arial', size=12)
        c3.alignment = _al_cwc
        c3.border = _border_all

        c4 = ws.cell(row=row, column=4, value=qty)
        c4.font = Font(name='宋体', size=14)
        c4.alignment = _al_cc
        c4.border = _border_all

        c5 = ws.cell(row=row, column=5, value='PC(S)')
        c5.font = Font(name='Arial', size=12, color=_red)
        c5.alignment = _al_cwc
        c5.border = _border_all
        c5.fill = _yellow_fill

        c6 = ws.cell(row=row, column=6, value=boxes)
        c6.font = Font(name='宋体', size=14)
        c6.alignment = _al_cc
        c6.border = _border_all

        c7 = ws.cell(row=row, column=7, value=round(nw, 1))
        c7.font = Font(name='Arial', size=12)
        c7.number_format = '0.00'
        c7.alignment = _al_cwc
        c7.border = _border_all

        c8 = ws.cell(row=row, column=8, value=round(gw, 1))
        c8.font = Font(name='Arial', size=12)
        c8.number_format = '0.00'
        c8.alignment = _al_cwc
        c8.border = _border_all

        c9 = ws.cell(row=row, column=9, value=round(volume, 5))
        c9.font = Font(name='Arial', size=12)
        c9.number_format = '0.00'
        c9.alignment = _al_cwc
        c9.border = _border_all

        # Picture column (col J) — empty with 宋体 14 bold
        c10 = ws.cell(row=row, column=10)
        c10.font = Font(name='宋体', size=14, bold=True)
        c10.alignment = _al_cc
        c10.border = _border_all

        tot_qty += qty
        tot_boxes += boxes
        tot_nw += nw
        tot_gw += gw
        tot_vol += volume
        idx += 1
        row += 1

    # Empty placeholder row
    placeholder_row = row
    ws.cell(row=placeholder_row, column=1, value=idx).font = Font(name='Arial', size=10)
    ws.cell(row=placeholder_row, column=1).alignment = _al_cwc
    ws.cell(row=placeholder_row, column=1).border = _border_all
    for c in range(2, 10):
        cell = ws.cell(row=placeholder_row, column=c)
        if c == 4:
            cell.font = Font(name='Arial', size=10)
        else:
            cell.font = Font(name='Arial', size=12)
        cell.alignment = _al_cwc
        cell.border = _border_all
    c5_ph = ws.cell(row=placeholder_row, column=5, value='PC(S)')
    c5_ph.font = Font(name='Arial', size=12, color=_red)
    c5_ph.fill = _yellow_fill
    row += 1

    # Totals row
    totals_row = row
    ws.cell(row=totals_row, column=1).font = Font(name='Arial', size=10)
    ws.cell(row=totals_row, column=1).alignment = _al_cwc
    ws.cell(row=totals_row, column=1).border = _border_all
    ws.cell(row=totals_row, column=2).font = Font(name='Arial', size=12, color=_red)
    ws.cell(row=totals_row, column=2).alignment = _al_cwc
    ws.cell(row=totals_row, column=2).border = _border_all
    ws.cell(row=totals_row, column=3).font = Font(name='Arial', size=12, color=_red)
    ws.cell(row=totals_row, column=3).alignment = _al_cwc
    ws.cell(row=totals_row, column=3).border = _border_all

    c4_tot = ws.cell(row=totals_row, column=4, value=tot_qty)
    c4_tot.font = Font(name='Arial', size=12)
    c4_tot.number_format = '#0'
    c4_tot.alignment = _al_cwc
    c4_tot.border = _border_all

    c5_tot = ws.cell(row=totals_row, column=5, value='/')
    c5_tot.font = Font(name='Arial', size=12)
    c5_tot.alignment = _al_cwc
    c5_tot.border = _border_all

    c6_tot = ws.cell(row=totals_row, column=6, value=round(tot_boxes))
    c6_tot.font = Font(name='Arial', size=12)
    c6_tot.number_format = '#0'
    c6_tot.alignment = _al_cwc
    c6_tot.border = _border_all

    c7_tot = ws.cell(row=totals_row, column=7, value=round(tot_nw, 1))
    c7_tot.font = Font(name='Arial', size=12)
    c7_tot.number_format = '#0.00'
    c7_tot.alignment = _al_cwc
    c7_tot.border = _border_all

    c8_tot = ws.cell(row=totals_row, column=8, value=round(tot_gw, 1))
    c8_tot.font = Font(name='Arial', size=12)
    c8_tot.number_format = '#0.00'
    c8_tot.alignment = _al_cwc
    c8_tot.border = _border_all

    c9_tot = ws.cell(row=totals_row, column=9, value=round(tot_vol, 5))
    c9_tot.font = Font(name='Arial', size=12)
    c9_tot.number_format = '#0.00'
    c9_tot.alignment = _al_cwc
    c9_tot.border = _border_all
