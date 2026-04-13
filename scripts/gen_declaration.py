#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Declaration Draft (报关单草稿) Excel document.
"""

import os
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

from helpers import sku_key


def gen_declaration(
    items: List[dict],
    contract: dict,
    kb: dict,
    cno: str,
    suffix: str,
    tq: Dict[str, int],
    ta: Dict[str, float],
    ship_alloc: Dict[str, float],
    total_ship: float,
    rate: float,
    price_term: str,
    out_dir: str,
) -> str:
    """Generate the declaration draft Excel file. Returns the output file path."""

    def _info(sku: str, item: dict) -> dict:
        if sku in kb:
            return kb[sku]
        return {
            'tariff_code': '',
            'english_name': item.get('name_en', ''),
            'declaration_elements': '0|0|塑料|塑料草坪|无品牌|无型号',
            'material': 'plastic',
        }

    wb = Workbook()
    ws = wb.active
    ws.title = '报关单草稿'

    # Set column widths (A-S: openpyxl character widths)
    col_widths = {
        'A': 6.27, 'B': 7.37, 'C': 6.0, 'D': 25.45, 'E': 10.91, 'F': 12.37, 'G': 11.0,
        'H': 6.73, 'I': 7.91, 'J': 7.09, 'K': 7.63, 'L': 5.37, 'M': 5.27, 'N': 5.09,
        'O': 7.27, 'P': 4.09, 'Q': 6.27, 'R': 7.63, 'S': 8.45
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set row heights (matching example template)
    ws.row_dimensions[1].height = 25.5
    ws.row_dimensions[2].height = 12.75
    ws.row_dimensions[3].height = 12.0
    ws.row_dimensions[4].height = 14.25
    ws.row_dimensions[5].height = 12.0
    ws.row_dimensions[7].height = 12.0
    ws.row_dimensions[8].height = 14.25
    ws.row_dimensions[9].height = 12.0
    ws.row_dimensions[10].height = 14.25
    ws.row_dimensions[11].height = 12.0
    ws.row_dimensions[12].height = 14.25
    ws.row_dimensions[13].height = 12.0
    ws.row_dimensions[14].height = 12.75
    ws.row_dimensions[15].height = 12.0
    ws.row_dimensions[17].height = 14.25
    ws.row_dimensions[18].height = 14.25

    # Title
    ws.merge_cells('A1:S1')
    ws['A1'] = '中华人民共和国海关出口货物报关单'
    ws['A1'].font = Font(name='宋体', size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Helper: border styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 2 (no borders, right-aligned, with merge cells)
    ws.merge_cells('A2:B2')
    c_a2 = ws['A2']
    c_a2.value = '预录入编号：'
    c_a2.font = Font(name='宋体', size=10)
    c_a2.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('C2:E2')

    ws['F2'] = '申报口岸:'
    ws['F2'].font = Font(name='宋体', size=10)
    ws['F2'].alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('G2:H2')

    ws['I2'] = '海关编号:'
    ws['I2'].font = Font(name='宋体', size=10)
    ws['I2'].alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('J2:N2')

    # Define border styles
    label_font = Font(name='宋体', size=10)
    label_font_red = Font(name='宋体', size=10, color='FF0000')
    value_font = Font(name='宋体', size=12, bold=True)
    value_font_red = Font(name='宋体', size=12, bold=True, color='FF0000')
    value_font_11 = Font(name='宋体', size=11, bold=True)
    label_align = Alignment(vertical='center')
    label_align_left = Alignment(horizontal='left', vertical='center')
    value_align_left = Alignment(horizontal='left', vertical='center')
    med_left = Side(style='medium')
    med_right = Side(style='medium')
    thin_side = Side(style='thin')

    # Row 3-4: 境内发货人
    ws['A3'] = '境内发货人'
    ws['A3'].font = label_font
    ws['A3'].alignment = label_align_left
    ws['A3'].border = Border(top=Side(style='medium'), left=Side(style='medium'))
    ws['B3'].border = Border(top=Side(style='medium'), left=Side(style='medium'))
    ws.merge_cells('A3:B3')
    ws['C3'] = '（91440300687577411Y）'
    ws['C3'].font = label_font
    ws['C3'].alignment = label_align_left
    ws['C3'].border = Border(top=Side(style='medium'), right=thin_side)
    ws.merge_cells('C3:D3')
    ws['E3'] = '出境关别'
    ws['E3'].font = label_font_red
    ws['E3'].alignment = label_align
    ws['E3'].border = Border(top=Side(style='medium'), left=thin_side)
    ws['F3'] = '(    )'
    ws['F3'].font = label_font
    ws['F3'].alignment = label_align
    ws['F3'].border = Border(top=Side(style='medium'), right=thin_side)
    ws['G3'] = '出口日期'
    ws['G3'].font = label_font
    ws['G3'].alignment = label_align
    ws['G3'].border = Border(top=Side(style='medium'), left=thin_side)
    ws.merge_cells('H3:J3')
    ws['K3'] = '申报日期'
    ws['K3'].font = label_font
    ws['K3'].alignment = label_align
    ws['K3'].border = Border(top=Side(style='medium'), left=thin_side)
    ws.merge_cells('L3:N3')
    ws['O3'] = '备案号'
    ws['O3'].font = label_font
    ws['O3'].alignment = label_align
    ws['O3'].border = Border(top=Side(style='medium'), left=thin_side)
    ws.merge_cells('P3:S3')

    # Ensure continuous medium top border across ALL cells in Row 3
    for col_idx in range(1, 20):
        cell = ws.cell(row=3, column=col_idx)
        existing = cell.border
        cell.border = Border(
            top=Side(style='medium'),
            left=existing.left if existing.left and existing.left.style else None,
            right=existing.right if existing.right and existing.right.style else None,
            bottom=existing.bottom if existing.bottom and existing.bottom.style else None,
        )
    s3 = ws['S3']
    s3.border = Border(top=Side(style='medium'), right=Side(style='medium'))

    # Row 4: values
    ws['A4'] = '深圳市艾进贸易有限公司'
    ws['A4'].font = value_font
    ws['A4'].alignment = value_align_left
    ws['A4'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
    ws.merge_cells('A4:D4')
    ws['E4'] = ''
    ws['E4'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('E4:F4')
    ws['G4'] = ''
    ws['G4'].border = Border(bottom=thin_side, left=thin_side)
    ws.merge_cells('G4:I4')
    ws['K4'] = ''
    ws['K4'].border = Border(bottom=thin_side)
    ws.merge_cells('K4:N4')
    ws['O4'] = ''
    ws['O4'].border = Border(bottom=thin_side)
    ws.merge_cells('O4:S4')

    # Row 5-6: 境外收货人
    ws['A5'] = '境外收货人'
    ws['A5'].font = label_font_red
    ws['A5'].alignment = label_align_left
    ws['A5'].border = Border(top=thin_side, left=Side(style='medium'))
    ws.merge_cells('A5:B5')
    ws['C5'] = ' '
    ws['C5'].font = label_font
    ws['C5'].border = Border(top=thin_side)
    ws['E5'] = '运输方式'
    ws['E5'].font = label_font
    ws['E5'].alignment = label_align
    ws['E5'].border = Border(top=thin_side, left=thin_side)
    ws['F5'] = '(    )'
    ws['F5'].font = label_font
    ws['F5'].alignment = label_align
    ws['F5'].border = Border(top=thin_side, right=thin_side)
    ws['G5'] = '运输工具名称及航次号'
    ws['G5'].font = label_font
    ws['G5'].alignment = Alignment(horizontal='left', vertical='center')
    ws['G5'].border = Border(top=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('G5:J5')
    ws['K5'] = '提运单号'
    ws['K5'].font = label_font
    ws['K5'].alignment = label_align
    ws['K5'].border = Border(top=thin_side, left=thin_side)
    ws.merge_cells('L5:N5')
    ws.merge_cells('O5:P5')
    ws.merge_cells('Q5:S5')

    # Row 6: values
    ws['A6'] = 'ZEATALINE INTERNATIONAL TRADING'
    ws['A6'].font = value_font_11
    ws['A6'].alignment = value_align_left
    ws['A6'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
    ws.merge_cells('A6:D6')
    ws['E6'] = '水路运输'
    ws['E6'].font = Font(name='宋体', size=11)
    ws['E6'].alignment = value_align_left
    ws['E6'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('E6:F6')
    ws['G6'] = ''
    ws['G6'].border = Border(bottom=thin_side, left=thin_side)
    ws.merge_cells('G6:J6')
    ws['K6'] = ''
    ws['K6'].border = Border(bottom=thin_side)
    ws.merge_cells('K6:N6')
    ws['O6'] = ''
    ws['O6'].border = Border(bottom=thin_side)
    ws.merge_cells('O6:S6')

    # Row 7-8: 生产销售单位
    ws['A7'] = '生产销售单位'
    ws['A7'].font = label_font
    ws['A7'].alignment = label_align_left
    ws['A7'].border = Border(top=thin_side, left=Side(style='medium'))
    ws.merge_cells('A7:B7')
    ws['C7'] = ' '
    ws['C7'].font = label_font
    ws['C7'].alignment = label_align_left
    ws['C7'].border = Border(top=thin_side, right=thin_side)
    ws.merge_cells('C7:D7')
    ws['E7'] = '监管方式'
    ws['E7'].font = label_font
    ws['E7'].alignment = label_align
    ws['E7'].border = Border(top=thin_side, left=thin_side)
    ws['F7'] = '(    )'
    ws['F7'].font = label_font
    ws['F7'].alignment = label_align
    ws['F7'].border = Border(top=thin_side, right=thin_side)
    ws['G7'] = '征免性质'
    ws['G7'].font = label_font
    ws['G7'].alignment = label_align
    ws['G7'].border = Border(top=thin_side, left=thin_side)
    ws['H7'] = '(    )'
    ws['H7'].font = label_font
    ws['H7'].alignment = label_align_left
    ws['H7'].border = Border(top=thin_side, right=thin_side)
    ws.merge_cells('H7:J7')
    ws['K7'] = '许可证号'
    ws['K7'].font = label_font
    ws['K7'].alignment = label_align
    ws['K7'].border = Border(top=thin_side, left=thin_side)
    ws.merge_cells('L7:N7')
    ws.merge_cells('P7:S7')

    # Row 8: values
    ws['A8'] = '深圳市艾进贸易有限公司'
    ws['A8'].font = value_font
    ws['A8'].alignment = value_align_left
    ws['A8'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
    ws.merge_cells('A8:D8')
    ws['E8'] = '一般贸易'
    ws['E8'].font = value_font
    ws['E8'].alignment = value_align_left
    ws['E8'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('E8:F8')
    ws['G8'] = '一般征税'
    ws['G8'].font = value_font
    ws['G8'].alignment = value_align_left
    ws['G8'].border = Border(bottom=thin_side, left=thin_side)
    ws.merge_cells('G8:I8')
    ws['K8'] = ''
    ws['K8'].border = Border(bottom=thin_side)
    ws.merge_cells('K8:N8')
    ws['O8'] = ''
    ws['O8'].border = Border(bottom=thin_side)
    ws.merge_cells('O8:S8')

    # Row 9-10: 合同协议号
    ws['A9'] = '合同协议号'
    ws['A9'].font = label_font
    ws['A9'].alignment = label_align_left
    ws['A9'].border = Border(top=thin_side, left=Side(style='medium'))
    ws.merge_cells('A9:B9')
    ws.merge_cells('C9:D9')
    ws['E9'] = '贸易国(地区)'
    ws['E9'].font = label_font
    ws['E9'].alignment = label_align
    ws['E9'].border = Border(top=thin_side, left=thin_side)
    ws['F9'] = '(  USA  )'
    ws['F9'].font = label_font
    ws['F9'].alignment = label_align
    ws['F9'].border = Border(top=thin_side, right=thin_side)
    ws['G9'] = '运抵国（地区）'
    ws['G9'].font = label_font
    ws['G9'].alignment = label_align
    ws['G9'].border = Border(top=thin_side, left=thin_side)
    ws['H9'] = '(  USA  )'
    ws['H9'].font = label_font
    ws['H9'].alignment = label_align_left
    ws.merge_cells('H9:J9')
    ws['K9'] = '指运港'
    ws['K9'].font = label_font
    ws['K9'].alignment = label_align
    ws['K9'].border = Border(top=thin_side, left=thin_side)
    ws['L9'] = '(    )'
    ws['L9'].font = label_font
    ws['L9'].alignment = label_align_left
    ws['L9'].border = Border(top=thin_side, right=thin_side)
    ws.merge_cells('L9:N9')
    ws['O9'] = '离境口岸'
    ws['O9'].font = label_font_red
    ws['O9'].alignment = label_align
    ws['O9'].border = Border(top=thin_side, left=thin_side)
    ws['P9'] = '(    )'
    ws['P9'].font = label_font
    ws['P9'].alignment = label_align_left
    ws['P9'].border = Border(top=thin_side, right=Side(style='medium'))
    ws.merge_cells('P9:S9')

    # Row 10: values
    ws['A10'] = cno
    ws['A10'].font = value_font
    ws['A10'].alignment = value_align_left
    ws['A10'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
    ws.merge_cells('A10:D10')
    ws['E10'] = '美国'
    ws['E10'].font = value_font_red
    ws['E10'].alignment = value_align_left
    ws['E10'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('E10:F10')
    ws['G10'] = '美国'
    ws['G10'].font = Font(name='宋体', size=11)
    ws['G10'].alignment = value_align_left
    ws['G10'].border = Border(bottom=thin_side, left=thin_side)
    ws.merge_cells('G10:J10')
    ws['K10'] = '美国'
    ws['K10'].font = Font(name='宋体', size=11)
    ws['K10'].alignment = value_align_left
    ws['K10'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('K10:N10')
    ws['O10'] = ''
    ws['O10'].border = Border(bottom=thin_side)
    ws.merge_cells('O10:S10')

    # Row 11-12: 包装/件数/毛重/运费
    ws['A11'] = '包装种类'
    ws['A11'].font = label_font
    ws['A11'].alignment = label_align_left
    ws['A11'].border = Border(top=thin_side, left=Side(style='medium'))
    ws.merge_cells('A11:B11')
    ws['C11'] = '( 22/06   )'
    ws['C11'].font = label_font
    ws['C11'].alignment = label_align_left
    ws['C11'].border = Border(top=thin_side, right=thin_side)
    ws.merge_cells('C11:D11')

    ws['E11'] = '件数'
    ws['E11'].font = label_font
    ws['E11'].alignment = label_align
    ws['E11'].border = Border(left=thin_side, right=thin_side)

    ws['F11'] = '毛重（千克）'
    ws['F11'].font = label_font
    ws['F11'].alignment = label_align
    ws['F11'].border = Border(top=thin_side, left=thin_side, right=thin_side)

    ws['G11'] = '净重（千克）'
    ws['G11'].font = label_font
    ws['G11'].alignment = Alignment(horizontal='left', vertical='center')
    ws['G11'].border = Border(top=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('G11:H11')

    ws['I11'] = '成交方式'
    ws['I11'].font = label_font
    ws['I11'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I11'].border = Border(top=thin_side, left=thin_side)

    ws['J11'] = '(  )'
    ws['J11'].font = label_font
    ws['J11'].alignment = label_align
    ws['J11'].border = Border(top=thin_side)

    ws['K11'] = '运费'
    ws['K11'].font = label_font
    ws['K11'].alignment = label_align
    ws['K11'].border = Border(top=thin_side, left=thin_side)
    ws.merge_cells('L11:M11')

    ws['N11'] = '保费'
    ws['N11'].font = label_font
    ws['N11'].alignment = label_align
    ws['N11'].border = Border(top=thin_side)
    ws.merge_cells('O11:P11')

    ws['Q11'] = '杂费'
    ws['Q11'].font = label_font
    ws['Q11'].alignment = label_align
    ws['Q11'].border = Border(top=thin_side, left=thin_side)
    ws.merge_cells('R11:S11')

    # Compute PL totals for the header
    tot_boxes = 0
    tot_nw = 0.0
    tot_gw = 0.0
    for item in items:
        sku = sku_key(item)
        qty = tq.get(sku, 0)
        if qty > 0:
            pr = item.get('packing_rate', 1) or 1
            bx = qty / pr
            tot_boxes += int(bx)
            tot_nw += item.get('net_weight_kg', 0) * bx
            tot_gw += item.get('gross_weight_kg', 0) * bx

    ws['A12'] = '纸制或纤维板制盒/箱/包/袋'
    ws['A12'].font = value_font
    ws['A12'].alignment = value_align_left
    ws['A12'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
    ws.merge_cells('A12:D12')

    ws['E12'] = tot_boxes
    ws['E12'].font = value_font
    ws['E12'].alignment = value_align_left
    ws['E12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)

    ws['F12'] = round(tot_gw, 1)
    ws['F12'].font = value_font
    ws['F12'].alignment = value_align_left
    ws['F12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)

    ws['G12'] = round(tot_nw, 1)
    ws['G12'].font = value_font
    ws['G12'].alignment = value_align_left
    ws['G12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
    ws.merge_cells('G12:H12')

    ws['I12'] = price_term
    ws['I12'].font = value_font
    ws['I12'].alignment = value_align_left
    ws['I12'].border = Border(bottom=thin_side, left=thin_side)
    ws.merge_cells('I12:J12')

    ws['K12'] = 'USD'
    ws['K12'].font = value_font
    ws['K12'].alignment = label_align
    ws['K12'].border = Border(bottom=thin_side, left=thin_side)

    ws['L12'] = round(total_ship / rate, 2)
    ws['L12'].font = value_font
    ws['L12'].alignment = Alignment(horizontal='center', vertical='center')
    ws['L12'].border = Border(bottom=thin_side, right=thin_side)
    ws.merge_cells('L12:M12')

    ws['N12'] = None
    ws['N12'].font = value_font
    ws['N12'].alignment = value_align_left
    ws['N12'].border = Border(bottom=thin_side, right=thin_side)
    ws.merge_cells('N12:P12')

    ws['Q12'] = None
    ws['Q12'].font = value_font
    ws['Q12'].alignment = value_align_left
    ws['Q12'].border = Border(bottom=thin_side, left=thin_side, right=Side(style='medium'))
    ws.merge_cells('Q12:S12')

    # Row 13: 随附单证及编号
    ws['A13'] = '随附单证及编号'
    ws['A13'].font = label_font
    ws['A13'].alignment = label_align
    ws['A13'].border = Border(top=thin_side, left=Side(style='medium'))
    ws.merge_cells('C13:S13')

    # Row 14
    ws.merge_cells('A14:C14')
    ws.merge_cells('D14:S14')

    # Row 15
    ws['A15'] = '标记唛码及备注：'
    ws['A15'].font = label_font
    ws['A15'].alignment = label_align
    ws['A15'].border = Border(left=Side(style='medium'))
    ws.merge_cells('C15:S15')

    # Row 16
    ws['A16'] = '备注：'
    ws['A16'].font = label_font
    ws['A16'].alignment = label_align
    ws['A16'].border = Border(left=Side(style='medium'))
    ws.merge_cells('B16:S16')

    # Row 17: section header
    ws.merge_cells('A17:K17')
    ws.merge_cells('L17:N17')
    ws.merge_cells('O17:S17')

    # Row 18: spacer row with specific formatting
    med_tb = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    med_ltb = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    med_lrtb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
    med_ltb_thin = Border(left=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
    med_rtb = Border(right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
    med_rtb_med = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    ws['A18'] = ' '
    ws['A18'].font = Font(name='宋体', size=11, bold=True)
    ws['A18'].alignment = label_align
    ws['A18'].border = med_ltb

    ws['B18'] = ' '
    ws['B18'].font = Font(name='宋体', size=11, bold=True)
    ws['B18'].alignment = label_align
    ws['B18'].border = med_tb

    ws['C18'] = None
    ws['C18'].font = Font(name='宋体', size=11, bold=True)
    ws['C18'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C18'].border = med_lrtb

    ws['D18'] = ' '
    ws['D18'].font = Font(name='宋体', size=11, bold=True)
    ws['D18'].alignment = Alignment(horizontal='right', vertical='center')
    ws['D18'].border = med_ltb_thin

    ws['E18'] = None
    ws['E18'].font = Font(name='宋体', size=11, bold=True)
    ws['E18'].alignment = Alignment(horizontal='left', vertical='center')
    ws['E18'].border = med_lrtb

    ws['F18'] = ' '
    ws['F18'].font = Font(name='宋体', size=11, bold=True)
    ws['F18'].alignment = Alignment(horizontal='right', vertical='center')
    ws['F18'].border = med_tb
    ws.merge_cells('F18:G18')

    ws['H18'] = None
    ws['H18'].font = Font(name='宋体', size=11, bold=True)
    ws['H18'].alignment = Alignment(horizontal='left', vertical='center')
    ws['H18'].border = med_lrtb

    ws['I18'] = ' '
    ws['I18'].font = Font(name='宋体', size=11, bold=True)
    ws['I18'].alignment = Alignment(horizontal='right', vertical='center')
    ws['I18'].border = med_ltb_thin
    ws.merge_cells('I18:J18')

    ws['K18'] = ' '
    ws['K18'].font = Font(name='宋体', size=10, bold=True)
    ws['K18'].alignment = label_align
    ws['K18'].border = med_rtb

    ws['L18'] = ' '
    ws['L18'].font = Font(name='宋体', size=11)
    ws['L18'].alignment = Alignment(horizontal='center', vertical='center')
    ws['L18'].border = med_rtb_med
    ws.merge_cells('L18:S18')

    # Row 19: Column headers with merges
    ws.row_dimensions[19].height = 12.0
    thin_bottom = Border(bottom=Side(style='thin'))
    hdr_font = Font(name='宋体', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center')

    ws.cell(row=19, column=1, value='项号').font = hdr_font
    ws.cell(row=19, column=1).alignment = hdr_align
    ws.cell(row=19, column=1).border = Border(bottom=Side(style='thin'), left=Side(style='thin'))

    ws.cell(row=19, column=2, value='商品编号').font = Font(name='宋体', size=10, color='FF0000')
    ws.cell(row=19, column=2).alignment = hdr_align
    ws.cell(row=19, column=2).border = thin_bottom
    ws.merge_cells('B19:C19')

    ws.cell(row=19, column=4, value='商品名称及规格型号').font = hdr_font
    ws.cell(row=19, column=4).alignment = hdr_align
    ws.cell(row=19, column=4).border = thin_bottom
    ws.merge_cells('D19:F19')

    ws.cell(row=19, column=7, value='数量及单位').font = hdr_font
    ws.cell(row=19, column=7).alignment = hdr_align
    ws.cell(row=19, column=7).border = thin_bottom
    ws.merge_cells('G19:H19')

    ws.cell(row=19, column=9, value='单价/总价/币制').font = Font(name='宋体', size=10, color='FF0000')
    ws.cell(row=19, column=9).alignment = hdr_align
    ws.cell(row=19, column=9).border = thin_bottom
    ws.merge_cells('I19:J19')

    ws.cell(row=19, column=11, value='原产国（地区）').font = hdr_font
    ws.cell(row=19, column=11).alignment = hdr_align
    ws.cell(row=19, column=11).border = thin_bottom
    ws.merge_cells('K19:L19')

    ws.cell(row=19, column=13, value='最终目的国（地区）').font = hdr_font
    ws.cell(row=19, column=13).alignment = hdr_align
    ws.cell(row=19, column=13).border = thin_bottom
    ws.merge_cells('M19:O19')

    ws.cell(row=19, column=16, value='境内货源地').font = Font(name='宋体', size=10, color='FF0000')
    ws.cell(row=19, column=16).alignment = hdr_align
    ws.cell(row=19, column=16).border = thin_bottom
    ws.merge_cells('P19:R19')

    ws.cell(row=19, column=19, value='征免').font = hdr_font
    ws.cell(row=19, column=19).alignment = hdr_align
    ws.cell(row=19, column=19).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    # Items: 3 rows per SKU
    row = 20
    item_no = 1
    supplier_city = contract.get('supplier', {}).get('city', '义乌')

    item_font = Font(name='宋体', size=11)
    item_font_12 = Font(name='宋体', size=12)

    for item in items:
        sku = sku_key(item)
        qty = tq.get(sku, 0)
        if qty == 0:
            continue

        amt = ta.get(sku, 0)
        shipping = ship_alloc.get(sku, 0)
        tax_excl = amt / 1.13
        cnf_total_rmb = tax_excl + shipping
        unit_usd = cnf_total_rmb / qty / rate if qty > 0 else 0
        total_usd = cnf_total_rmb / rate

        info = _info(sku, item)
        nw_kg = item.get('net_weight_kg', 0) * (qty / (item.get('packing_rate', 1) or 1))

        # Set row heights: row1=14.25
        ws.row_dimensions[row].height = 14.25

        # Row 1: main info with merges
        c1 = ws.cell(row=row, column=1, value=item_no)
        c1.font = item_font
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.border = Border(left=Side(style='thin'))

        c2 = ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000)
        c2.font = item_font
        c2.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{row}:C{row}')

        c4 = ws.cell(row=row, column=4, value=item.get('name_cn', ''))
        c4.font = item_font
        c4.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'D{row}:F{row}')

        c7 = ws.cell(row=row, column=7, value=qty)
        c7.font = item_font
        c7.alignment = Alignment(vertical='center')

        ws.cell(row=row, column=8, value='个').font = item_font
        ws.cell(row=row, column=8).alignment = Alignment(vertical='center')

        c9 = ws.cell(row=row, column=9, value=round(unit_usd, 6))
        c9.font = item_font
        c9.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'I{row}:J{row}')

        c11 = ws.cell(row=row, column=11, value='中国')
        c11.font = item_font
        c11.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'K{row}:L{row}')

        c13 = ws.cell(row=row, column=13, value='美国')
        c13.font = item_font_12
        c13.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'M{row}:O{row}')

        c16 = ws.cell(row=row, column=16, value=f'{supplier_city}(33189)')
        c16.font = item_font_12
        c16.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'P{row}:R{row}')

        c19 = ws.cell(row=row, column=19, value='照章征税')
        c19.font = item_font
        c19.alignment = Alignment(vertical='center')
        c19.border = Border(right=Side(style='thin'))

        # Row 2: declaration elements + net weight + total price
        row += 1
        ws.merge_cells(f'B{row}:C{row}')
        c4_r2 = ws.cell(row=row, column=4, value=info['declaration_elements'])
        c4_r2.font = Font(name='宋体', size=10)
        c4_r2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c4_r2.border = Border(bottom=Side(style='thin'))
        ws.merge_cells(f'D{row}:F{row+1}')

        c7_r2 = ws.cell(row=row, column=7, value=round(nw_kg, 1))
        c7_r2.font = item_font
        c7_r2.alignment = Alignment(vertical='center')

        ws.cell(row=row, column=8, value='千克').font = item_font
        ws.cell(row=row, column=8).alignment = Alignment(vertical='center')

        c9_r2 = ws.cell(row=row, column=9, value=round(total_usd, 2))
        c9_r2.font = item_font
        c9_r2.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'I{row}:J{row}')

        ws.cell(row=row, column=11, value='（CHN）').font = item_font
        ws.cell(row=row, column=11).alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'K{row}:L{row}')
        ws.cell(row=row, column=13, value='（USA）').font = item_font
        ws.cell(row=row, column=13).alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'M{row}:O{row}')
        ws.merge_cells(f'P{row}:R{row}')

        # Row 3: qty in 个 + currency (with bottom borders)
        row += 1
        ws.merge_cells(f'B{row}:C{row}')

        c7_r3 = ws.cell(row=row, column=7, value=qty)
        c7_r3.font = item_font
        c7_r3.alignment = Alignment(vertical='center')
        c7_r3.border = Border(bottom=Side(style='thin'))

        ws.cell(row=row, column=8, value='个').font = item_font
        ws.cell(row=row, column=8).alignment = Alignment(vertical='center')
        ws.cell(row=row, column=8).border = Border(bottom=Side(style='thin'))

        c9_r3 = ws.cell(row=row, column=9, value='美元')
        c9_r3.font = item_font
        c9_r3.alignment = Alignment(horizontal='right', vertical='center')
        c9_r3.border = Border(bottom=Side(style='thin'))
        ws.merge_cells(f'I{row}:J{row}')

        ws.merge_cells(f'K{row}:L{row}')
        ws.merge_cells(f'M{row}:O{row}')
        ws.merge_cells(f'P{row}:R{row}')

        row += 1
        item_no += 1

    fn = f'{suffix}出口报关单草稿.xlsx' if suffix else '出口报关单草稿.xlsx'
    fp = os.path.join(out_dir, fn)
    wb.save(fp)
    return fp
