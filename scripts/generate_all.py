#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate all customs declaration documents:
1. Export Contract (出口合同)
2. Invoice & Packing List (IV&PL)
3. Declaration Draft (报关单草稿)

Input: parsed purchase_contract.json + fba_shipments.json + user parameters
Output: 3 Excel files per selected ticket group
"""

import argparse
import json
import math
import os
import sys
from typing import Dict, List, Any, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ─── helpers ───────────────────────────────────────────────────────────

def _get_lwh(item: dict) -> Tuple[float, float, float]:
    """Extract length, width, height from package_size_cm array."""
    sz = item.get('package_size_cm', [0, 0, 0])
    if len(sz) >= 3:
        return float(sz[0]), float(sz[1]), float(sz[2])
    return 0.0, 0.0, 0.0


def _sku_key(item: dict) -> str:
    """Get the SKU identifier from a contract item."""
    return item.get('fba_sku', item.get('sku', ''))


# ─── SKU matching between FBA matrix and purchase contract ─────────

def build_sku_mapping(contract_items: list, fba_matrix: dict) -> Dict[str, str]:
    """
    Build a mapping from contract SKU → FBA matrix SKU.
    First tries exact match, then falls back to quantity-based matching.
    Returns {contract_sku: matrix_sku}
    """
    contract_skus = {_sku_key(item): item['quantity'] for item in contract_items}
    matrix_skus = {sku: sum(wh.values()) for sku, wh in fba_matrix.items()}

    mapping = {}
    unmatched_contract = {}
    unmatched_matrix = dict(matrix_skus)

    # Phase 1: exact match
    for c_sku, c_qty in contract_skus.items():
        if c_sku in matrix_skus:
            mapping[c_sku] = c_sku
            unmatched_matrix.pop(c_sku, None)
        else:
            unmatched_contract[c_sku] = c_qty

    # Phase 2: match remaining by quantity
    for c_sku, c_qty in unmatched_contract.items():
        for m_sku, m_qty in list(unmatched_matrix.items()):
            if c_qty == m_qty:
                mapping[c_sku] = m_sku
                unmatched_matrix.pop(m_sku)
                break

    # Phase 3: if still unmatched, warn and skip
    for c_sku in unmatched_contract:
        if c_sku not in mapping:
            print(f"WARNING: Contract SKU '{c_sku}' has no match in FBA matrix", file=sys.stderr)
            mapping[c_sku] = c_sku  # use as-is, will get 0 qty from matrix

    return mapping


# ─── core class ────────────────────────────────────────────────────────

class CustomsDeclarationGenerator:
    def __init__(
        self,
        contract_json: str,
        shipments_json: str,
        groups: List[Dict[str, Any]],
        selected_group_indices: List[int],
        exchange_rate: float,
        shipping_rate: float,
        output_dir: str,
        price_term: str = "CNF",
        knowledge_base: Optional[str] = None,
        template_dir: Optional[str] = None,
    ):
        self.contract = self._load_json(contract_json)
        self.shipments = self._load_json(shipments_json)
        self.groups = groups
        self.selected = selected_group_indices
        self.rate = exchange_rate
        self.ship_rate = shipping_rate
        self.out_dir = output_dir
        self.price_term = price_term
        self.kb = self._load_kb(knowledge_base) if knowledge_base else {}
        self.template_dir = template_dir

        os.makedirs(output_dir, exist_ok=True)

        # Items shorthand
        self.items = self.contract.get('items', [])
        # FBA matrix: {sku: {warehouse: qty}}
        self.matrix = self.shipments.get('matrix', {})
        # SKU mapping: contract_sku → matrix_sku
        self.sku_map = build_sku_mapping(self.items, self.matrix)

    @staticmethod
    def _load_json(path: str) -> dict:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_kb(self, path: str) -> Dict[str, dict]:
        """Load knowledge base Excel (SKU → tariff/name/elements/material)."""
        from openpyxl import load_workbook
        kb = {}
        try:
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                kb[str(row[0])] = {
                    'tariff_code': row[1] if len(row) > 1 and row[1] else '',
                    'english_name': row[2] if len(row) > 2 and row[2] else '',
                    'declaration_elements': row[3] if len(row) > 3 and row[3] else '0|0|塑料|塑料草坪|无品牌|无型号',
                    'material': row[4] if len(row) > 4 and row[4] else 'plastic',
                }
        except Exception as e:
            print(f"Warning: KB load failed: {e}", file=sys.stderr)
        return kb

    def _info(self, sku: str, item: dict) -> dict:
        """Get SKU info from KB or contract item fallback."""
        if sku in self.kb:
            return self.kb[sku]
        return {
            'tariff_code': '',
            'english_name': item.get('name_en', ''),
            'declaration_elements': '0|0|塑料|塑料草坪|无品牌|无型号',
            'material': 'plastic',
        }

    # ─── quantity computation ──────────────────────────────────────────

    def _group_qty(self, group_idx: int) -> Dict[str, int]:
        """Compute {contract_sku: qty} for a ticket group from the FBA matrix."""
        whs = self.groups[group_idx]['warehouses']
        result = {}
        for item in self.items:
            c_sku = _sku_key(item)
            m_sku = self.sku_map.get(c_sku, c_sku)
            wh_data = self.matrix.get(m_sku, {})
            total = sum(wh_data.get(w, 0) for w in whs)
            if total > 0:
                result[c_sku] = total
        return result

    # ─── chargeable weight & shipping ──────────────────────────────────

    def _chargeable(self) -> Tuple[float, float, float, float]:
        """
        Compute chargeable weight from ALL contract items (full quantities).
        Returns (total_gross, total_vol, chargeable_weight, total_shipping_rmb).
        """
        total_gross = 0.0
        total_vol = 0.0
        for item in self.items:
            qty = item['quantity']
            pr = item.get('packing_rate', 1) or 1
            boxes = qty / pr
            l, w, h = _get_lwh(item)
            total_gross += item.get('gross_weight_kg', 0) * boxes
            total_vol += (l * w * h / 6000) * boxes
        chargeable = max(total_gross, total_vol)
        return total_gross, total_vol, chargeable, chargeable * self.ship_rate

    def _shipping_alloc(self, total_ship: float, total_vol: float, total_gross: float) -> Dict[str, float]:
        """Allocate shipping to each contract SKU by proportion."""
        use_vol = total_vol > total_gross
        alloc = {}
        for item in self.items:
            sku = _sku_key(item)
            qty = item['quantity']
            pr = item.get('packing_rate', 1) or 1
            boxes = qty / pr
            l, w, h = _get_lwh(item)
            sku_vol = (l * w * h / 6000) * boxes
            sku_gross = item.get('gross_weight_kg', 0) * boxes
            if use_vol:
                prop = sku_vol / total_vol if total_vol > 0 else 0
            else:
                prop = sku_gross / total_gross if total_gross > 0 else 0
            alloc[sku] = total_ship * prop
        return alloc

    # ─── amount allocation across tickets ──────────────────────────────

    def _amount_alloc(self) -> Dict[int, Dict[str, float]]:
        """
        Allocate purchase amounts across selected tickets proportionally.
        {group_idx: {contract_sku: allocated_rmb_amount}}
        """
        # Total qty per SKU across ALL selected tickets
        totals = {}
        group_qtys = {}
        for gi in self.selected:
            gq = self._group_qty(gi)
            group_qtys[gi] = gq
            for sku, q in gq.items():
                totals[sku] = totals.get(sku, 0) + q

        result = {}
        for gi in self.selected:
            result[gi] = {}
            for item in self.items:
                sku = _sku_key(item)
                tq = totals.get(sku, 0)
                gq = group_qtys[gi].get(sku, 0)
                if tq > 0 and gq > 0:
                    result[gi][sku] = item['total_amount'] * (gq / tq)
                else:
                    result[gi][sku] = 0
        return result

    # ─── main generate entry ───────────────────────────────────────────

    def generate(self) -> str:
        total_gross, total_vol, chargeable, total_ship = self._chargeable()
        ship_alloc = self._shipping_alloc(total_ship, total_vol, total_gross)
        amt_alloc = self._amount_alloc()

        cno = self.contract.get('contract_no', 'UNKNOWN')
        files = []
        group_names = []

        for gi in self.selected:
            gname = self.groups[gi]['name']
            group_names.append(gname)
            suffix = f'【{gname}】' if len(self.selected) > 1 else ''

            tq = self._group_qty(gi)
            ta = amt_alloc[gi]

            # 1) Export contract
            f1 = self._gen_export_contract(cno, suffix, tq, ta, ship_alloc,
                                           total_gross, total_vol, chargeable, total_ship)
            files.append(os.path.basename(f1))

            # 2) IV & PL
            f2 = self._gen_iv_pl(cno, suffix, tq, ta, ship_alloc)
            files.append(os.path.basename(f2))

            # 3) Declaration draft
            f3 = self._gen_declaration(cno, suffix, tq, ta, ship_alloc, total_ship)
            files.append(os.path.basename(f3))

        # Summary
        all_qty = {}
        total_usd = 0
        for gi in self.selected:
            tq = self._group_qty(gi)
            ta = amt_alloc[gi]
            for item in self.items:
                sku = _sku_key(item)
                q = tq.get(sku, 0)
                a = ta.get(sku, 0)
                all_qty[sku] = all_qty.get(sku, 0) + q
                if q > 0:
                    tax_excl = a / 1.13
                    total_usd += (tax_excl + ship_alloc.get(sku, 0)) / self.rate

        total_boxes = 0
        total_nw = 0.0
        total_gw = 0.0
        for item in self.items:
            sku = _sku_key(item)
            q = all_qty.get(sku, 0)
            if q > 0:
                pr = item.get('packing_rate', 1) or 1
                bx = q / pr
                total_boxes += int(bx)
                total_nw += item.get('net_weight_kg', 0) * bx
                total_gw += item.get('gross_weight_kg', 0) * bx

        summary = {
            'groups_generated': group_names,
            'files': files,
            'summary': {
                'total_qty': sum(all_qty.values()),
                'total_usd': round(total_usd, 2),
                'total_boxes': total_boxes,
                'total_gross_weight': round(total_gw, 1),
                'total_net_weight': round(total_nw, 1),
                'exchange_rate_check': round(total_usd * self.rate / (sum(amt_alloc[self.selected[0]].values()) / 1.13 + total_ship), 4) if total_usd > 0 else 0,
                'chargeable_weight': round(chargeable, 2),
                'total_shipping_rmb': round(total_ship, 2),
            },
        }
        return json.dumps(summary, ensure_ascii=False, indent=2)

    # ─── Export Contract ───────────────────────────────────────────────

    def _gen_export_contract(self, cno, suffix, tq, ta, ship_alloc,
                             total_gross, total_vol, chargeable, total_ship):
        wb = Workbook()
        ws = wb.active
        ws.title = "出口合同"

        # Set column widths (converted from xlrd units ~256 ≈ openpyxl char width)
        col_widths = {
            0: 5441/256, 1: 4053/256, 2: 3626/256, 3: 4864/256, 4: 2304/256, 5: 2474/256,
            6: 2218/256, 7: 4266/256, 8: 4736/256, 9: 3584/256, 10: 3754/256, 11: 4053/256,
            13: 6869/256, 14: 2730/256, 18: 2773/256, 19: 3242/256, 20: 3669/256, 21: 3114/256
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        # Set row heights
        row_heights = {
            0: 42, 1: 20, 2: 20, 3: 20, 4: 27, 5: 20, 6: 20, 7: 20, 8: 20, 9: 20, 10: 20,
            11: 24, 12: 24, 13: 24, 14: 24, 15: 24, 16: 13.5, 17: 13.5, 18: 25, 19: 25
        }
        for row_idx, height in row_heights.items():
            ws.row_dimensions[row_idx + 1].height = height

        # Header section (mimic purchase contract layout)
        ws['A1'] = '采购合同'
        ws['A1'].font = Font(name='宋体', size=24, bold=False)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:L1')

        ws['H2'] = '合同编号：'
        ws['H2'].font = Font(name='宋体', size=10)
        ws['I2'] = cno
        ws['I2'].font = Font(name='宋体', size=10)
        ws['A3'] = '日期：'
        ws['A3'].font = Font(name='宋体', size=10)
        ws['D3'] = self.contract.get('date', '')
        ws['D3'].font = Font(name='宋体', size=10)

        supplier = self.contract.get('supplier', {})
        buyer = self.contract.get('buyer', {})
        ws['A4'] = '供方：'
        ws['A4'].font = Font(name='宋体', size=10)
        ws['D4'] = supplier.get('name', '')
        ws['D4'].font = Font(name='宋体', size=10)
        ws['H4'] = '需方：'
        ws['H4'].font = Font(name='宋体', size=10)
        ws['I4'] = buyer.get('name', '')
        ws['I4'].font = Font(name='宋体', size=10)

        # Column headers (row 11, 1-indexed in openpyxl)
        headers_main = ['产品名称', '产品图片', '规格型号', 'FBA SKU', '单位', '数量',
                        '箱率', '含税单价/元', '包装尺寸/CM', '外箱净重/KG', '外箱毛重/KG', '总额/元']
        headers_calc = ['', '计算逻辑', '', '箱数-计算总海运费使用', '', '', '体积重', '运费平摊', 'C&F总价', 'C&F单价']

        for c, h in enumerate(headers_main, 1):
            cell = ws.cell(row=11, column=c, value=h)
            cell.font = Font(name='宋体', size=10, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for c, h in enumerate(headers_calc, 13):
            cell = ws.cell(row=11, column=c, value=h)
            cell.font = Font(name='宋体', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Item rows
        row = 12
        sum_qty = 0
        sum_amt = 0.0
        calc_labels_written = False

        for item in self.items:
            sku = _sku_key(item)
            qty = tq.get(sku, 0)
            if qty == 0:
                continue

            amt = ta.get(sku, 0)
            unit_price = amt / qty if qty > 0 else 0
            l, w, h_ = _get_lwh(item)
            size_str = f'{int(l)}*{int(w)}*{int(h_)}'

            # Full contract boxes (for shipping calc)
            pr = item.get('packing_rate', 1) or 1
            full_boxes = item['quantity'] / pr
            vol_weight = (l * w * h_ / 6000) * full_boxes
            shipping = ship_alloc.get(sku, 0)
            tax_excl = amt / 1.13
            cnf_total = tax_excl + shipping
            cnf_unit = cnf_total / qty if qty > 0 else 0

            name_cn = item.get('name_cn', '')
            name_en = item.get('name_en', '')
            c1 = ws.cell(row=row, column=1, value=f'{name_cn}\n{name_en}')
            c1.font = Font(name='宋体', size=10)
            c1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            ws.cell(row=row, column=3, value=item.get('spec', '')).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=4, value=sku).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=5, value=item.get('unit', '件')).font = Font(name='宋体', size=10)
            c6 = ws.cell(row=row, column=6, value=qty)
            c6.font = Font(name='宋体', size=10)
            c6.number_format = '#,##0'
            c6.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=7, value=int(pr)).font = Font(name='宋体', size=10)
            c8 = ws.cell(row=row, column=8, value=round(unit_price, 2))
            c8.font = Font(name='宋体', size=10)
            c8.number_format = '0.00_ '
            c8.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=9, value=size_str).font = Font(name='宋体', size=10)
            c10 = ws.cell(row=row, column=10, value=item.get('net_weight_kg', 0))
            c10.font = Font(name='宋体', size=10)
            c10.number_format = '0.00_ '
            c11 = ws.cell(row=row, column=11, value=item.get('gross_weight_kg', 0))
            c11.font = Font(name='宋体', size=10)
            c11.number_format = '0.00_ '
            c12 = ws.cell(row=row, column=12, value=round(amt, 2))
            c12.font = Font(name='宋体', size=10)
            c12.number_format = '0.0000_ '
            c12.alignment = Alignment(horizontal='center', vertical='center')

            # Right-side calculation columns
            ws.cell(row=row, column=16, value=round(full_boxes)).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=19, value=round(vol_weight, 2)).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=20, value=round(shipping, 2)).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=21, value=round(cnf_total, 2)).font = Font(name='宋体', size=10)
            ws.cell(row=row, column=22, value=round(cnf_unit, 2)).font = Font(name='宋体', size=10)

            # Write calculation labels on first data rows
            if not calc_labels_written:
                # Labels go in rows 12-15 col N-O (14-15)
                ws.cell(row=12, column=14, value='').font = Font(name='宋体', size=10)
                ws.cell(row=13, column=14, value='总毛重').font = Font(name='宋体', size=10)
                ws.cell(row=13, column=15, value=round(total_gross, 1)).font = Font(name='宋体', size=10)
                ws.cell(row=14, column=14, value='总体积重').font = Font(name='宋体', size=10)
                ws.cell(row=14, column=15, value=round(total_vol, 2)).font = Font(name='宋体', size=10)
                ws.cell(row=15, column=14, value='计费重').font = Font(name='宋体', size=10)
                ws.cell(row=15, column=15, value=round(chargeable, 2)).font = Font(name='宋体', size=10)
                ws.cell(row=16, column=14, value=f'总海运费（单价{self.ship_rate}元/kg）').font = Font(name='宋体', size=10)
                ws.cell(row=16, column=15, value=round(total_ship, 2)).font = Font(name='宋体', size=10)
                calc_labels_written = True

            sum_qty += qty
            sum_amt += amt
            row += 1

        # Totals
        ws.cell(row=row, column=1, value='小写合计').font = Font(name='宋体', size=10)
        ws.cell(row=row, column=6, value=sum_qty).font = Font(name='宋体', size=10)
        row += 1
        ws.cell(row=row, column=1, value='大写合计').font = Font(name='宋体', size=10)
        row += 2
        ws.cell(row=row, column=1, value='共计').font = Font(name='宋体', size=10)
        ws.cell(row=row, column=12, value=round(sum_amt, 2)).font = Font(name='宋体', size=10)

        # Exchange rate validation
        ws.cell(row=row, column=14, value=f'报关汇率{self.rate}').font = Font(name='宋体', size=10)
        # Validation: (tax_excl_total + total_shipping) / invoice_usd ≈ exchange_rate
        tax_excl_total = sum_amt / 1.13
        inv_usd = (tax_excl_total + total_ship) / self.rate
        ws.cell(row=row, column=15, value=round(tax_excl_total + total_ship, 2)).font = Font(name='宋体', size=10)

        fn = f'【{cno}】{suffix}出口合同.xlsx'
        fp = os.path.join(self.out_dir, fn)
        wb.save(fp)
        return fp

    # ─── Invoice & Packing List ────────────────────────────────────────

    def _gen_iv_pl(self, cno, suffix, tq, ta, ship_alloc):
        wb = Workbook()
        wb.remove(wb.active)

        iv = wb.create_sheet('IV')
        self._fill_iv(iv, cno, tq, ta, ship_alloc)

        pl = wb.create_sheet('PL')
        self._fill_pl(pl, cno, tq)

        fn = f'【{cno}】{suffix}IV&PL.xlsx'
        fp = os.path.join(self.out_dir, fn)
        wb.save(fp)
        return fp

    def _fill_iv(self, ws, cno, tq, ta, ship_alloc):
        # Set column widths
        col_widths = {0: 4793/256, 1: 6400/256, 2: 6609/256, 3: 2669/256, 4: 3257/256,
                      5: 3840/256, 6: 3072/256, 7: 3769/256, 8: 5678/256, 9: 4430/256}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        # Set row heights
        row_heights = {0: 30, 1: 30, 2: 70, 3: 30, 4: 30, 5: 67, 6: 30, 7: 30, 8: 51, 9: 56}
        for row_idx, height in row_heights.items():
            ws.row_dimensions[row_idx + 1].height = height
        for r in range(10, 13):
            ws.row_dimensions[r + 1].height = 59
        ws.row_dimensions[14].height = 30
        ws.row_dimensions[15].height = 30

        # Title
        ws['A1'] = 'INVOICE'
        ws['A1'].font = Font(name='Arial', size=16, bold=False)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='top')
        ws.merge_cells('A1:I1')

        # Header - with merges (each pair: 2 columns merged)
        ws['A2'] = 'Shipper:公司名称'
        ws['A2'].font = Font(name='Arial', size=12)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B2:C2')
        ws['B2'] = 'Shenzhen Adhoc Trading Co., Ltd.'
        ws['B2'].font = Font(name='Arial', size=10)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D2:E2')
        ws['D2'] = 'INVOICE NO:'
        ws['D2'].font = Font(name='Arial', size=12)
        ws['D2'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F2:G2')
        ws['F2'] = cno
        ws['F2'].font = Font(name='Arial', size=12)
        ws['F2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A3'] = 'ADD:地址'
        ws['A3'].font = Font(name='Arial', size=12)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B3:C3')
        ws['B3'] = 'Flat 1006, Zhenye International Business Centre,No.3101-90,Qianhai Road, Nanshan District, Shenzhen,China.'
        ws['B3'].font = Font(name='Arial', size=10)
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D3:E3')
        ws['D3'] = 'DATE:'
        ws['D3'].font = Font(name='Arial', size=12)
        ws['D3'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F3:G3')
        ws['F3'] = datetime.now().strftime('%Y/%m/%d')
        ws['F3'].font = Font(name='Arial', size=12)
        ws['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('D4:E4')
        ws['D4'] = 'ORIGIN COUNTRY :'
        ws['D4'].font = Font(name='Arial', size=12)
        ws['D4'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F4:G4')
        ws['F4'] = 'CHINA'
        ws['F4'].font = Font(name='Arial', size=12)
        ws['F4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A5'] = 'CNEE:公司名称'
        ws['A5'].font = Font(name='Arial', size=12)
        ws['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells('B5:C5')
        ws['B5'] = 'ZEATALINE INTERNATIONAL TRADING'
        ws['B5'].font = Font(name='Arial', size=10)
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D5:E5')
        ws['D5'] = 'PRICE TERMS:'
        ws['D5'].font = Font(name='Arial', size=12)
        ws['D5'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['F5'] = 'C&F'
        ws['F5'].font = Font(name='Arial', size=11)
        ws['F5'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['G5'] = 'USA'
        ws['G5'].font = Font(name='宋体', size=10)
        ws['G5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A6'] = 'ADD:地址'
        ws['A6'].font = Font(name='Arial', size=12)
        ws['A6'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B6:C6')
        ws['B6'] = 'zouzhichang\n890 S Azusa Ave\nCity of Industry, CA 91748\nUS'
        ws['B6'].font = Font(name='Arial', size=10)
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D6:E6')
        ws['D6'] = 'Reference No.'
        ws['D6'].font = Font(name='Arial', size=12)
        ws['D6'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # Column headers - English
        hdrs = ['No.', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
                'Unit Price', 'USD', 'Total Amount', 'material quality', 'picture']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=7, column=c, value=h)
            cell.font = Font(name='Arial', size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Column headers - Chinese
        cn_hdrs = ['', '海关编码（清关用的）', '英文品名', '数量', 'PC(S)',
                   '单价', 'USD', '总价', '材质', '产品图片']
        for c, h in enumerate(cn_hdrs, 1):
            cell = ws.cell(row=8, column=c, value=h)
            cell.font = Font(name='宋体', size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row = 9
        total_usd = 0.0
        total_qty = 0
        idx = 1

        for item in self.items:
            sku = _sku_key(item)
            qty = tq.get(sku, 0)
            if qty == 0:
                continue

            amt = ta.get(sku, 0)
            shipping = ship_alloc.get(sku, 0)
            tax_excl = amt / 1.13
            cnf_total_rmb = tax_excl + shipping
            unit_usd = cnf_total_rmb / qty / self.rate if qty > 0 else 0
            total_item_usd = cnf_total_rmb / self.rate

            info = self._info(sku, item)

            c1 = ws.cell(row=row, column=1, value=idx)
            c1.font = Font(name='Arial', size=10)
            c1.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000).font = Font(name='Arial', size=12)
            ws.cell(row=row, column=3, value=info['english_name'] or item.get('name_en', '')).font = Font(name='Arial', size=12)

            c4 = ws.cell(row=row, column=4, value=qty)
            c4.font = Font(name='Arial', size=12)
            c4.number_format = '0_);[Red]\\(0\\)'
            c4.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=5, value='PC(S)').font = Font(name='Arial', size=12)

            c6 = ws.cell(row=row, column=6, value=unit_usd)
            c6.font = Font(name='Arial', size=12)
            c6.number_format = '0.00_);[Red]\\(0.00\\)'
            c6.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=7, value='USD').font = Font(name='Arial', size=12)

            c8 = ws.cell(row=row, column=8, value=total_item_usd)
            c8.font = Font(name='Arial', size=12)
            c8.number_format = '0.00_);[Red]\\(0.00\\)'
            c8.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=9, value=info['material'] or 'plastic').font = Font(name='Arial', size=12)

            total_usd += total_item_usd
            total_qty += qty
            idx += 1
            row += 1

        # Empty placeholder row
        ws.cell(row=row, column=1, value=idx).font = Font(name='Arial', size=12)
        row += 1

        # Totals
        c4_tot = ws.cell(row=row, column=4, value=total_qty)
        c4_tot.font = Font(name='Arial', size=12)
        c4_tot.number_format = '#0'
        c4_tot.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row, column=5, value='/').font = Font(name='Arial', size=12)
        ws.cell(row=row, column=6, value='/').font = Font(name='Arial', size=12)
        ws.cell(row=row, column=7, value='/').font = Font(name='Arial', size=12)

        c8_tot = ws.cell(row=row, column=8, value=total_usd)
        c8_tot.font = Font(name='Arial', size=12)
        c8_tot.number_format = '0.00_);[Red]\\(0.00\\)'
        c8_tot.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row, column=9, value='/').font = Font(name='Arial', size=12)
        ws.cell(row=row, column=10, value='/').font = Font(name='Arial', size=12)

    def _fill_pl(self, ws, cno, tq):
        # Set column widths
        col_widths = {0: 4025/256, 1: 7097/256, 2: 6609/256, 4: 3257/256, 5: 3840/256,
                      6: 3072/256, 7: 3118/256, 8: 3140/256, 9: 5032/256}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        # Set row heights
        row_heights = {0: 30, 1: 30, 2: 61, 3: 30, 4: 30, 5: 67, 6: 30, 7: 30, 8: 77}
        for row_idx, height in row_heights.items():
            ws.row_dimensions[row_idx + 1].height = height
        for r in range(9, 13):
            ws.row_dimensions[r + 1].height = 68 + (r - 9) * 2
        ws.row_dimensions[14].height = 30
        ws.row_dimensions[15].height = 30
        ws.row_dimensions[16].height = 30

        # Title
        ws['A1'] = 'PACKING LIST'
        ws['A1'].font = Font(name='Arial', size=16, bold=False)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='top')
        ws.merge_cells('A1:I1')

        # Header - with merges (each pair: 2 columns merged)
        ws['A2'] = 'Shipper:'
        ws['A2'].font = Font(name='Arial', size=12)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B2:C2')
        ws['B2'] = 'Shenzhen Adhoc Trading Co., Ltd.'
        ws['B2'].font = Font(name='Arial', size=10)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D2:E2')
        ws['D2'] = 'INVOICE NO:'
        ws['D2'].font = Font(name='Arial', size=12)
        ws['D2'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F2:G2')
        ws['F2'] = cno
        ws['F2'].font = Font(name='Arial', size=12)
        ws['F2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A3'] = 'ADD:'
        ws['A3'].font = Font(name='Arial', size=12)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B3:C3')
        ws['B3'] = 'Flat 1006, Zhenye International Business Centre,No.3101-90,Qianhai Road, Nanshan District, Shenzhen,China.'
        ws['B3'].font = Font(name='Arial', size=10)
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D3:E3')
        ws['D3'] = 'DATE:'
        ws['D3'].font = Font(name='Arial', size=12)
        ws['D3'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F3:G3')
        ws['F3'] = datetime.now().strftime('%Y/%m/%d')
        ws['F3'].font = Font(name='宋体', size=12)
        ws['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('D4:E4')
        ws['D4'] = 'ORIGIN COUNTRY :'
        ws['D4'].font = Font(name='Arial', size=12)
        ws['D4'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.merge_cells('F4:G4')
        ws['F4'] = 'CHINA'
        ws['F4'].font = Font(name='Arial', size=12)
        ws['F4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A5'] = 'TO:'
        ws['A5'].font = Font(name='Arial', size=12)
        ws['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells('B5:C5')
        ws['B5'] = 'ZEATALINE INTERNATIONAL TRADING'
        ws['B5'].font = Font(name='Arial', size=10)
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D5:E5')
        ws['D5'] = 'PRICE TERMS:'
        ws['D5'].font = Font(name='Arial', size=12)
        ws['D5'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['F5'] = 'C&F'
        ws['F5'].font = Font(name='Arial', size=11)
        ws['F5'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws['G5'] = 'USA'
        ws['G5'].font = Font(name='宋体', size=10)
        ws['G5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A6'] = 'ADD:'
        ws['A6'].font = Font(name='Arial', size=12)
        ws['A6'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('B6:C6')
        ws['B6'] = 'zouzhichang\n890 S Azusa Ave\nCity of Industry, CA 91748\nUS'
        ws['B6'].font = Font(name='Arial', size=10)
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D6:E6')
        ws['D6'] = 'Reference No.'
        ws['D6'].font = Font(name='Arial', size=12)
        ws['D6'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # Column headers - English
        hdrs = ['NO', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
                'Box Qty\n(CTNS)', 'N.W.\n(KG)', 'G.W.\n(KG)', 'VOLUME (CBM)', '']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=7, column=c, value=h)
            cell.font = Font(name='Arial', size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Column headers - Chinese
        cn_hdrs = ['', '海关编码（清关用的）', '英文品名', '数量', 'PC(S)',
                   '箱数', '净重', '毛重', '方数', '产品图片']
        for c, h in enumerate(cn_hdrs, 1):
            cell = ws.cell(row=8, column=c, value=h)
            cell.font = Font(name='宋体', size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row = 9
        tot_qty = 0
        tot_boxes = 0
        tot_nw = 0.0
        tot_gw = 0.0
        tot_vol = 0.0
        idx = 1

        for item in self.items:
            sku = _sku_key(item)
            qty = tq.get(sku, 0)
            if qty == 0:
                continue

            pr = item.get('packing_rate', 1) or 1
            boxes = qty / pr  # may be fractional for display
            l, w, h_ = _get_lwh(item)
            nw = item.get('net_weight_kg', 0) * boxes
            gw = item.get('gross_weight_kg', 0) * boxes
            volume = (l * w * h_ / 1_000_000) * boxes

            info = self._info(sku, item)

            c1 = ws.cell(row=row, column=1, value=idx)
            c1.font = Font(name='Arial', size=10)
            c1.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000).font = Font(name='Arial', size=12)
            ws.cell(row=row, column=3, value=info['english_name'] or item.get('name_en', '')).font = Font(name='Arial', size=12)

            c4 = ws.cell(row=row, column=4, value=qty)
            c4.font = Font(name='宋体', size=14)
            c4.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=5, value='PC(S)').font = Font(name='Arial', size=12)

            c6 = ws.cell(row=row, column=6, value=boxes)
            c6.font = Font(name='宋体', size=14)
            c6.alignment = Alignment(horizontal='center', vertical='center')

            c7 = ws.cell(row=row, column=7, value=round(nw, 1))
            c7.font = Font(name='Arial', size=12)
            c7.number_format = '0.00'
            c7.alignment = Alignment(horizontal='center', vertical='center')

            c8 = ws.cell(row=row, column=8, value=round(gw, 1))
            c8.font = Font(name='Arial', size=12)
            c8.number_format = '0.00'
            c8.alignment = Alignment(horizontal='center', vertical='center')

            c9 = ws.cell(row=row, column=9, value=round(volume, 5))
            c9.font = Font(name='Arial', size=12)
            c9.number_format = '0.00'
            c9.alignment = Alignment(horizontal='center', vertical='center')

            tot_qty += qty
            tot_boxes += boxes
            tot_nw += nw
            tot_gw += gw
            tot_vol += volume
            idx += 1
            row += 1

        # Empty placeholder
        ws.cell(row=row, column=1, value=idx).font = Font(name='Arial', size=12)
        ws.cell(row=row, column=5, value='PC(S)').font = Font(name='Arial', size=12)
        row += 1

        # Totals
        c4_tot = ws.cell(row=row, column=4, value=tot_qty)
        c4_tot.font = Font(name='Arial', size=12)
        c4_tot.number_format = '#0'
        c4_tot.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row, column=5, value='/').font = Font(name='Arial', size=12)

        c6_tot = ws.cell(row=row, column=6, value=round(tot_boxes))
        c6_tot.font = Font(name='Arial', size=12)
        c6_tot.number_format = '#0'
        c6_tot.alignment = Alignment(horizontal='center', vertical='center')

        c7_tot = ws.cell(row=row, column=7, value=round(tot_nw, 1))
        c7_tot.font = Font(name='Arial', size=12)
        c7_tot.number_format = '#0.00'
        c7_tot.alignment = Alignment(horizontal='center', vertical='center')

        c8_tot = ws.cell(row=row, column=8, value=round(tot_gw, 1))
        c8_tot.font = Font(name='Arial', size=12)
        c8_tot.number_format = '#0.00'
        c8_tot.alignment = Alignment(horizontal='center', vertical='center')

        c9_tot = ws.cell(row=row, column=9, value=round(tot_vol, 5))
        c9_tot.font = Font(name='Arial', size=12)
        c9_tot.number_format = '#0.00'
        c9_tot.alignment = Alignment(horizontal='center', vertical='center')

    # ─── Declaration Draft ─────────────────────────────────────────────

    def _gen_declaration(self, cno, suffix, tq, ta, ship_alloc, total_ship):
        wb = Workbook()
        ws = wb.active
        ws.title = '报关单草稿'

        # Set column widths (A-S: openpyxl character widths)
        col_widths = {
            'A': 6.27, 'B': 7.37, 'C': 6.0, 'D': 25.45, 'E': 10.91, 'F': 12.37, 'G': 11.0,
            'H': 6.73, 'I': 7.91, 'J': 7.09, 'K': 7.63, 'L': 5.37, 'M': 5.27, 'N': 5.09,
            'O': 7.27, 'P': 4.09, 'Q': 6.27, 'R': 7.63, 'S': 8.45
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Set row heights (matching example template)
        ws.row_dimensions[1].height = 25.5
        ws.row_dimensions[2].height = 12.75
        ws.row_dimensions[3].height = 12.0
        ws.row_dimensions[4].height = 14.25
        ws.row_dimensions[5].height = 12.0
        ws.row_dimensions[7].height = 12.0
        ws.row_dimensions[8].height = 14.25
        ws.row_dimensions[9].height = 12.0
        ws.row_dimensions[10].height = 14.25
        ws.row_dimensions[11].height = 12.0
        ws.row_dimensions[12].height = 14.25
        ws.row_dimensions[13].height = 12.0
        ws.row_dimensions[14].height = 12.75
        ws.row_dimensions[15].height = 12.0
        ws.row_dimensions[17].height = 14.25
        ws.row_dimensions[18].height = 14.25

        # Title
        ws.merge_cells('A1:S1')
        ws['A1'] = '中华人民共和国海关出口货物报关单'
        ws['A1'].font = Font(name='宋体', size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Helper: create bordered label cell
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Row 2 (no borders, right-aligned, with merge cells)
        ws.merge_cells('A2:B2')
        c_a2 = ws['A2']
        c_a2.value = '预录入编号：'
        c_a2.font = Font(name='宋体', size=10)
        c_a2.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('C2:E2')

        ws['F2'] = '申报口岸:'
        ws['F2'].font = Font(name='宋体', size=10)
        ws['F2'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('G2:H2')

        ws['I2'] = '海关编号:'
        ws['I2'].font = Font(name='宋体', size=10)
        ws['I2'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('J2:N2')

        # Define border styles
        label_font = Font(name='宋体', size=10)
        label_font_red = Font(name='宋体', size=10, color='FF0000')
        value_font = Font(name='宋体', size=12, bold=True)
        value_font_red = Font(name='宋体', size=12, bold=True, color='FF0000')
        value_font_11 = Font(name='宋体', size=11, bold=True)
        label_align = Alignment(vertical='center')
        label_align_left = Alignment(horizontal='left', vertical='center')
        value_align_left = Alignment(horizontal='left', vertical='center')
        med_left = Side(style='medium')
        med_right = Side(style='medium')
        thin_side = Side(style='thin')

        # Row 3-4: 境内发货人 (merges: A3:B3, C3:D3, H3:J3, L3:N3, P3:S3)
        ws['A3'] = '境内发货人'
        ws['A3'].font = label_font
        ws['A3'].alignment = label_align_left
        ws['A3'].border = Border(top=Side(style='medium'), left=Side(style='medium'))
        ws.merge_cells('A3:B3')
        ws['C3'] = '（91440300687577411Y）'
        ws['C3'].font = label_font
        ws['C3'].alignment = label_align_left
        ws['C3'].border = Border(top=Side(style='medium'), right=thin_side)
        ws.merge_cells('C3:D3')
        ws['E3'] = '出境关别'
        ws['E3'].font = label_font_red
        ws['E3'].alignment = label_align
        ws['E3'].border = Border(top=Side(style='medium'), left=thin_side)
        ws['F3'] = '(    )'
        ws['F3'].font = label_font
        ws['F3'].alignment = label_align
        ws['F3'].border = Border(top=Side(style='medium'), right=thin_side)
        ws['G3'] = '出口日期'
        ws['G3'].font = label_font
        ws['G3'].alignment = label_align
        ws['G3'].border = Border(top=Side(style='medium'), left=thin_side)
        ws.merge_cells('H3:J3')
        ws['K3'] = '申报日期'
        ws['K3'].font = label_font
        ws['K3'].alignment = label_align
        ws['K3'].border = Border(top=Side(style='medium'), left=thin_side)
        ws.merge_cells('L3:N3')
        ws['O3'] = '备案号'
        ws['O3'].font = label_font
        ws['O3'].alignment = label_align
        ws['O3'].border = Border(top=Side(style='medium'), left=thin_side)
        ws.merge_cells('P3:S3')

        # Ensure continuous medium top border across ALL cells in Row 3
        # (merged cells need individual border settings for openpyxl to render properly)
        for col_idx in range(1, 20):  # columns A(1) to S(19)
            cell = ws.cell(row=3, column=col_idx)
            existing = cell.border
            cell.border = Border(
                top=Side(style='medium'),
                left=existing.left if existing.left and existing.left.style else None,
                right=existing.right if existing.right and existing.right.style else None,
                bottom=existing.bottom if existing.bottom and existing.bottom.style else None,
            )
        # Right edge of Row 3 needs medium right border
        s3 = ws['S3']
        s3.border = Border(top=Side(style='medium'), right=Side(style='medium'))

        # Row 4: values
        ws['A4'] = '深圳市艾进贸易有限公司'
        ws['A4'].font = value_font
        ws['A4'].alignment = value_align_left
        ws['A4'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
        ws.merge_cells('A4:D4')
        ws['E4'] = ''
        ws['E4'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('E4:F4')
        ws['G4'] = ''
        ws['G4'].border = Border(bottom=thin_side, left=thin_side)
        ws.merge_cells('G4:I4')
        ws['K4'] = ''
        ws['K4'].border = Border(bottom=thin_side)
        ws.merge_cells('K4:N4')
        ws['O4'] = ''
        ws['O4'].border = Border(bottom=thin_side)
        ws.merge_cells('O4:S4')

        # Row 5-6: 境外收货人 (merges: A5:B5, G5:J5, L5:N5, O5:P5, Q5:S5)
        ws['A5'] = '境外收货人'
        ws['A5'].font = label_font_red
        ws['A5'].alignment = label_align_left
        ws['A5'].border = Border(top=thin_side, left=Side(style='medium'))
        ws.merge_cells('A5:B5')
        ws['C5'] = ' '
        ws['C5'].font = label_font
        ws['C5'].border = Border(top=thin_side)
        ws['E5'] = '运输方式'
        ws['E5'].font = label_font
        ws['E5'].alignment = label_align
        ws['E5'].border = Border(top=thin_side, left=thin_side)
        ws['F5'] = '(    )'
        ws['F5'].font = label_font
        ws['F5'].alignment = label_align
        ws['F5'].border = Border(top=thin_side, right=thin_side)
        ws['G5'] = '运输工具名称及航次号'
        ws['G5'].font = label_font
        ws['G5'].alignment = Alignment(horizontal='left', vertical='center')
        ws['G5'].border = Border(top=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('G5:J5')
        ws['K5'] = '提运单号'
        ws['K5'].font = label_font
        ws['K5'].alignment = label_align
        ws['K5'].border = Border(top=thin_side, left=thin_side)
        ws.merge_cells('L5:N5')
        ws.merge_cells('O5:P5')
        ws.merge_cells('Q5:S5')

        # Row 6: values
        ws['A6'] = 'ZEATALINE INTERNATIONAL TRADING'
        ws['A6'].font = value_font_11
        ws['A6'].alignment = value_align_left
        ws['A6'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
        ws.merge_cells('A6:D6')
        ws['E6'] = '水路运输'
        ws['E6'].font = Font(name='宋体', size=11)
        ws['E6'].alignment = value_align_left
        ws['E6'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('E6:F6')
        ws['G6'] = ''
        ws['G6'].border = Border(bottom=thin_side, left=thin_side)
        ws.merge_cells('G6:J6')
        ws['K6'] = ''
        ws['K6'].border = Border(bottom=thin_side)
        ws.merge_cells('K6:N6')
        ws['O6'] = ''
        ws['O6'].border = Border(bottom=thin_side)
        ws.merge_cells('O6:S6')

        # Row 7-8: 生产销售单位
        ws['A7'] = '生产销售单位'
        ws['A7'].font = label_font
        ws['A7'].alignment = label_align_left
        ws['A7'].border = Border(top=thin_side, left=Side(style='medium'))
        ws.merge_cells('A7:B7')
        ws['C7'] = ' '
        ws['C7'].font = label_font
        ws['C7'].alignment = label_align_left
        ws['C7'].border = Border(top=thin_side, right=thin_side)
        ws.merge_cells('C7:D7')
        ws['E7'] = '监管方式'
        ws['E7'].font = label_font
        ws['E7'].alignment = label_align
        ws['E7'].border = Border(top=thin_side, left=thin_side)
        ws['F7'] = '(    )'
        ws['F7'].font = label_font
        ws['F7'].alignment = label_align
        ws['F7'].border = Border(top=thin_side, right=thin_side)
        ws['G7'] = '征免性质'
        ws['G7'].font = label_font
        ws['G7'].alignment = label_align
        ws['G7'].border = Border(top=thin_side, left=thin_side)
        ws['H7'] = '(    )'
        ws['H7'].font = label_font
        ws['H7'].alignment = label_align_left
        ws['H7'].border = Border(top=thin_side, right=thin_side)
        ws.merge_cells('H7:J7')
        ws['K7'] = '许可证号'
        ws['K7'].font = label_font
        ws['K7'].alignment = label_align
        ws['K7'].border = Border(top=thin_side, left=thin_side)
        ws.merge_cells('L7:N7')
        ws.merge_cells('P7:S7')

        # Row 8: values
        ws['A8'] = '深圳市艾进贸易有限公司'
        ws['A8'].font = value_font
        ws['A8'].alignment = value_align_left
        ws['A8'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
        ws.merge_cells('A8:D8')
        ws['E8'] = '一般贸易'
        ws['E8'].font = value_font
        ws['E8'].alignment = value_align_left
        ws['E8'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('E8:F8')
        ws['G8'] = '一般征税'
        ws['G8'].font = value_font
        ws['G8'].alignment = value_align_left
        ws['G8'].border = Border(bottom=thin_side, left=thin_side)
        ws.merge_cells('G8:I8')
        ws['K8'] = ''
        ws['K8'].border = Border(bottom=thin_side)
        ws.merge_cells('K8:N8')
        ws['O8'] = ''
        ws['O8'].border = Border(bottom=thin_side)
        ws.merge_cells('O8:S8')

        # Row 9-10: 合同协议号
        ws['A9'] = '合同协议号'
        ws['A9'].font = label_font
        ws['A9'].alignment = label_align_left
        ws['A9'].border = Border(top=thin_side, left=Side(style='medium'))
        ws.merge_cells('A9:B9')
        ws.merge_cells('C9:D9')
        ws['E9'] = '贸易国(地区)'
        ws['E9'].font = label_font
        ws['E9'].alignment = label_align
        ws['E9'].border = Border(top=thin_side, left=thin_side)
        ws['F9'] = '(  USA  )'
        ws['F9'].font = label_font
        ws['F9'].alignment = label_align
        ws['F9'].border = Border(top=thin_side, right=thin_side)
        ws['G9'] = '运抵国（地区）'
        ws['G9'].font = label_font
        ws['G9'].alignment = label_align
        ws['G9'].border = Border(top=thin_side, left=thin_side)
        ws['H9'] = '(  USA  )'
        ws['H9'].font = label_font
        ws['H9'].alignment = label_align_left
        ws.merge_cells('H9:J9')
        ws['K9'] = '指运港'
        ws['K9'].font = label_font
        ws['K9'].alignment = label_align
        ws['K9'].border = Border(top=thin_side, left=thin_side)
        ws['L9'] = '(    )'
        ws['L9'].font = label_font
        ws['L9'].alignment = label_align_left
        ws['L9'].border = Border(top=thin_side, right=thin_side)
        ws.merge_cells('L9:N9')
        ws['O9'] = '离境口岸'
        ws['O9'].font = label_font_red
        ws['O9'].alignment = label_align
        ws['O9'].border = Border(top=thin_side, left=thin_side)
        ws['P9'] = '(    )'
        ws['P9'].font = label_font
        ws['P9'].alignment = label_align_left
        ws['P9'].border = Border(top=thin_side, right=Side(style='medium'))
        ws.merge_cells('P9:S9')

        # Row 10: values
        ws['A10'] = cno
        ws['A10'].font = value_font
        ws['A10'].alignment = value_align_left
        ws['A10'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
        ws.merge_cells('A10:D10')
        ws['E10'] = '美国'
        ws['E10'].font = value_font_red
        ws['E10'].alignment = value_align_left
        ws['E10'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('E10:F10')
        ws['G10'] = '美国'
        ws['G10'].font = Font(name='宋体', size=11)
        ws['G10'].alignment = value_align_left
        ws['G10'].border = Border(bottom=thin_side, left=thin_side)
        ws.merge_cells('G10:J10')
        ws['K10'] = '美国'
        ws['K10'].font = Font(name='宋体', size=11)
        ws['K10'].alignment = value_align_left
        ws['K10'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('K10:N10')
        ws['O10'] = ''
        ws['O10'].border = Border(bottom=thin_side)
        ws.merge_cells('O10:S10')

        # Row 11-12: 包装/件数/毛重/运费
        ws['A11'] = '包装种类'
        ws['A11'].font = label_font
        ws['A11'].alignment = label_align_left
        ws['A11'].border = Border(top=thin_side, left=Side(style='medium'))
        ws.merge_cells('A11:B11')
        ws['C11'] = '( 22/06   )'
        ws['C11'].font = label_font
        ws['C11'].alignment = label_align_left
        ws['C11'].border = Border(top=thin_side, right=thin_side)
        ws.merge_cells('C11:D11')

        ws['E11'] = '件数'
        ws['E11'].font = label_font
        ws['E11'].alignment = label_align
        ws['E11'].border = Border(left=thin_side, right=thin_side)

        ws['F11'] = '毛重（千克）'
        ws['F11'].font = label_font
        ws['F11'].alignment = label_align
        ws['F11'].border = Border(top=thin_side, left=thin_side, right=thin_side)

        ws['G11'] = '净重（千克）'
        ws['G11'].font = label_font
        ws['G11'].alignment = Alignment(horizontal='left', vertical='center')
        ws['G11'].border = Border(top=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('G11:H11')

        ws['I11'] = '成交方式'
        ws['I11'].font = label_font
        ws['I11'].alignment = Alignment(horizontal='right', vertical='center')
        ws['I11'].border = Border(top=thin_side, left=thin_side)

        ws['J11'] = '(  )'
        ws['J11'].font = label_font
        ws['J11'].alignment = label_align
        ws['J11'].border = Border(top=thin_side)

        ws['K11'] = '运费'
        ws['K11'].font = label_font
        ws['K11'].alignment = label_align
        ws['K11'].border = Border(top=thin_side, left=thin_side)
        ws.merge_cells('L11:M11')

        ws['N11'] = '保费'
        ws['N11'].font = label_font
        ws['N11'].alignment = label_align
        ws['N11'].border = Border(top=thin_side)
        ws.merge_cells('O11:P11')

        ws['Q11'] = '杂费'
        ws['Q11'].font = label_font
        ws['Q11'].alignment = label_align
        ws['Q11'].border = Border(top=thin_side, left=thin_side)
        ws.merge_cells('R11:S11')

        # Compute PL totals for the header
        tot_boxes = 0
        tot_nw = 0.0
        tot_gw = 0.0
        for item in self.items:
            sku = _sku_key(item)
            qty = tq.get(sku, 0)
            if qty > 0:
                pr = item.get('packing_rate', 1) or 1
                bx = qty / pr
                tot_boxes += int(bx)
                tot_nw += item.get('net_weight_kg', 0) * bx
                tot_gw += item.get('gross_weight_kg', 0) * bx

        ws['A12'] = '纸制或纤维板制盒/箱/包/袋'
        ws['A12'].font = value_font
        ws['A12'].alignment = value_align_left
        ws['A12'].border = Border(bottom=thin_side, left=Side(style='medium'), right=thin_side)
        ws.merge_cells('A12:D12')

        ws['E12'] = tot_boxes
        ws['E12'].font = value_font
        ws['E12'].alignment = value_align_left
        ws['E12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)

        ws['F12'] = round(tot_gw, 1)
        ws['F12'].font = value_font
        ws['F12'].alignment = value_align_left
        ws['F12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)

        ws['G12'] = round(tot_nw, 1)
        ws['G12'].font = value_font
        ws['G12'].alignment = value_align_left
        ws['G12'].border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        ws.merge_cells('G12:H12')

        ws['I12'] = self.price_term
        ws['I12'].font = value_font
        ws['I12'].alignment = value_align_left
        ws['I12'].border = Border(bottom=thin_side, left=thin_side)
        ws.merge_cells('I12:J12')

        ws['K12'] = 'USD'
        ws['K12'].font = value_font
        ws['K12'].alignment = label_align
        ws['K12'].border = Border(bottom=thin_side, left=thin_side)

        ws['L12'] = round(total_ship / self.rate, 2)
        ws['L12'].font = value_font
        ws['L12'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L12'].border = Border(bottom=thin_side, right=thin_side)
        ws.merge_cells('L12:M12')

        ws['N12'] = None
        ws['N12'].font = value_font
        ws['N12'].alignment = value_align_left
        ws['N12'].border = Border(bottom=thin_side, right=thin_side)
        ws.merge_cells('N12:P12')

        ws['Q12'] = None
        ws['Q12'].font = value_font
        ws['Q12'].alignment = value_align_left
        ws['Q12'].border = Border(bottom=thin_side, left=thin_side, right=Side(style='medium'))
        ws.merge_cells('Q12:S12')

        # Row 13: 随附单证及编号
        ws['A13'] = '随附单证及编号'
        ws['A13'].font = label_font
        ws['A13'].alignment = label_align
        ws['A13'].border = Border(top=thin_side, left=Side(style='medium'))
        ws.merge_cells('C13:S13')

        # Row 14
        ws.merge_cells('A14:C14')
        ws.merge_cells('D14:S14')

        # Row 15
        ws['A15'] = '标记唛码及备注：'
        ws['A15'].font = label_font
        ws['A15'].alignment = label_align
        ws['A15'].border = Border(left=Side(style='medium'))
        ws.merge_cells('C15:S15')

        # Row 16
        ws['A16'] = '备注：'
        ws['A16'].font = label_font
        ws['A16'].alignment = label_align
        ws['A16'].border = Border(left=Side(style='medium'))
        ws.merge_cells('B16:S16')

        # Row 17: section header (项目列表区域标题)
        ws.merge_cells('A17:K17')
        ws.merge_cells('L17:N17')
        ws.merge_cells('O17:S17')

        # Row 18: spacer row with specific formatting per cell (matching example)
        med_tb = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        med_ltb = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        med_lrtb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
        med_ltb_thin = Border(left=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
        med_rtb = Border(right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium'))
        med_rtb_med = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        ws['A18'] = ' '
        ws['A18'].font = Font(name='宋体', size=11, bold=True)
        ws['A18'].alignment = label_align
        ws['A18'].border = med_ltb

        ws['B18'] = ' '
        ws['B18'].font = Font(name='宋体', size=11, bold=True)
        ws['B18'].alignment = label_align
        ws['B18'].border = med_tb

        ws['C18'] = None
        ws['C18'].font = Font(name='宋体', size=11, bold=True)
        ws['C18'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C18'].border = med_lrtb

        ws['D18'] = ' '
        ws['D18'].font = Font(name='宋体', size=11, bold=True)
        ws['D18'].alignment = Alignment(horizontal='right', vertical='center')
        ws['D18'].border = med_ltb_thin

        ws['E18'] = None
        ws['E18'].font = Font(name='宋体', size=11, bold=True)
        ws['E18'].alignment = Alignment(horizontal='left', vertical='center')
        ws['E18'].border = med_lrtb

        ws['F18'] = ' '
        ws['F18'].font = Font(name='宋体', size=11, bold=True)
        ws['F18'].alignment = Alignment(horizontal='right', vertical='center')
        ws['F18'].border = med_tb
        ws.merge_cells('F18:G18')

        ws['H18'] = None
        ws['H18'].font = Font(name='宋体', size=11, bold=True)
        ws['H18'].alignment = Alignment(horizontal='left', vertical='center')
        ws['H18'].border = med_lrtb

        ws['I18'] = ' '
        ws['I18'].font = Font(name='宋体', size=11, bold=True)
        ws['I18'].alignment = Alignment(horizontal='right', vertical='center')
        ws['I18'].border = med_ltb_thin
        ws.merge_cells('I18:J18')

        ws['K18'] = ' '
        ws['K18'].font = Font(name='宋体', size=10, bold=True)
        ws['K18'].alignment = label_align
        ws['K18'].border = med_rtb

        ws['L18'] = ' '
        ws['L18'].font = Font(name='宋体', size=11)
        ws['L18'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L18'].border = med_rtb_med
        ws.merge_cells('L18:S18')

        # Row 19: Column headers with merges
        ws.row_dimensions[19].height = 12.0
        thin_bottom = Border(bottom=Side(style='thin'))
        hdr_font = Font(name='宋体', size=10)
        hdr_align = Alignment(horizontal='center', vertical='center')

        ws.cell(row=19, column=1, value='项号').font = hdr_font
        ws.cell(row=19, column=1).alignment = hdr_align
        ws.cell(row=19, column=1).border = Border(bottom=Side(style='thin'), left=Side(style='thin'))

        ws.cell(row=19, column=2, value='商品编号').font = Font(name='宋体', size=10, color='FF0000')
        ws.cell(row=19, column=2).alignment = hdr_align
        ws.cell(row=19, column=2).border = thin_bottom
        ws.merge_cells('B19:C19')

        ws.cell(row=19, column=4, value='商品名称及规格型号').font = hdr_font
        ws.cell(row=19, column=4).alignment = hdr_align
        ws.cell(row=19, column=4).border = thin_bottom
        ws.merge_cells('D19:F19')

        ws.cell(row=19, column=7, value='数量及单位').font = hdr_font
        ws.cell(row=19, column=7).alignment = hdr_align
        ws.cell(row=19, column=7).border = thin_bottom
        ws.merge_cells('G19:H19')

        ws.cell(row=19, column=9, value='单价/总价/币制').font = Font(name='宋体', size=10, color='FF0000')
        ws.cell(row=19, column=9).alignment = hdr_align
        ws.cell(row=19, column=9).border = thin_bottom
        ws.merge_cells('I19:J19')

        ws.cell(row=19, column=11, value='原产国（地区）').font = hdr_font
        ws.cell(row=19, column=11).alignment = hdr_align
        ws.cell(row=19, column=11).border = thin_bottom
        ws.merge_cells('K19:L19')

        ws.cell(row=19, column=13, value='最终目的国（地区）').font = hdr_font
        ws.cell(row=19, column=13).alignment = hdr_align
        ws.cell(row=19, column=13).border = thin_bottom
        ws.merge_cells('M19:O19')

        ws.cell(row=19, column=16, value='境内货源地').font = Font(name='宋体', size=10, color='FF0000')
        ws.cell(row=19, column=16).alignment = hdr_align
        ws.cell(row=19, column=16).border = thin_bottom
        ws.merge_cells('P19:R19')

        ws.cell(row=19, column=19, value='征免').font = hdr_font
        ws.cell(row=19, column=19).alignment = hdr_align
        ws.cell(row=19, column=19).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

        # Items: 3 rows per SKU
        row = 20
        item_no = 1
        supplier_city = self.contract.get('supplier', {}).get('city', '义乌')

        item_font = Font(name='宋体', size=11)
        item_font_12 = Font(name='宋体', size=12)

        for item in self.items:
            sku = _sku_key(item)
            qty = tq.get(sku, 0)
            if qty == 0:
                continue

            amt = ta.get(sku, 0)
            shipping = ship_alloc.get(sku, 0)
            tax_excl = amt / 1.13
            cnf_total_rmb = tax_excl + shipping
            unit_usd = cnf_total_rmb / qty / self.rate if qty > 0 else 0
            total_usd = cnf_total_rmb / self.rate

            info = self._info(sku, item)
            nw_kg = item.get('net_weight_kg', 0) * (qty / (item.get('packing_rate', 1) or 1))

            # Set row heights: row1=14.25, row2&row3=default
            ws.row_dimensions[row].height = 14.25

            # Row 1: main info with merges
            c1 = ws.cell(row=row, column=1, value=item_no)
            c1.font = item_font
            c1.alignment = Alignment(horizontal='center', vertical='center')
            c1.border = Border(left=Side(style='thin'))

            c2 = ws.cell(row=row, column=2, value=info['tariff_code'] or 3918909000)
            c2.font = item_font
            c2.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'B{row}:C{row}')

            c4 = ws.cell(row=row, column=4, value=item.get('name_cn', ''))
            c4.font = item_font
            c4.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'D{row}:F{row}')

            c7 = ws.cell(row=row, column=7, value=qty)
            c7.font = item_font
            c7.alignment = Alignment(vertical='center')

            ws.cell(row=row, column=8, value='个').font = item_font
            ws.cell(row=row, column=8).alignment = Alignment(vertical='center')

            c9 = ws.cell(row=row, column=9, value=round(unit_usd, 6))
            c9.font = item_font
            c9.alignment = Alignment(horizontal='right', vertical='center')
            ws.merge_cells(f'I{row}:J{row}')

            c11 = ws.cell(row=row, column=11, value='中国')
            c11.font = item_font
            c11.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'K{row}:L{row}')

            c13 = ws.cell(row=row, column=13, value='美国')
            c13.font = item_font_12
            c13.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'M{row}:O{row}')

            c16 = ws.cell(row=row, column=16, value=f'{supplier_city}(33189)')
            c16.font = item_font_12
            c16.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'P{row}:R{row}')

            c19 = ws.cell(row=row, column=19, value='照章征税')
            c19.font = item_font
            c19.alignment = Alignment(vertical='center')
            c19.border = Border(right=Side(style='thin'))

            # Row 2: declaration elements + net weight + total price
            row += 1
            ws.merge_cells(f'B{row}:C{row}')
            c4_r2 = ws.cell(row=row, column=4, value=info['declaration_elements'])
            c4_r2.font = Font(name='宋体', size=10)
            c4_r2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c4_r2.border = Border(bottom=Side(style='thin'))
            ws.merge_cells(f'D{row}:F{row+1}')

            c7_r2 = ws.cell(row=row, column=7, value=round(nw_kg, 1))
            c7_r2.font = item_font
            c7_r2.alignment = Alignment(vertical='center')

            ws.cell(row=row, column=8, value='千克').font = item_font
            ws.cell(row=row, column=8).alignment = Alignment(vertical='center')

            c9_r2 = ws.cell(row=row, column=9, value=round(total_usd, 2))
            c9_r2.font = item_font
            c9_r2.alignment = Alignment(horizontal='right', vertical='center')
            ws.merge_cells(f'I{row}:J{row}')

            ws.cell(row=row, column=11, value='（CHN）').font = item_font
            ws.cell(row=row, column=11).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'K{row}:L{row}')
            ws.cell(row=row, column=13, value='（USA）').font = item_font
            ws.cell(row=row, column=13).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'M{row}:O{row}')
            ws.merge_cells(f'P{row}:R{row}')

            # Row 3: qty in 个 + currency (with bottom borders)
            row += 1
            ws.merge_cells(f'B{row}:C{row}')

            c7_r3 = ws.cell(row=row, column=7, value=qty)
            c7_r3.font = item_font
            c7_r3.alignment = Alignment(vertical='center')
            c7_r3.border = Border(bottom=Side(style='thin'))

            ws.cell(row=row, column=8, value='个').font = item_font
            ws.cell(row=row, column=8).alignment = Alignment(vertical='center')
            ws.cell(row=row, column=8).border = Border(bottom=Side(style='thin'))

            c9_r3 = ws.cell(row=row, column=9, value='美元')
            c9_r3.font = item_font
            c9_r3.alignment = Alignment(horizontal='right', vertical='center')
            c9_r3.border = Border(bottom=Side(style='thin'))
            ws.merge_cells(f'I{row}:J{row}')

            ws.merge_cells(f'K{row}:L{row}')
            ws.merge_cells(f'M{row}:O{row}')
            ws.merge_cells(f'P{row}:R{row}')

            row += 1
            item_no += 1

        fn = f'{suffix}出口报关单草稿.xlsx' if suffix else '出口报关单草稿.xlsx'
        fp = os.path.join(self.out_dir, fn)
        wb.save(fp)
        return fp


# ─── CLI ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Generate customs declaration documents')
    parser.add_argument('--contract', required=True, help='Path to purchase_contract.json')
    parser.add_argument('--shipments', required=True, help='Path to fba_shipments.json')
    parser.add_argument('--knowledge-base', help='Path to knowledge base Excel')
    parser.add_argument('--groups', required=True, help='JSON string defining ticket groups')
    parser.add_argument('--selected-groups', required=True, help='Comma-separated group indices')
    parser.add_argument('--exchange-rate', type=float, required=True)
    parser.add_argument('--shipping-rate', type=float, required=True, help='RMB per kg')
    parser.add_argument('--output-dir', required=True)
    parser.add_argument('--template-dir', help='(reserved for future)')
    parser.add_argument('--price-term', default='CNF', choices=['CNF', 'FOB'])

    args = parser.parse_args()

    groups = json.loads(args.groups)
    selected = [int(x.strip()) for x in args.selected_groups.split(',')]

    gen = CustomsDeclarationGenerator(
        contract_json=args.contract,
        shipments_json=args.shipments,
        groups=groups,
        selected_group_indices=selected,
        exchange_rate=args.exchange_rate,
        shipping_rate=args.shipping_rate,
        output_dir=args.output_dir,
        price_term=args.price_term,
        knowledge_base=args.knowledge_base,
        template_dir=args.template_dir,
    )

    print(gen.generate())


if __name__ == '__main__':
    main()
