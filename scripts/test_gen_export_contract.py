#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script for gen_export_contract.py — generates a sample 出货合同 (export contract)
and verifies the output Excel structure and styling.
"""

import os
import sys
import tempfile

# Ensure the scripts directory is on the import path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from gen_export_contract import gen_export_contract
from helpers import sku_key, get_lwh

# ─── Sample data (modeled after PO2603230466) ────────────────────────────────

SAMPLE_ITEMS = [
    {
        "name_cn": "人造草坪拼接地板",
        "name_en": "Artificial Grass Interlocking Floor Tiles",
        "spec": "30*30cm 9pcs",
        "fba_sku": "AG-TILE-3030-9P",
        "unit": "件",
        "quantity": 3600,
        "packing_rate": 12,
        "unit_price_with_tax": 35.50,
        "package_size_cm": [43, 45, 54],
        "net_weight_kg": 8.5,
        "gross_weight_kg": 9.2,
        "total_amount": 127800.00,
    },
    {
        "name_cn": "塑料草坪围栏",
        "name_en": "Plastic Grass Fence Panel",
        "spec": "60*40cm",
        "fba_sku": "PG-FENCE-6040",
        "unit": "件",
        "quantity": 2400,
        "packing_rate": 8,
        "unit_price_with_tax": 28.00,
        "package_size_cm": [62, 42, 35],
        "net_weight_kg": 7.0,
        "gross_weight_kg": 7.8,
        "total_amount": 67200.00,
    },
    {
        "name_cn": "仿真植物墙装饰",
        "name_en": "Artificial Plant Wall Decor",
        "spec": "40*60cm",
        "fba_sku": "AP-WALL-4060",
        "unit": "件",
        "quantity": 1800,
        "packing_rate": 6,
        "unit_price_with_tax": 42.00,
        "package_size_cm": [65, 45, 30],
        "net_weight_kg": 6.5,
        "gross_weight_kg": 7.2,
        "total_amount": 75600.00,
    },
]

SAMPLE_CONTRACT = {
    "contract_no": "PO2603230466",
    "date": "2026-03-23",
    "supplier": {
        "name": "义乌市绿美工艺品有限公司",
        "city": "义乌",
        "address": "",
        "contact": "",
        "phone": "",
    },
    "buyer": {
        "name": "深圳市艾进贸易有限公司",
        "address": "",
        "contact": "",
        "phone": "",
    },
    "grand_total": 270600.00,
}

SAMPLE_KB = {
    "AG-TILE-3030-9P": {
        "tariff_code": "3918909000",
        "english_name": "Artificial Grass Interlocking Floor Tiles",
        "declaration_elements": "0|0|塑料|人造草坪拼接地板|无品牌|无型号",
        "material": "plastic",
    },
    "PG-FENCE-6040": {
        "tariff_code": "3926909090",
        "english_name": "Plastic Grass Fence Panel",
        "declaration_elements": "0|0|塑料|塑料草坪围栏|无品牌|无型号",
        "material": "plastic",
    },
    "AP-WALL-4060": {
        "tariff_code": "6702100000",
        "english_name": "Artificial Plant Wall Decor",
        "declaration_elements": "0|0|塑料|仿真植物墙|无品牌|无型号",
        "material": "plastic",
    },
}

EXCHANGE_RATE = 7.25
SHIPPING_RATE = 8.5

TQ = {sku_key(item): item["quantity"] for item in SAMPLE_ITEMS}
TA = {sku_key(item): item["total_amount"] for item in SAMPLE_ITEMS}


def _compute_shipping(items, ship_rate):
    total_gross = 0.0
    total_vol = 0.0
    for item in items:
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        total_gross += item["gross_weight_kg"] * boxes
        total_vol += (l * w * h / 6000) * boxes
    chargeable = max(total_gross, total_vol)
    total_ship = chargeable * ship_rate
    use_vol = total_vol > total_gross
    alloc = {}
    for item in items:
        sku = sku_key(item)
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        sku_vol = (l * w * h / 6000) * boxes
        sku_gross = item["gross_weight_kg"] * boxes
        if use_vol:
            prop = sku_vol / total_vol if total_vol > 0 else 0
        else:
            prop = sku_gross / total_gross if total_gross > 0 else 0
        alloc[sku] = total_ship * prop
    return alloc, total_gross, total_vol, chargeable, total_ship


SHIP_ALLOC, TOTAL_GROSS, TOTAL_VOL, CHARGEABLE, TOTAL_SHIP = _compute_shipping(SAMPLE_ITEMS, SHIPPING_RATE)


# ─── Test ─────────────────────────────────────────────────────────────────────

def test_gen_export_contract():
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = gen_export_contract(
            items=SAMPLE_ITEMS,
            contract=SAMPLE_CONTRACT,
            kb=SAMPLE_KB,
            cno="PO2603230466",
            suffix="",
            tq=TQ,
            ta=TA,
            ship_alloc=SHIP_ALLOC,
            total_gross=TOTAL_GROSS,
            total_vol=TOTAL_VOL,
            chargeable=CHARGEABLE,
            total_ship=TOTAL_SHIP,
            rate=EXCHANGE_RATE,
            ship_rate=SHIPPING_RATE,
            out_dir=tmpdir,
        )

        # 1) File exists
        assert os.path.exists(fp), f"Output file not created: {fp}"
        print(f"[PASS] File created: {os.path.basename(fp)}")

        # 2) Load workbook
        wb = load_workbook(fp)
        ws = wb.active
        assert ws.title == "出口合同", f"Unexpected sheet title: {ws.title}"
        print(f"[PASS] Sheet title: {ws.title}")

        # 3) Title: '出货合同'
        assert ws["A1"].value == "出货合同"
        assert ws["A1"].font.size == 24
        assert ws["A1"].font.name == "宋体"
        assert ws["A1"].alignment.horizontal == "center"
        print("[PASS] Title: '出货合同', font 宋体 size 24, centered")

        # 4) Contract number
        assert ws["I2"].value == "PO2603230466"
        print(f"[PASS] Contract number: {ws['I2'].value}")

        # 5) Date
        assert ws["D3"].value == "2026-03-23"
        print(f"[PASS] Date: {ws['D3'].value}")

        # 6) Supplier and buyer
        assert ws["D4"].value == "义乌市绿美工艺品有限公司"
        assert ws["I4"].value == "深圳市艾进贸易有限公司"
        print(f"[PASS] Supplier: {ws['D4'].value}")
        print(f"[PASS] Buyer: {ws['I4'].value}")

        # 6b) Header rows: 地址/联系人/电话
        assert ws["A5"].value == "地址："
        assert ws["H5"].value == "地址："
        assert ws["A6"].value == "联系人："
        assert ws["A7"].value == "电话："
        print("[PASS] Header section: 地址/联系人/电话 rows present")

        # 6c) Section label
        assert ws["A9"].value == "一、项目名称、规格型号、数量、金额"
        print("[PASS] Section label at A9")

        # 7) Column headers at row 11
        expected_headers = ['产品名称', '产品图片', '规格型号', 'FBA SKU', '单位', '数量',
                           '箱率', '含税单价/元', '包装尺寸/CM', '外箱净重/KG', '外箱毛重/KG', '总额/元']
        for c, h in enumerate(expected_headers, 1):
            actual = ws.cell(row=11, column=c).value
            assert actual == h, f"Header col {c} mismatch: {actual} != {h}"
        print("[PASS] All 12 column headers correct at row 11")

        # 7b) Header borders check
        h1 = ws.cell(row=11, column=1)
        assert h1.border.top.style == 'medium', f"Header row top border should be medium, got {h1.border.top.style}"
        assert h1.border.bottom.style == 'medium', f"Header row bottom border should be medium, got {h1.border.bottom.style}"
        assert h1.border.left.style == 'medium', f"Header col A left border should be medium, got {h1.border.left.style}"
        h12 = ws.cell(row=11, column=12)
        assert h12.border.right.style == 'medium', f"Header col L right border should be medium, got {h12.border.right.style}"
        print("[PASS] Header row borders: medium top/bottom, medium left/right edges")

        # 7c) Right-side calc headers
        assert ws.cell(row=11, column=15).value == '实重'
        assert ws.cell(row=11, column=16).value == '体积重'
        assert ws.cell(row=11, column=17).value == '海运费平摊'
        assert ws.cell(row=11, column=18).value == 'C&F总价'
        assert ws.cell(row=11, column=19).value == 'C&F单价'
        print("[PASS] Right-side calc headers correct")

        # 8) Item data rows starting at row 12
        row = 12
        sum_qty = 0
        sum_amt = 0.0
        for idx, item in enumerate(SAMPLE_ITEMS):
            sku = sku_key(item)
            qty = TQ[sku]
            amt = TA[sku]

            # Column A: name_cn + name_en
            name_val = ws.cell(row=row, column=1).value
            assert item["name_cn"] in str(name_val), f"Item {idx+1}: name_cn not found"
            assert item["name_en"] in str(name_val), f"Item {idx+1}: name_en not found"

            # Column C: spec
            assert ws.cell(row=row, column=3).value == item["spec"]

            # Column D: SKU
            assert ws.cell(row=row, column=4).value == sku

            # Column E: unit
            assert ws.cell(row=row, column=5).value == item["unit"]

            # Column F: quantity
            assert ws.cell(row=row, column=6).value == qty

            # Column G: packing rate
            assert ws.cell(row=row, column=7).value == item["packing_rate"]

            # Column H: unit price
            unit_price = ws.cell(row=row, column=8).value
            expected_up = round(amt / qty, 2)
            assert abs(unit_price - expected_up) < 0.01, f"Item {idx+1}: unit price {unit_price} != {expected_up}"

            # Column I: package size
            l, w, h = item["package_size_cm"]
            expected_size = f"{int(l)}*{int(w)}*{int(h)}"
            assert ws.cell(row=row, column=9).value == expected_size

            # Column J: net weight
            assert ws.cell(row=row, column=10).value == item["net_weight_kg"]

            # Column K: gross weight
            assert ws.cell(row=row, column=11).value == item["gross_weight_kg"]

            # Column L: total amount
            total_amt = ws.cell(row=row, column=12).value
            assert abs(total_amt - round(amt, 2)) < 0.01

            # Data row borders
            d1 = ws.cell(row=row, column=1)
            assert d1.border.top.style == 'medium', f"Data row {row} col A top border should be medium"
            assert d1.border.bottom.style == 'thin', f"Data row {row} col A bottom border should be thin"
            assert d1.border.left.style == 'medium', f"Data row {row} col A left border should be medium"
            d12 = ws.cell(row=row, column=12)
            assert d12.border.right.style == 'medium', f"Data row {row} col L right border should be medium"

            # Alignment: center/center with wrap
            assert d1.alignment.horizontal == 'center'
            assert d1.alignment.vertical == 'center'
            assert d1.alignment.wrap_text is True

            # Right-side calc columns
            real_w = ws.cell(row=row, column=15).value
            assert real_w > 0, f"Item {idx+1}: real weight should be > 0"
            vol_w = ws.cell(row=row, column=16).value
            assert vol_w > 0, f"Item {idx+1}: volume weight should be > 0"
            ship_val = ws.cell(row=row, column=17).value
            assert ship_val > 0, f"Item {idx+1}: shipping alloc should be > 0"
            cnf_total = ws.cell(row=row, column=18).value
            assert cnf_total > 0, f"Item {idx+1}: C&F total should be > 0"
            cnf_unit = ws.cell(row=row, column=19).value
            assert cnf_unit > 0, f"Item {idx+1}: C&F unit price should be > 0"

            print(f"[PASS] Item {idx+1} ({sku}): all fields + borders + alignment correct, C&F unit={cnf_unit:.2f}")

            sum_qty += qty
            sum_amt += amt
            row += 1

        # 9) Totals row — "小写合计"
        assert ws.cell(row=row, column=1).value == "小写合计"
        # Check border on 小写合计
        t1 = ws.cell(row=row, column=1)
        assert t1.border.top.style == 'medium'
        assert t1.border.left.style == 'medium'
        print(f"[PASS] 小写合计 row with borders")

        # Right-side: 总重
        assert ws.cell(row=row, column=14).value == '总重'
        assert ws.cell(row=row, column=15).value == round(TOTAL_GROSS, 1)
        print("[PASS] 总重 label and value")

        # 10) "大写合计" row
        row += 1
        assert ws.cell(row=row, column=1).value == "大写合计"
        # Right-side: 总海运费
        assert ws.cell(row=row, column=14).value == '总海运费'
        assert ws.cell(row=row, column=15).value == round(TOTAL_SHIP, 2)
        print("[PASS] 大写合计 + 总海运费")

        # 11) Empty bordered row
        row += 1
        for ci in range(1, 13):
            cell = ws.cell(row=row, column=ci)
            assert cell.border.top.style == 'medium', f"Empty row col {ci} top border should be medium"
        print("[PASS] Empty bordered row")

        # 12) "共计" row
        row += 1
        assert ws.cell(row=row, column=1).value == "共计"
        grand_total = ws.cell(row=row, column=12).value
        assert abs(grand_total - round(sum_amt, 2)) < 0.01
        # Bottom border should be medium
        g12 = ws.cell(row=row, column=12)
        assert g12.border.bottom.style == 'medium'
        assert g12.border.right.style == 'medium'
        print(f"[PASS] 共计: {grand_total}, borders correct")

        # 13) Row heights
        assert ws.row_dimensions[1].height == 42
        assert ws.row_dimensions[5].height == 27
        print("[PASS] Row heights correct")

        # 14) Contract terms text
        # Find the first terms row (should be 2 rows after 共计)
        terms_row = row + 2
        term1 = ws.cell(row=terms_row, column=1).value
        assert term1 is not None and '交期' in str(term1), f"Expected contract term at row {terms_row}, got: {term1}"
        print(f"[PASS] Contract terms start at row {terms_row}")

        # Check a few key terms exist
        found_terms = set()
        for r in range(terms_row, terms_row + 40):
            v = ws.cell(row=r, column=1).value
            if v:
                if '交期' in str(v): found_terms.add('交期')
                if '账期' in str(v): found_terms.add('账期')
                if '交货地点' in str(v): found_terms.add('交货地点')
                if '保密协议' in str(v): found_terms.add('保密协议')
                if '验货标准' in str(v): found_terms.add('验货标准')
                if '包装标准' in str(v): found_terms.add('包装标准')
                if '售后保障' in str(v): found_terms.add('售后保障')
                if '纠纷' in str(v): found_terms.add('纠纷')
                if '开票' in str(v): found_terms.add('开票')
                if '壹式贰份' in str(v): found_terms.add('壹式贰份')
                if '供方：' in str(v): found_terms.add('供方签章')
                if '盖章：' in str(v): found_terms.add('盖章')

        expected_terms = {'交期', '账期', '交货地点', '保密协议', '验货标准', '包装标准',
                         '售后保障', '纠纷', '开票', '壹式贰份', '供方签章', '盖章'}
        missing = expected_terms - found_terms
        assert not missing, f"Missing contract terms: {missing}"
        print(f"[PASS] All {len(expected_terms)} contract terms/signatures found")

        # 15) Font checks
        assert ws.cell(row=12, column=1).font.name == "宋体"
        assert ws.cell(row=12, column=1).font.size == 10
        print("[PASS] Data font: 宋体 size 10")

        print("\n" + "=" * 60)
        print("ALL TESTS PASSED for gen_export_contract")
        print("=" * 60)


if __name__ == "__main__":
    test_gen_export_contract()
