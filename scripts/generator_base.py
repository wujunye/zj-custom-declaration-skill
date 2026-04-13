#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Core CustomsDeclarationGenerator class with shared computation logic.
Delegates document generation to separate modules.
"""

import json
import os
import sys
from typing import Dict, List, Any, Tuple, Optional

from helpers import get_lwh, sku_key, build_sku_mapping
from gen_export_contract import gen_export_contract
from gen_iv_pl import gen_iv_pl
from gen_declaration import gen_declaration


class CustomsDeclarationGenerator:
    def __init__(
        self,
        contract_json: str,
        shipments_json: str,
        groups: List[Dict[str, Any]],
        selected_group_indices: List[int],
        exchange_rate: float,
        shipping_rate: float,
        output_dir: str,
        price_term: str = "CNF",
        knowledge_base: Optional[str] = None,
        template_dir: Optional[str] = None,
    ):
        self.contract = self._load_json(contract_json)
        self.shipments = self._load_json(shipments_json)
        self.groups = groups
        self.selected = selected_group_indices
        self.rate = exchange_rate
        self.ship_rate = shipping_rate
        self.out_dir = output_dir
        self.price_term = price_term
        self.kb = self._load_kb(knowledge_base) if knowledge_base else {}
        self.template_dir = template_dir

        os.makedirs(output_dir, exist_ok=True)

        # Items shorthand
        self.items = self.contract.get('items', [])
        # FBA matrix: {sku: {warehouse: qty}}
        self.matrix = self.shipments.get('matrix', {})
        # SKU mapping: contract_sku → matrix_sku
        self.sku_map = build_sku_mapping(self.items, self.matrix)

    @staticmethod
    def _load_json(path: str) -> dict:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_kb(self, path: str) -> Dict[str, dict]:
        """Load knowledge base Excel (SKU → tariff/name/elements/material)."""
        from openpyxl import load_workbook
        kb = {}
        try:
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                kb[str(row[0])] = {
                    'tariff_code': row[1] if len(row) > 1 and row[1] else '',
                    'english_name': row[2] if len(row) > 2 and row[2] else '',
                    'declaration_elements': row[3] if len(row) > 3 and row[3] else '0|0|塑料|塑料草坪|无品牌|无型号',
                    'material': row[4] if len(row) > 4 and row[4] else 'plastic',
                }
        except Exception as e:
            print(f"Warning: KB load failed: {e}", file=sys.stderr)
        return kb

    # ─── quantity computation ──────────────────────────────────────────

    def _group_qty(self, group_idx: int) -> Dict[str, int]:
        """Compute {contract_sku: qty} for a ticket group from the FBA matrix."""
        whs = self.groups[group_idx]['warehouses']
        result = {}
        for item in self.items:
            c_sku = sku_key(item)
            m_sku = self.sku_map.get(c_sku, c_sku)
            wh_data = self.matrix.get(m_sku, {})
            total = sum(wh_data.get(w, 0) for w in whs)
            if total > 0:
                result[c_sku] = total
        return result

    # ─── chargeable weight & shipping ──────────────────────────────────

    def _chargeable(self) -> Tuple[float, float, float, float]:
        """
        Compute chargeable weight from ALL contract items (full quantities).
        Returns (total_gross, total_vol, chargeable_weight, total_shipping_rmb).
        """
        total_gross = 0.0
        total_vol = 0.0
        for item in self.items:
            qty = item['quantity']
            pr = item.get('packing_rate', 1) or 1
            boxes = qty / pr
            l, w, h = get_lwh(item)
            total_gross += item.get('gross_weight_kg', 0) * boxes
            total_vol += (l * w * h / 6000) * boxes
        chargeable = max(total_gross, total_vol)
        return total_gross, total_vol, chargeable, chargeable * self.ship_rate

    def _shipping_alloc(self, total_ship: float, total_vol: float, total_gross: float) -> Dict[str, float]:
        """Allocate shipping to each contract SKU by proportion."""
        use_vol = total_vol > total_gross
        alloc = {}
        for item in self.items:
            sku = sku_key(item)
            qty = item['quantity']
            pr = item.get('packing_rate', 1) or 1
            boxes = qty / pr
            l, w, h = get_lwh(item)
            sku_vol = (l * w * h / 6000) * boxes
            sku_gross = item.get('gross_weight_kg', 0) * boxes
            if use_vol:
                prop = sku_vol / total_vol if total_vol > 0 else 0
            else:
                prop = sku_gross / total_gross if total_gross > 0 else 0
            alloc[sku] = total_ship * prop
        return alloc

    # ─── amount allocation across tickets ──────────────────────────────

    def _amount_alloc(self) -> Dict[int, Dict[str, float]]:
        """
        Allocate purchase amounts across selected tickets proportionally.
        {group_idx: {contract_sku: allocated_rmb_amount}}
        """
        totals = {}
        group_qtys = {}
        for gi in self.selected:
            gq = self._group_qty(gi)
            group_qtys[gi] = gq
            for sku, q in gq.items():
                totals[sku] = totals.get(sku, 0) + q

        result = {}
        for gi in self.selected:
            result[gi] = {}
            for item in self.items:
                sku = sku_key(item)
                tq = totals.get(sku, 0)
                gq = group_qtys[gi].get(sku, 0)
                if tq > 0 and gq > 0:
                    result[gi][sku] = item['total_amount'] * (gq / tq)
                else:
                    result[gi][sku] = 0
        return result

    # ─── main generate entry ───────────────────────────────────────────

    def generate(self) -> str:
        total_gross, total_vol, chargeable, total_ship = self._chargeable()
        ship_alloc = self._shipping_alloc(total_ship, total_vol, total_gross)
        amt_alloc = self._amount_alloc()

        cno = self.contract.get('contract_no', 'UNKNOWN')
        files = []
        group_names = []

        for gi in self.selected:
            gname = self.groups[gi]['name']
            group_names.append(gname)
            suffix = f'【{gname}】' if len(self.selected) > 1 else ''

            tq = self._group_qty(gi)
            ta = amt_alloc[gi]

            # 1) Export contract
            f1 = gen_export_contract(
                items=self.items, contract=self.contract, kb=self.kb,
                cno=cno, suffix=suffix, tq=tq, ta=ta, ship_alloc=ship_alloc,
                total_gross=total_gross, total_vol=total_vol,
                chargeable=chargeable, total_ship=total_ship,
                rate=self.rate, ship_rate=self.ship_rate, out_dir=self.out_dir,
            )
            files.append(os.path.basename(f1))

            # 2) IV & PL
            f2 = gen_iv_pl(
                items=self.items, kb=self.kb,
                cno=cno, suffix=suffix, tq=tq, ta=ta, ship_alloc=ship_alloc,
                rate=self.rate, out_dir=self.out_dir,
            )
            files.append(os.path.basename(f2))

            # 3) Declaration draft
            f3 = gen_declaration(
                items=self.items, contract=self.contract, kb=self.kb,
                cno=cno, suffix=suffix, tq=tq, ta=ta, ship_alloc=ship_alloc,
                total_ship=total_ship, rate=self.rate,
                price_term=self.price_term, out_dir=self.out_dir,
            )
            files.append(os.path.basename(f3))

        # Summary
        all_qty = {}
        total_usd = 0
        for gi in self.selected:
            tq = self._group_qty(gi)
            ta = amt_alloc[gi]
            for item in self.items:
                sku = sku_key(item)
                q = tq.get(sku, 0)
                a = ta.get(sku, 0)
                all_qty[sku] = all_qty.get(sku, 0) + q
                if q > 0:
                    tax_excl = a / 1.13
                    total_usd += (tax_excl + ship_alloc.get(sku, 0)) / self.rate

        total_boxes = 0
        total_nw = 0.0
        total_gw = 0.0
        for item in self.items:
            sku = sku_key(item)
            q = all_qty.get(sku, 0)
            if q > 0:
                pr = item.get('packing_rate', 1) or 1
                bx = q / pr
                total_boxes += int(bx)
                total_nw += item.get('net_weight_kg', 0) * bx
                total_gw += item.get('gross_weight_kg', 0) * bx

        summary = {
            'groups_generated': group_names,
            'files': files,
            'summary': {
                'total_qty': sum(all_qty.values()),
                'total_usd': round(total_usd, 2),
                'total_boxes': total_boxes,
                'total_gross_weight': round(total_gw, 1),
                'total_net_weight': round(total_nw, 1),
                'exchange_rate_check': round(total_usd * self.rate / (sum(amt_alloc[self.selected[0]].values()) / 1.13 + total_ship), 4) if total_usd > 0 else 0,
                'chargeable_weight': round(chargeable, 2),
                'total_shipping_rmb': round(total_ship, 2),
            },
        }
        return json.dumps(summary, ensure_ascii=False, indent=2)
