#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified pytest for all three document generators:
  - gen_declaration (报关单草稿)
  - gen_iv_pl (IV & Packing List)
  - gen_export_contract (出货合同)

Generates real Excel files to output/test/ for manual inspection,
then validates cell values, calculations, fonts, borders, and layout.

Run:  cd scripts && python -m pytest test_all_generators.py -v
"""

import os
import sys
import pytest

# Ensure the scripts directory is on the import path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from gen_declaration import gen_declaration
from gen_iv_pl import gen_iv_pl
from gen_export_contract import gen_export_contract
from helpers import sku_key, get_lwh


# ═══════════════════════════════════════════════════════════════════════════════
# Shared sample data (modeled after PO2603230466)
# ═══════════════════════════════════════════════════════════════════════════════

SAMPLE_ITEMS = [
    {
        "name_cn": "人造草坪拼接地板",

        "spec": "30*30cm 9pcs",
        "fba_sku": "AG-TILE-3030-9P",
        "unit": "件",
        "quantity": 3600,
        "packing_rate": 12,
        "unit_price_with_tax": 35.50,
        "package_size_cm": [43, 45, 54],
        "net_weight_kg": 8.5,
        "gross_weight_kg": 9.2,
        "total_amount": 127800.00,
    },
    {
        "name_cn": "塑料草坪围栏",

        "spec": "60*40cm",
        "fba_sku": "PG-FENCE-6040",
        "unit": "件",
        "quantity": 2400,
        "packing_rate": 8,
        "unit_price_with_tax": 28.00,
        "package_size_cm": [62, 42, 35],
        "net_weight_kg": 7.0,
        "gross_weight_kg": 7.8,
        "total_amount": 67200.00,
    },
    {
        "name_cn": "仿真植物墙装饰",

        "spec": "40*60cm",
        "fba_sku": "AP-WALL-4060",
        "unit": "件",
        "quantity": 1800,
        "packing_rate": 6,
        "unit_price_with_tax": 42.00,
        "package_size_cm": [65, 45, 30],
        "net_weight_kg": 6.5,
        "gross_weight_kg": 7.2,
        "total_amount": 75600.00,
    },
]

# Declaration test needs unit_1/unit_2 in KB; IV/PL and contract don't
SAMPLE_KB = {
    "AG-TILE-3030-9P": {
        "tariff_code": "3918909000",
        "english_name": "Artificial Grass Interlocking Floor Tiles",
        "declaration_elements": "0|0|塑料|人造草坪拼接地板|无品牌|无型号",
        "material": "plastic",
        "unit_1": "千克",
        "unit_2": "",  # empty → fallback to contract unit "件"
    },
    "PG-FENCE-6040": {
        "tariff_code": "3926909090",
        "english_name": "Plastic Grass Fence Panel",
        "declaration_elements": "0|0|塑料|塑料草坪围栏|无品牌|无型号",
        "material": "plastic",
        "unit_1": "千克",
        "unit_2": "件",
    },
    "AP-WALL-4060": {
        "tariff_code": "6702100000",
        "english_name": "Artificial Plant Wall Decor",
        "declaration_elements": "0|0|塑料|仿真植物墙|无品牌|无型号",
        "material": "plastic",
        "unit_1": "件",
        "unit_2": "千克",
    },
}

# Contract with all fields (superset needed by declaration + export contract)
SAMPLE_CONTRACT = {
    "contract_no": "PO2603230466",
    "date": "2026-03-23",
    "supplier": {
        "name": "义乌市绿美工艺品有限公司",
        "city": "义乌",
        "address": "",
        "contact": "",
        "phone": "",
    },
    "buyer": {
        "name": "深圳市艾进贸易有限公司",
        "address": "",
        "contact": "",
        "phone": "",
    },
    "grand_total": 270600.00,
}

# Declaration uses a different buyer name
SAMPLE_CONTRACT_DECL = {
    **SAMPLE_CONTRACT,
    "buyer": {"name": "ZEATALINE INTERNATIONAL TRADING"},
}

EXCHANGE_RATE = 7.25
SHIPPING_RATE = 8.5
CNO = "PO2603230466"

# ─── Computed values ────────────────────────────────────────────────────────

TQ = {sku_key(item): item["quantity"] for item in SAMPLE_ITEMS}
TA = {sku_key(item): item["total_amount"] for item in SAMPLE_ITEMS}


def _compute_shipping(items, ship_rate):
    """Compute shipping allocation, total_gross, total_vol, chargeable, total_ship."""
    total_gross = 0.0
    total_vol = 0.0
    for item in items:
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        total_gross += item["gross_weight_kg"] * boxes
        total_vol += (l * w * h / 6000) * boxes
    chargeable = max(total_gross, total_vol)
    total_ship = chargeable * ship_rate
    use_vol = total_vol > total_gross
    alloc = {}
    for item in items:
        sku = sku_key(item)
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        sku_vol = (l * w * h / 6000) * boxes
        sku_gross = item["gross_weight_kg"] * boxes
        if use_vol:
            prop = sku_vol / total_vol if total_vol > 0 else 0
        else:
            prop = sku_gross / total_gross if total_gross > 0 else 0
        alloc[sku] = total_ship * prop
    return alloc, total_gross, total_vol, chargeable, total_ship


SHIP_ALLOC, TOTAL_GROSS, TOTAL_VOL, CHARGEABLE, TOTAL_SHIP = _compute_shipping(
    SAMPLE_ITEMS, SHIPPING_RATE
)


# ═══════════════════════════════════════════════════════════════════════════════
# Fixtures
# ═══════════════════════════════════════════════════════════════════════════════

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", "test"
)


@pytest.fixture(scope="session")
def out_dir():
    """Return output/test/ directory, creating it if needed. Files are kept for manual inspection."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return OUTPUT_DIR


@pytest.fixture(scope="class")
def declaration_wb(out_dir):
    """Generate declaration draft and return (workbook, filepath, warnings)."""
    fp, warnings = gen_declaration(
        items=SAMPLE_ITEMS,
        contract=SAMPLE_CONTRACT_DECL,
        kb=SAMPLE_KB,
        cno=CNO,
        suffix="",
        tq=TQ,
        ta=TA,
        ship_alloc=SHIP_ALLOC,
        total_ship=TOTAL_SHIP,
        rate=EXCHANGE_RATE,
        price_term="CNF",
        out_dir=out_dir,
    )
    wb = load_workbook(fp)
    return wb, fp, warnings


@pytest.fixture(scope="class")
def iv_pl_wb(out_dir):
    """Generate IV&PL and return (workbook, filepath)."""
    fp = gen_iv_pl(
        items=SAMPLE_ITEMS,
        kb=SAMPLE_KB,
        cno=CNO,
        suffix="",
        tq=TQ,
        ta=TA,
        ship_alloc=SHIP_ALLOC,
        rate=EXCHANGE_RATE,
        out_dir=out_dir,
    )
    wb = load_workbook(fp)
    return wb, fp


@pytest.fixture(scope="class")
def export_contract_wb(out_dir):
    """Generate export contract and return (workbook, filepath)."""
    fp = gen_export_contract(
        items=SAMPLE_ITEMS,
        contract=SAMPLE_CONTRACT,
        kb=SAMPLE_KB,
        cno=CNO,
        suffix="",
        tq=TQ,
        ta=TA,
        ship_alloc=SHIP_ALLOC,
        total_gross=TOTAL_GROSS,
        total_vol=TOTAL_VOL,
        chargeable=CHARGEABLE,
        total_ship=TOTAL_SHIP,
        rate=EXCHANGE_RATE,
        ship_rate=SHIPPING_RATE,
        out_dir=out_dir,
    )
    wb = load_workbook(fp)
    return wb, fp


# ═══════════════════════════════════════════════════════════════════════════════
# Test: 报关单草稿 (Declaration Draft)
# ═══════════════════════════════════════════════════════════════════════════════


class TestDeclaration:
    def test_file_created(self, declaration_wb):
        _, fp, _ = declaration_wb
        assert os.path.exists(fp)

    def test_sheet_title(self, declaration_wb):
        wb, _, _ = declaration_wb
        assert wb.active.title == "报关单草稿"

    def test_title_row(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["A1"].value == "中华人民共和国海关出口货物报关单"
        assert ws["A1"].font.size == 20
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].font.name == "宋体"

    def test_contract_number(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["A10"].value == CNO
        assert ws["A10"].font.bold is True

    def test_trade_country_red(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["E10"].value == "美国"
        assert ws["E10"].font.color and ws["E10"].font.color.rgb == "00FF0000"

    def test_price_term(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["I12"].value == "CNF"

    def test_shipping_cost_usd(self, declaration_wb):
        ws = declaration_wb[0].active
        expected = round(TOTAL_SHIP / EXCHANGE_RATE, 2)
        assert ws["L12"].value == expected

    def test_totals_row12(self, declaration_wb):
        ws = declaration_wb[0].active
        assert isinstance(ws["E12"].value, int) and ws["E12"].value > 0  # boxes
        assert ws["F12"].value > 0  # gross weight
        assert ws["G12"].value > 0  # net weight

    def test_item_rows(self, declaration_wb):
        ws = declaration_wb[0].active
        row = 20
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            kb_entry = SAMPLE_KB[sku]
            contract_unit = item["unit"]
            u1 = kb_entry["unit_1"] or "千克"
            u2 = kb_entry["unit_2"] or contract_unit

            qty_val = item["quantity"]
            pr = item.get("packing_rate", 1) or 1
            nw_kg = item["net_weight_kg"] * (qty_val / pr)

            def expected_qty(unit):
                if unit == "千克":
                    return round(nw_kg, 1)
                elif unit == contract_unit:
                    return qty_val
                else:
                    return None

            exp_qty_1 = expected_qty(u1)
            exp_qty_2 = expected_qty(u2)

            # Row 1
            assert ws.cell(row=row, column=1).value == idx
            assert ws.cell(row=row, column=2).value == kb_entry["tariff_code"]
            assert ws.cell(row=row, column=4).value == item["name_cn"]
            assert ws.cell(row=row, column=8).value == u1
            if exp_qty_1 is not None:
                assert ws.cell(row=row, column=7).value == exp_qty_1
            else:
                assert ws.cell(row=row, column=7).value is None
            unit_usd = ws.cell(row=row, column=9).value
            assert isinstance(unit_usd, float) and unit_usd > 0
            assert ws.cell(row=row, column=11).value == "中国"
            assert ws.cell(row=row, column=13).value == "美国"
            assert "义乌" in str(ws.cell(row=row, column=16).value)
            assert ws.cell(row=row, column=19).value == "照章征税"

            # Row 2
            assert ws.cell(row=row + 1, column=4).value == kb_entry["declaration_elements"]
            assert ws.cell(row=row + 1, column=8).value == u2
            if exp_qty_2 is not None:
                assert ws.cell(row=row + 1, column=7).value == exp_qty_2
            else:
                assert ws.cell(row=row + 1, column=7).value is None
            total_usd = ws.cell(row=row + 1, column=9).value
            assert isinstance(total_usd, float) and total_usd > 0

            # Row 3
            assert ws.cell(row=row + 2, column=8).value == u1
            if exp_qty_1 is not None:
                assert ws.cell(row=row + 2, column=7).value == exp_qty_1
            assert ws.cell(row=row + 2, column=9).value == "美元"

            row += 3

    def test_no_warnings(self, declaration_wb):
        _, _, warnings = declaration_wb
        assert len(warnings) == 0, f"Unexpected warnings: {warnings}"

    def test_row3_borders(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["A3"].border.top.style == "medium"
        assert ws["A3"].border.left.style == "medium"

    def test_row19_headers(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws.cell(row=19, column=1).value == "项号"
        assert ws.cell(row=19, column=2).value == "商品编号"
        assert ws.cell(row=19, column=4).value == "商品名称及规格型号"
        c_b19 = ws.cell(row=19, column=2)
        assert c_b19.font.color and c_b19.font.color.rgb == "00FF0000"

    def test_font_family(self, declaration_wb):
        ws = declaration_wb[0].active
        assert ws["A1"].font.name == "宋体"
        assert ws["A3"].font.name == "宋体"


# ═══════════════════════════════════════════════════════════════════════════════
# Test: IV & Packing List
# ═══════════════════════════════════════════════════════════════════════════════


class TestIvPl:
    # ─── Sheet structure ────────────────────────────────────────────

    def test_file_created(self, iv_pl_wb):
        _, fp = iv_pl_wb
        assert os.path.exists(fp)

    def test_sheets(self, iv_pl_wb):
        wb, _ = iv_pl_wb
        assert wb.sheetnames == ["IV", "PL"]

    # ─── Invoice (IV) ──────────────────────────────────────────────

    def test_iv_title(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert iv["A1"].value == "INVOICE"
        assert iv["A1"].font.size == 16
        assert iv["A1"].font.name == "Arial"
        assert iv["A1"].font.bold is True
        assert iv["A1"].alignment.horizontal == "center"

    def test_iv_shipper(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert "Shenzhen Adhoc" in str(iv["B2"].value)

    def test_iv_invoice_no(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert iv["F2"].value == CNO

    def test_iv_cnee(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert "ZEATALINE" in str(iv["B5"].value)

    def test_iv_price_terms(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert iv["F5"].value == "C&F"
        assert iv["G5"].value == "USA"

    def test_iv_english_headers(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        expected = [
            "No.", "Tariff Code", "Descriptions", "Qty", "Unit",
            "Unit Price", "USD", "Total Amount", "material quality", "picture",
        ]
        for c, h in enumerate(expected, 1):
            assert iv.cell(row=7, column=c).value == h

    def test_iv_chinese_headers(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        expected = [
            None, "海关编码（清关用的）", "英文品名", "数量", "PC(S)",
            "单价", "USD", "总价", "材质", "产品图片",
        ]
        for c, h in enumerate(expected, 1):
            actual = iv.cell(row=8, column=c).value
            if h is None:
                assert actual is None
            else:
                assert actual == h

    def test_iv_item_rows(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        row = 9
        total_usd_check = 0.0
        total_qty_check = 0
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            qty = TQ[sku]
            amt = TA[sku]
            shipping = SHIP_ALLOC[sku]
            tax_excl = amt / 1.13
            cnf_total_rmb = tax_excl + shipping
            expected_unit_usd = cnf_total_rmb / qty / EXCHANGE_RATE
            expected_total_usd = cnf_total_rmb / EXCHANGE_RATE

            assert iv.cell(row=row, column=1).value == idx
            assert str(iv.cell(row=row, column=2).value) == SAMPLE_KB[sku]["tariff_code"]
            assert iv.cell(row=row, column=3).value == SAMPLE_KB[sku]["english_name"]
            assert iv.cell(row=row, column=4).value == qty
            assert iv.cell(row=row, column=5).value == "PC(S)"
            assert abs(iv.cell(row=row, column=6).value - expected_unit_usd) < 0.01
            assert iv.cell(row=row, column=7).value == "USD"
            assert abs(iv.cell(row=row, column=8).value - expected_total_usd) < 0.01
            assert iv.cell(row=row, column=9).value == SAMPLE_KB[sku]["material"]

            total_usd_check += expected_total_usd
            total_qty_check += qty
            row += 1

        # Empty placeholder row
        assert iv.cell(row=row, column=1).value == len(SAMPLE_ITEMS) + 1
        row += 1

        # Totals row
        assert iv.cell(row=row, column=4).value == total_qty_check
        assert iv.cell(row=row, column=5).value == "/"
        assert iv.cell(row=row, column=6).value == "/"
        assert abs(iv.cell(row=row, column=8).value - total_usd_check) < 0.01

    def test_iv_row_heights(self, iv_pl_wb):
        iv = iv_pl_wb[0]["IV"]
        assert iv.row_dimensions[1].height == 30
        assert iv.row_dimensions[3].height == 70

    # ─── Packing List (PL) ─────────────────────────────────────────

    def test_pl_title(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        assert pl["A1"].value == "PACKING LIST"
        assert pl["A1"].font.size == 16
        assert pl["A1"].font.name == "Arial"
        assert pl["A1"].font.bold is True

    def test_pl_shipper(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        assert "Shenzhen Adhoc" in str(pl["B2"].value)

    def test_pl_invoice_no(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        assert pl["F2"].value == CNO

    def test_pl_headers(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        expected = [
            "NO", "Tariff Code", "Descriptions", "Qty", "Unit",
            "Box Qty\n(CTNS)", "N.W.\n(KG)", "G.W.\n(KG)", "VOLUME (CBM)", None,
        ]
        for c, h in enumerate(expected, 1):
            actual = pl.cell(row=7, column=c).value
            if h is None:
                assert actual is None or actual == ""
            else:
                assert actual == h

    def test_pl_item_rows(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        row = 9
        tot_qty, tot_boxes, tot_nw, tot_gw, tot_vol = 0, 0, 0.0, 0.0, 0.0
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            qty = TQ[sku]
            pr = item.get("packing_rate", 1) or 1
            boxes = qty / pr
            l, w, h = get_lwh(item)
            nw = item["net_weight_kg"] * boxes
            gw = item["gross_weight_kg"] * boxes
            volume = (l * w * h / 1_000_000) * boxes

            assert pl.cell(row=row, column=1).value == idx
            assert str(pl.cell(row=row, column=2).value) == SAMPLE_KB[sku]["tariff_code"]
            assert pl.cell(row=row, column=3).value == SAMPLE_KB[sku]["english_name"]
            assert pl.cell(row=row, column=4).value == qty
            assert pl.cell(row=row, column=4).font.name == "宋体"
            assert pl.cell(row=row, column=4).font.size == 14
            assert pl.cell(row=row, column=5).value == "PC(S)"
            assert pl.cell(row=row, column=6).value == boxes
            assert pl.cell(row=row, column=6).font.name == "宋体"
            assert pl.cell(row=row, column=6).font.size == 14
            assert abs(pl.cell(row=row, column=7).value - round(nw, 1)) < 0.1
            assert abs(pl.cell(row=row, column=8).value - round(gw, 1)) < 0.1
            assert abs(pl.cell(row=row, column=9).value - round(volume, 5)) < 0.001

            tot_qty += qty
            tot_boxes += boxes
            tot_nw += nw
            tot_gw += gw
            tot_vol += volume
            row += 1

        # Placeholder
        assert pl.cell(row=row, column=1).value == len(SAMPLE_ITEMS) + 1
        row += 1

        # Totals
        assert pl.cell(row=row, column=4).value == tot_qty
        assert pl.cell(row=row, column=6).value == round(tot_boxes)
        assert abs(pl.cell(row=row, column=7).value - round(tot_nw, 1)) < 0.1
        assert abs(pl.cell(row=row, column=8).value - round(tot_gw, 1)) < 0.1
        assert abs(pl.cell(row=row, column=9).value - round(tot_vol, 5)) < 0.01

    def test_pl_row_heights(self, iv_pl_wb):
        pl = iv_pl_wb[0]["PL"]
        assert pl.row_dimensions[1].height == 30
        assert pl.row_dimensions[3].height == 61


# ═══════════════════════════════════════════════════════════════════════════════
# Test: 出货合同 (Export Contract)
# ═══════════════════════════════════════════════════════════════════════════════


class TestExportContract:
    def test_file_created(self, export_contract_wb):
        _, fp = export_contract_wb
        assert os.path.exists(fp)

    def test_sheet_title(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws.title == "出口合同"

    def test_title_row(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["A1"].value == "出货合同"
        assert ws["A1"].font.size == 24
        assert ws["A1"].font.name == "宋体"
        assert ws["A1"].alignment.horizontal == "center"

    def test_contract_number(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["I2"].value == CNO

    def test_date(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["D3"].value == "2026-03-23"

    def test_supplier_buyer(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["D4"].value == "义乌市绿美工艺品有限公司"
        assert ws["I4"].value == "深圳市艾进贸易有限公司"

    def test_header_labels(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["A5"].value == "地址："
        assert ws["H5"].value == "地址："
        assert ws["A6"].value == "联系人："
        assert ws["A7"].value == "电话："

    def test_section_label(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws["A9"].value == "一、项目名称、规格型号、数量、金额"

    def test_column_headers(self, export_contract_wb):
        ws = export_contract_wb[0].active
        expected = [
            "产品名称", "产品图片", "规格型号", "FBA SKU", "单位", "数量",
            "箱率", "含税单价/元", "包装尺寸/CM", "外箱净重/KG", "外箱毛重/KG", "总额/元",
        ]
        for c, h in enumerate(expected, 1):
            assert ws.cell(row=11, column=c).value == h

    def test_header_borders(self, export_contract_wb):
        ws = export_contract_wb[0].active
        h1 = ws.cell(row=11, column=1)
        assert h1.border.top.style == "medium"
        assert h1.border.bottom.style == "medium"
        assert h1.border.left.style == "medium"
        h12 = ws.cell(row=11, column=12)
        assert h12.border.right.style == "medium"

    def test_right_side_headers(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws.cell(row=11, column=15).value == "实重"
        assert ws.cell(row=11, column=16).value == "体积重"
        assert ws.cell(row=11, column=17).value == "海运费平摊"
        assert ws.cell(row=11, column=18).value == "C&F总价"
        assert ws.cell(row=11, column=19).value == "C&F单价"

    def test_item_data_rows(self, export_contract_wb):
        ws = export_contract_wb[0].active
        row = 12
        sum_qty = 0
        sum_amt = 0.0
        for idx, item in enumerate(SAMPLE_ITEMS):
            sku = sku_key(item)
            qty = TQ[sku]
            amt = TA[sku]

            # Name (CN + EN from KB)
            name_val = ws.cell(row=row, column=1).value
            assert item["name_cn"] in str(name_val)
            kb_en = SAMPLE_KB[sku]["english_name"]
            assert kb_en in str(name_val)

            assert ws.cell(row=row, column=3).value == item["spec"]
            assert ws.cell(row=row, column=4).value == sku
            assert ws.cell(row=row, column=5).value == item["unit"]
            assert ws.cell(row=row, column=6).value == qty
            assert ws.cell(row=row, column=7).value == item["packing_rate"]

            # Unit price
            expected_up = round(amt / qty, 2)
            assert abs(ws.cell(row=row, column=8).value - expected_up) < 0.01

            # Package size
            l, w, h = item["package_size_cm"]
            assert ws.cell(row=row, column=9).value == f"{int(l)}*{int(w)}*{int(h)}"

            assert ws.cell(row=row, column=10).value == item["net_weight_kg"]
            assert ws.cell(row=row, column=11).value == item["gross_weight_kg"]
            assert abs(ws.cell(row=row, column=12).value - round(amt, 2)) < 0.01

            # Borders
            d1 = ws.cell(row=row, column=1)
            assert d1.border.top.style == "medium"
            assert d1.border.bottom.style == "thin"
            assert d1.border.left.style == "medium"
            d12 = ws.cell(row=row, column=12)
            assert d12.border.right.style == "medium"

            # Alignment
            assert d1.alignment.horizontal == "center"
            assert d1.alignment.vertical == "center"
            assert d1.alignment.wrap_text is True

            # Right-side calculations
            assert ws.cell(row=row, column=15).value > 0  # real weight
            assert ws.cell(row=row, column=16).value > 0  # volume weight
            assert ws.cell(row=row, column=17).value > 0  # shipping alloc
            assert ws.cell(row=row, column=18).value > 0  # C&F total
            assert ws.cell(row=row, column=19).value > 0  # C&F unit

            sum_qty += qty
            sum_amt += amt
            row += 1

    def test_subtotal_row(self, export_contract_wb):
        ws = export_contract_wb[0].active
        row = 12 + len(SAMPLE_ITEMS)  # 小写合计 row
        assert ws.cell(row=row, column=1).value == "小写合计"
        t1 = ws.cell(row=row, column=1)
        assert t1.border.top.style == "medium"
        assert t1.border.left.style == "medium"
        assert ws.cell(row=row, column=14).value == "总重"
        assert ws.cell(row=row, column=15).value == round(TOTAL_GROSS, 1)

    def test_capital_total_row(self, export_contract_wb):
        ws = export_contract_wb[0].active
        row = 12 + len(SAMPLE_ITEMS) + 1  # 大写合计 row
        assert ws.cell(row=row, column=1).value == "大写合计"
        assert ws.cell(row=row, column=14).value == "总海运费"
        assert ws.cell(row=row, column=15).value == round(TOTAL_SHIP, 2)

    def test_empty_bordered_row(self, export_contract_wb):
        ws = export_contract_wb[0].active
        row = 12 + len(SAMPLE_ITEMS) + 2  # empty bordered row
        for ci in range(1, 13):
            assert ws.cell(row=row, column=ci).border.top.style == "medium"

    def test_grand_total_row(self, export_contract_wb):
        ws = export_contract_wb[0].active
        row = 12 + len(SAMPLE_ITEMS) + 3  # 共计 row
        assert ws.cell(row=row, column=1).value == "共计"
        sum_amt = sum(TA.values())
        assert abs(ws.cell(row=row, column=12).value - round(sum_amt, 2)) < 0.01
        g12 = ws.cell(row=row, column=12)
        assert g12.border.bottom.style == "medium"
        assert g12.border.right.style == "medium"

    def test_row_heights(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws.row_dimensions[1].height == 42
        assert ws.row_dimensions[5].height == 27

    def test_contract_terms(self, export_contract_wb):
        ws = export_contract_wb[0].active
        terms_row = 12 + len(SAMPLE_ITEMS) + 3 + 2  # 2 rows after 共计
        found_terms = set()
        for r in range(terms_row, terms_row + 40):
            v = ws.cell(row=r, column=1).value
            if v:
                for keyword in ["交期", "账期", "交货地点", "保密协议", "验货标准",
                                "包装标准", "售后保障", "纠纷", "开票", "壹式贰份"]:
                    if keyword in str(v):
                        found_terms.add(keyword)
                if "供方：" in str(v):
                    found_terms.add("供方签章")
                if "盖章：" in str(v):
                    found_terms.add("盖章")

        expected = {
            "交期", "账期", "交货地点", "保密协议", "验货标准", "包装标准",
            "售后保障", "纠纷", "开票", "壹式贰份", "供方签章", "盖章",
        }
        missing = expected - found_terms
        assert not missing, f"Missing contract terms: {missing}"

    def test_data_font(self, export_contract_wb):
        ws = export_contract_wb[0].active
        assert ws.cell(row=12, column=1).font.name == "宋体"
        assert ws.cell(row=12, column=1).font.size == 10
