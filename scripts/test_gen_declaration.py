#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script for gen_declaration.py — generates a sample 报关单草稿 (customs declaration draft)
and verifies the output Excel structure and styling.
"""

import os
import sys
import tempfile

# Ensure the scripts directory is on the import path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border
from gen_declaration import gen_declaration
from helpers import sku_key

# ─── Sample data (modeled after PO2603230466) ────────────────────────────────

SAMPLE_ITEMS = [
    {
        "name_cn": "人造草坪拼接地板",
        "name_en": "Artificial Grass Interlocking Floor Tiles",
        "spec": "30*30cm 9pcs",
        "fba_sku": "AG-TILE-3030-9P",
        "unit": "件",
        "quantity": 3600,
        "packing_rate": 12,
        "unit_price_with_tax": 35.50,
        "package_size_cm": [43, 45, 54],
        "net_weight_kg": 8.5,
        "gross_weight_kg": 9.2,
        "total_amount": 127800.00,
    },
    {
        "name_cn": "塑料草坪围栏",
        "name_en": "Plastic Grass Fence Panel",
        "spec": "60*40cm",
        "fba_sku": "PG-FENCE-6040",
        "unit": "件",
        "quantity": 2400,
        "packing_rate": 8,
        "unit_price_with_tax": 28.00,
        "package_size_cm": [62, 42, 35],
        "net_weight_kg": 7.0,
        "gross_weight_kg": 7.8,
        "total_amount": 67200.00,
    },
    {
        "name_cn": "仿真植物墙装饰",
        "name_en": "Artificial Plant Wall Decor",
        "spec": "40*60cm",
        "fba_sku": "AP-WALL-4060",
        "unit": "件",
        "quantity": 1800,
        "packing_rate": 6,
        "unit_price_with_tax": 42.00,
        "package_size_cm": [65, 45, 30],
        "net_weight_kg": 6.5,
        "gross_weight_kg": 7.2,
        "total_amount": 75600.00,
    },
]

SAMPLE_CONTRACT = {
    "contract_no": "PO2603230466",
    "date": "2026-03-23",
    "supplier": {"name": "义乌市绿美工艺品有限公司", "city": "义乌"},
    "buyer": {"name": "ZEATALINE INTERNATIONAL TRADING"},
    "grand_total": 270600.00,
}

SAMPLE_KB = {
    "AG-TILE-3030-9P": {
        "tariff_code": "3918909000",
        "english_name": "Artificial Grass Interlocking Floor Tiles",
        "declaration_elements": "0|0|塑料|人造草坪拼接地板|无品牌|无型号",
        "material": "plastic",
    },
    "PG-FENCE-6040": {
        "tariff_code": "3926909090",
        "english_name": "Plastic Grass Fence Panel",
        "declaration_elements": "0|0|塑料|塑料草坪围栏|无品牌|无型号",
        "material": "plastic",
    },
    "AP-WALL-4060": {
        "tariff_code": "6702100000",
        "english_name": "Artificial Plant Wall Decor",
        "declaration_elements": "0|0|塑料|仿真植物墙|无品牌|无型号",
        "material": "plastic",
    },
}

# Simulated computed values (normally from generator_base.py)
EXCHANGE_RATE = 7.25
SHIPPING_RATE = 8.5

# Build ticket quantities — use full quantities for a single-group scenario
TQ = {sku_key(item): item["quantity"] for item in SAMPLE_ITEMS}

# Build ticket amounts — use full amounts
TA = {sku_key(item): item["total_amount"] for item in SAMPLE_ITEMS}

# Compute shipping allocation (simplified proportional by gross weight)
def _compute_ship_alloc(items, ship_rate):
    total_gross = 0.0
    total_vol = 0.0
    for item in items:
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = item["package_size_cm"]
        total_gross += item["gross_weight_kg"] * boxes
        total_vol += (l * w * h / 6000) * boxes
    chargeable = max(total_gross, total_vol)
    total_ship = chargeable * ship_rate
    use_vol = total_vol > total_gross
    alloc = {}
    for item in items:
        sku = sku_key(item)
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = item["package_size_cm"]
        sku_vol = (l * w * h / 6000) * boxes
        sku_gross = item["gross_weight_kg"] * boxes
        if use_vol:
            prop = sku_vol / total_vol if total_vol > 0 else 0
        else:
            prop = sku_gross / total_gross if total_gross > 0 else 0
        alloc[sku] = total_ship * prop
    return alloc, total_ship

SHIP_ALLOC, TOTAL_SHIP = _compute_ship_alloc(SAMPLE_ITEMS, SHIPPING_RATE)


# ─── Test ─────────────────────────────────────────────────────────────────────

def test_gen_declaration():
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = gen_declaration(
            items=SAMPLE_ITEMS,
            contract=SAMPLE_CONTRACT,
            kb=SAMPLE_KB,
            cno="PO2603230466",
            suffix="",
            tq=TQ,
            ta=TA,
            ship_alloc=SHIP_ALLOC,
            total_ship=TOTAL_SHIP,
            rate=EXCHANGE_RATE,
            price_term="CNF",
            out_dir=tmpdir,
        )

        # 1) File exists
        assert os.path.exists(fp), f"Output file not created: {fp}"
        print(f"[PASS] File created: {os.path.basename(fp)}")

        # 2) Load workbook
        wb = load_workbook(fp)
        ws = wb.active
        assert ws.title == "报关单草稿", f"Unexpected sheet title: {ws.title}"
        print(f"[PASS] Sheet title: {ws.title}")

        # 3) Title row
        assert ws["A1"].value == "中华人民共和国海关出口货物报关单"
        assert ws["A1"].font.size == 20
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"
        print("[PASS] Title: text, font size 20 bold, center-aligned")

        # 4) Contract number in row 10
        assert ws["A10"].value == "PO2603230466"
        assert ws["A10"].font.bold is True
        print(f"[PASS] Contract number in A10: {ws['A10'].value}")

        # 5) Trade country
        assert ws["E10"].value == "美国"
        assert ws["E10"].font.color and ws["E10"].font.color.rgb == "00FF0000"
        print("[PASS] Trade country '美国' in red font")

        # 6) Transaction terms
        assert ws["I12"].value == "CNF"
        print(f"[PASS] Price term: {ws['I12'].value}")

        # 7) Shipping cost in USD
        ship_usd = ws["L12"].value
        expected_ship_usd = round(TOTAL_SHIP / EXCHANGE_RATE, 2)
        assert ship_usd == expected_ship_usd, f"Shipping USD mismatch: {ship_usd} != {expected_ship_usd}"
        print(f"[PASS] Shipping cost USD: {ship_usd}")

        # 8) Total boxes, gross weight, net weight in row 12
        assert isinstance(ws["E12"].value, int), f"Boxes should be int, got {type(ws['E12'].value)}"
        assert ws["E12"].value > 0
        print(f"[PASS] Total boxes: {ws['E12'].value}")
        assert ws["F12"].value > 0
        print(f"[PASS] Gross weight: {ws['F12'].value}")
        assert ws["G12"].value > 0
        print(f"[PASS] Net weight: {ws['G12'].value}")

        # 9) Item rows start at row 20, each item takes 3 rows
        row = 20
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            # Row 1: item number
            assert ws.cell(row=row, column=1).value == idx, f"Item {idx} number mismatch"
            # Row 1: tariff code
            tariff = ws.cell(row=row, column=2).value
            assert tariff == SAMPLE_KB[sku]["tariff_code"], f"Item {idx} tariff mismatch: {tariff}"
            # Row 1: Chinese name
            assert ws.cell(row=row, column=4).value == item["name_cn"]
            # Row 1: quantity
            assert ws.cell(row=row, column=7).value == item["quantity"]
            # Row 1: unit price in USD
            unit_usd = ws.cell(row=row, column=9).value
            assert isinstance(unit_usd, float) and unit_usd > 0
            # Row 1: origin country
            assert ws.cell(row=row, column=11).value == "中国"
            # Row 1: destination
            assert ws.cell(row=row, column=13).value == "美国"
            # Row 1: supplier city
            city_val = ws.cell(row=row, column=16).value
            assert "义乌" in str(city_val)
            # Row 1: tax type
            assert ws.cell(row=row, column=19).value == "照章征税"

            # Row 2: declaration elements
            decl_elem = ws.cell(row=row + 1, column=4).value
            assert decl_elem == SAMPLE_KB[sku]["declaration_elements"]
            # Row 2: net weight in kg
            nw_val = ws.cell(row=row + 1, column=7).value
            assert nw_val > 0
            # Row 2: total USD
            total_usd = ws.cell(row=row + 1, column=9).value
            assert isinstance(total_usd, float) and total_usd > 0

            # Row 3: currency label
            assert ws.cell(row=row + 2, column=9).value == "美元"

            print(f"[PASS] Item {idx} ({sku}): 3-row layout, tariff={tariff}, qty={item['quantity']}")
            row += 3

        # 10) Border checks — title row merged, medium borders on row 3
        row3_a = ws["A3"]
        assert row3_a.border.top.style == "medium", f"Row 3 top border should be medium, got {row3_a.border.top.style}"
        assert row3_a.border.left.style == "medium"
        print("[PASS] Row 3 borders: medium top and left")

        # 11) Row 19 column headers
        assert ws.cell(row=19, column=1).value == "项号"
        assert ws.cell(row=19, column=2).value == "商品编号"
        assert ws.cell(row=19, column=4).value == "商品名称及规格型号"
        assert "商品编号" in str(ws.cell(row=19, column=2).value)
        # Red font for 商品编号
        c_b19 = ws.cell(row=19, column=2)
        assert c_b19.font.color and c_b19.font.color.rgb == "00FF0000"
        print("[PASS] Row 19 column headers with correct fonts")

        # 12) Font family checks
        assert ws["A1"].font.name == "宋体"
        assert ws["A3"].font.name == "宋体"
        print("[PASS] Font family: 宋体 throughout")

        print("\n" + "=" * 60)
        print("ALL TESTS PASSED for gen_declaration")
        print("=" * 60)


if __name__ == "__main__":
    test_gen_declaration()
