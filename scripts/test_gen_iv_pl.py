#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script for gen_iv_pl.py — generates a sample IV&PL (Invoice & Packing List)
and verifies the output Excel structure and styling.
"""

import os
import sys
import tempfile

# Ensure the scripts directory is on the import path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from gen_iv_pl import gen_iv_pl
from helpers import sku_key, get_lwh

# ─── Sample data (modeled after PO2603230466) ────────────────────────────────

SAMPLE_ITEMS = [
    {
        "name_cn": "人造草坪拼接地板",
        "name_en": "Artificial Grass Interlocking Floor Tiles",
        "spec": "30*30cm 9pcs",
        "fba_sku": "AG-TILE-3030-9P",
        "unit": "件",
        "quantity": 3600,
        "packing_rate": 12,
        "unit_price_with_tax": 35.50,
        "package_size_cm": [43, 45, 54],
        "net_weight_kg": 8.5,
        "gross_weight_kg": 9.2,
        "total_amount": 127800.00,
    },
    {
        "name_cn": "塑料草坪围栏",
        "name_en": "Plastic Grass Fence Panel",
        "spec": "60*40cm",
        "fba_sku": "PG-FENCE-6040",
        "unit": "件",
        "quantity": 2400,
        "packing_rate": 8,
        "unit_price_with_tax": 28.00,
        "package_size_cm": [62, 42, 35],
        "net_weight_kg": 7.0,
        "gross_weight_kg": 7.8,
        "total_amount": 67200.00,
    },
    {
        "name_cn": "仿真植物墙装饰",
        "name_en": "Artificial Plant Wall Decor",
        "spec": "40*60cm",
        "fba_sku": "AP-WALL-4060",
        "unit": "件",
        "quantity": 1800,
        "packing_rate": 6,
        "unit_price_with_tax": 42.00,
        "package_size_cm": [65, 45, 30],
        "net_weight_kg": 6.5,
        "gross_weight_kg": 7.2,
        "total_amount": 75600.00,
    },
]

SAMPLE_KB = {
    "AG-TILE-3030-9P": {
        "tariff_code": "3918909000",
        "english_name": "Artificial Grass Interlocking Floor Tiles",
        "declaration_elements": "0|0|塑料|人造草坪拼接地板|无品牌|无型号",
        "material": "plastic",
    },
    "PG-FENCE-6040": {
        "tariff_code": "3926909090",
        "english_name": "Plastic Grass Fence Panel",
        "declaration_elements": "0|0|塑料|塑料草坪围栏|无品牌|无型号",
        "material": "plastic",
    },
    "AP-WALL-4060": {
        "tariff_code": "6702100000",
        "english_name": "Artificial Plant Wall Decor",
        "declaration_elements": "0|0|塑料|仿真植物墙|无品牌|无型号",
        "material": "plastic",
    },
}

EXCHANGE_RATE = 7.25
SHIPPING_RATE = 8.5

TQ = {sku_key(item): item["quantity"] for item in SAMPLE_ITEMS}
TA = {sku_key(item): item["total_amount"] for item in SAMPLE_ITEMS}


def _compute_ship_alloc(items, ship_rate):
    total_gross = 0.0
    total_vol = 0.0
    for item in items:
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        total_gross += item["gross_weight_kg"] * boxes
        total_vol += (l * w * h / 6000) * boxes
    chargeable = max(total_gross, total_vol)
    total_ship = chargeable * ship_rate
    use_vol = total_vol > total_gross
    alloc = {}
    for item in items:
        sku = sku_key(item)
        pr = item.get("packing_rate", 1) or 1
        boxes = item["quantity"] / pr
        l, w, h = get_lwh(item)
        sku_vol = (l * w * h / 6000) * boxes
        sku_gross = item["gross_weight_kg"] * boxes
        if use_vol:
            prop = sku_vol / total_vol if total_vol > 0 else 0
        else:
            prop = sku_gross / total_gross if total_gross > 0 else 0
        alloc[sku] = total_ship * prop
    return alloc


SHIP_ALLOC = _compute_ship_alloc(SAMPLE_ITEMS, SHIPPING_RATE)


# ─── Test ─────────────────────────────────────────────────────────────────────

def test_gen_iv_pl():
    with tempfile.TemporaryDirectory() as tmpdir:
        fp = gen_iv_pl(
            items=SAMPLE_ITEMS,
            kb=SAMPLE_KB,
            cno="PO2603230466",
            suffix="",
            tq=TQ,
            ta=TA,
            ship_alloc=SHIP_ALLOC,
            rate=EXCHANGE_RATE,
            out_dir=tmpdir,
        )

        # 1) File exists
        assert os.path.exists(fp), f"Output file not created: {fp}"
        print(f"[PASS] File created: {os.path.basename(fp)}")

        # 2) Load workbook — should have IV and PL sheets, no default sheet
        wb = load_workbook(fp)
        sheet_names = wb.sheetnames
        assert "IV" in sheet_names, f"Missing 'IV' sheet, got: {sheet_names}"
        assert "PL" in sheet_names, f"Missing 'PL' sheet, got: {sheet_names}"
        assert len(sheet_names) == 2, f"Expected 2 sheets, got {len(sheet_names)}: {sheet_names}"
        print(f"[PASS] Sheets: {sheet_names}")

        # ─── Invoice (IV) sheet ─────────────────────────────────────
        iv = wb["IV"]

        # 3) Title
        assert iv["A1"].value == "INVOICE"
        assert iv["A1"].font.size == 16
        assert iv["A1"].font.name == "Arial"
        assert iv["A1"].alignment.horizontal == "center"
        print("[PASS] IV title: 'INVOICE', Arial 16, centered")

        # 4) Shipper info
        assert "Shenzhen Adhoc" in str(iv["B2"].value)
        print(f"[PASS] IV shipper: {iv['B2'].value}")

        # 5) Invoice number
        assert iv["F2"].value == "PO2603230466"
        print(f"[PASS] IV invoice no: {iv['F2'].value}")

        # 6) CNEE
        assert "ZEATALINE" in str(iv["B5"].value)
        print(f"[PASS] IV CNEE: {iv['B5'].value}")

        # 7) Price terms
        assert iv["F5"].value == "C&F"
        assert iv["G5"].value == "USA"
        print("[PASS] IV price terms: C&F USA")

        # 8) Column headers (row 7 English, row 8 Chinese)
        en_hdrs = ['No.', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
                   'Unit Price', 'USD', 'Total Amount', 'material quality', 'picture']
        for c, h in enumerate(en_hdrs, 1):
            actual = iv.cell(row=7, column=c).value
            assert actual == h, f"IV header col {c}: {actual} != {h}"
        print("[PASS] IV English column headers correct")

        cn_hdrs_7 = [None, '海关编码（清关用的）', '英文品名', '数量', 'PC(S)',
                     '单价', 'USD', '总价', '材质', '产品图片']
        for c, h in enumerate(cn_hdrs_7, 1):
            actual = iv.cell(row=8, column=c).value
            if h is None:
                assert actual is None, f"IV Chinese header col {c}: {actual} != None"
            else:
                assert actual == h, f"IV Chinese header col {c}: {actual} != {h}"
        print("[PASS] IV Chinese column headers correct")

        # 9) Item rows starting at row 9
        row = 9
        total_usd_check = 0.0
        total_qty_check = 0
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            qty = TQ[sku]
            amt = TA[sku]
            shipping = SHIP_ALLOC[sku]
            tax_excl = amt / 1.13
            cnf_total_rmb = tax_excl + shipping
            expected_unit_usd = cnf_total_rmb / qty / EXCHANGE_RATE
            expected_total_usd = cnf_total_rmb / EXCHANGE_RATE

            # No. column
            assert iv.cell(row=row, column=1).value == idx

            # Tariff code
            tariff = iv.cell(row=row, column=2).value
            assert str(tariff) == SAMPLE_KB[sku]["tariff_code"], f"IV item {idx} tariff: {tariff}"

            # English name
            eng_name = iv.cell(row=row, column=3).value
            assert eng_name == SAMPLE_KB[sku]["english_name"]

            # Qty
            assert iv.cell(row=row, column=4).value == qty

            # Unit
            assert iv.cell(row=row, column=5).value == "PC(S)"

            # Unit price USD
            unit_usd = iv.cell(row=row, column=6).value
            assert abs(unit_usd - expected_unit_usd) < 0.01, \
                f"IV item {idx} unit USD: {unit_usd:.4f} != {expected_unit_usd:.4f}"

            # Currency label
            assert iv.cell(row=row, column=7).value == "USD"

            # Total amount USD
            total_amt_usd = iv.cell(row=row, column=8).value
            assert abs(total_amt_usd - expected_total_usd) < 0.01

            # Material
            mat = iv.cell(row=row, column=9).value
            assert mat == SAMPLE_KB[sku]["material"]

            print(f"[PASS] IV item {idx} ({sku}): tariff={tariff}, unit_usd={unit_usd:.4f}, total={total_amt_usd:.2f}")

            total_usd_check += expected_total_usd
            total_qty_check += qty
            row += 1

        # 10) Empty placeholder row
        assert iv.cell(row=row, column=1).value == len(SAMPLE_ITEMS) + 1
        row += 1

        # 11) Totals row
        assert iv.cell(row=row, column=4).value == total_qty_check
        assert iv.cell(row=row, column=5).value == "/"
        assert iv.cell(row=row, column=6).value == "/"
        total_usd_cell = iv.cell(row=row, column=8).value
        assert abs(total_usd_cell - total_usd_check) < 0.01
        print(f"[PASS] IV totals: qty={total_qty_check}, USD={total_usd_cell:.2f}")

        # ─── Packing List (PL) sheet ────────────────────────────────
        pl = wb["PL"]

        # 12) Title
        assert pl["A1"].value == "PACKING LIST"
        assert pl["A1"].font.size == 16
        assert pl["A1"].font.name == "Arial"
        print("[PASS] PL title: 'PACKING LIST', Arial 16")

        # 13) Shipper info
        assert "Shenzhen Adhoc" in str(pl["B2"].value)
        print(f"[PASS] PL shipper: {pl['B2'].value}")

        # 14) Invoice number
        assert pl["F2"].value == "PO2603230466"
        print(f"[PASS] PL invoice no: {pl['F2'].value}")

        # 15) PL column headers
        pl_hdrs = ['NO', 'Tariff Code', 'Descriptions', 'Qty', 'Unit',
                   'Box Qty\n(CTNS)', 'N.W.\n(KG)', 'G.W.\n(KG)', 'VOLUME (CBM)', None]
        for c, h in enumerate(pl_hdrs, 1):
            actual = pl.cell(row=7, column=c).value
            if h is None:
                assert actual is None or actual == '', f"PL header col {c}: expected empty, got '{actual}'"
            else:
                assert actual == h, f"PL header col {c}: '{actual}' != '{h}'"
        print("[PASS] PL English column headers correct")

        # 16) PL item rows
        row = 9
        tot_qty = 0
        tot_boxes = 0
        tot_nw = 0.0
        tot_gw = 0.0
        tot_vol = 0.0
        for idx, item in enumerate(SAMPLE_ITEMS, 1):
            sku = sku_key(item)
            qty = TQ[sku]
            pr = item.get("packing_rate", 1) or 1
            boxes = qty / pr
            l, w, h = get_lwh(item)
            nw = item["net_weight_kg"] * boxes
            gw = item["gross_weight_kg"] * boxes
            volume = (l * w * h / 1_000_000) * boxes

            # No.
            assert pl.cell(row=row, column=1).value == idx

            # Tariff code
            assert str(pl.cell(row=row, column=2).value) == SAMPLE_KB[sku]["tariff_code"]

            # English name
            assert pl.cell(row=row, column=3).value == SAMPLE_KB[sku]["english_name"]

            # Qty
            assert pl.cell(row=row, column=4).value == qty
            # Qty font should be 宋体 14
            assert pl.cell(row=row, column=4).font.name == "宋体"
            assert pl.cell(row=row, column=4).font.size == 14

            # Unit
            assert pl.cell(row=row, column=5).value == "PC(S)"

            # Box qty
            assert pl.cell(row=row, column=6).value == boxes
            # Box font should be 宋体 14
            assert pl.cell(row=row, column=6).font.name == "宋体"
            assert pl.cell(row=row, column=6).font.size == 14

            # Net weight
            assert abs(pl.cell(row=row, column=7).value - round(nw, 1)) < 0.1

            # Gross weight
            assert abs(pl.cell(row=row, column=8).value - round(gw, 1)) < 0.1

            # Volume
            vol_val = pl.cell(row=row, column=9).value
            assert abs(vol_val - round(volume, 5)) < 0.001

            print(f"[PASS] PL item {idx} ({sku}): boxes={boxes}, NW={round(nw,1)}, GW={round(gw,1)}, vol={round(volume,5)}")

            tot_qty += qty
            tot_boxes += boxes
            tot_nw += nw
            tot_gw += gw
            tot_vol += volume
            row += 1

        # 17) PL empty placeholder
        assert pl.cell(row=row, column=1).value == len(SAMPLE_ITEMS) + 1
        row += 1

        # 18) PL totals
        assert pl.cell(row=row, column=4).value == tot_qty
        assert pl.cell(row=row, column=6).value == round(tot_boxes)
        assert abs(pl.cell(row=row, column=7).value - round(tot_nw, 1)) < 0.1
        assert abs(pl.cell(row=row, column=8).value - round(tot_gw, 1)) < 0.1
        assert abs(pl.cell(row=row, column=9).value - round(tot_vol, 5)) < 0.01
        print(f"[PASS] PL totals: qty={tot_qty}, boxes={round(tot_boxes)}, "
              f"NW={round(tot_nw,1)}, GW={round(tot_gw,1)}, vol={round(tot_vol,5)}")

        # 19) Row heights for IV
        assert iv.row_dimensions[1].height == 30
        assert iv.row_dimensions[3].height == 70
        print(f"[PASS] IV row heights: row1={iv.row_dimensions[1].height}, row3={iv.row_dimensions[3].height}")

        # 20) Row heights for PL
        assert pl.row_dimensions[1].height == 30
        assert pl.row_dimensions[3].height == 61
        print(f"[PASS] PL row heights: row1={pl.row_dimensions[1].height}, row3={pl.row_dimensions[3].height}")

        print("\n" + "=" * 60)
        print("ALL TESTS PASSED for gen_iv_pl")
        print("=" * 60)


if __name__ == "__main__":
    test_gen_iv_pl()
