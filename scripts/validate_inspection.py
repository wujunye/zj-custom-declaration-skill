#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import argparse
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
import os

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def load_declaration(declaration_path: str) -> Dict[str, Any]:
    """Load purchase contract JSON declaration file."""
    try:
        with open(declaration_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        raise ValueError(f"Could not parse declaration file: {e}")


def parse_inspection_xls(file_path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Parse .xls inspection certificate using xlrd."""
    if xlrd is None:
        raise ValueError("xlrd not installed. Install with: pip install xlrd")

    try:
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)

        # Find header row and column mappings
        headers = [cell.value for cell in sheet.row(0)]
        col_mapping = find_columns(headers)

        # Extract data rows
        rows = []
        for row_idx in range(1, sheet.nrows):
            row = sheet.row(row_idx)
            row_data = extract_row_data(row, col_mapping)
            if row_data:  # Skip empty rows
                rows.append(row_data)

        # Calculate totals
        totals = calculate_totals(rows)

        return rows, totals
    except Exception as e:
        raise ValueError(f"Could not parse .xls file: {e}")


def parse_inspection_xlsx(file_path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Parse .xlsx inspection certificate using openpyxl."""
    if load_workbook is None:
        raise ValueError("openpyxl not installed. Install with: pip install openpyxl")

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Find header row
        headers = [cell.value for cell in sheet[1]]
        col_mapping = find_columns(headers)

        # Extract data rows
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_cells = sheet[row_idx]
            row_data = extract_row_data_xlsx(row_cells, col_mapping)
            if row_data:  # Skip empty rows
                rows.append(row_data)

        # Calculate totals
        totals = calculate_totals(rows)

        return rows, totals
    except Exception as e:
        raise ValueError(f"Could not parse .xlsx file: {e}")


def find_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    Find column indices matching expected field names.
    Returns mapping of field names to column indices.
    """
    mapping = {
        'product_name': None,
        'quantity': None,
        'net_weight': None,
        'gross_weight': None,
        'boxes': None,
        'value': None
    }

    # Chinese column name patterns
    patterns = {
        'product_name': ['品名', '产品名称', '商品名称', '物品名称'],
        'quantity': ['数量', '件数', '数目'],
        'net_weight': ['净重', '净重(kg)', '净重kg', '网重'],
        'gross_weight': ['毛重', '毛重(kg)', '毛重kg', '总重'],
        'boxes': ['箱数', '包数', '包装数', '箱'],
        'value': ['金额', '总额', '价格', '单价', '总价', '金额合计']
    }

    headers_lower = [str(h).lower() if h else '' for h in headers]

    for idx, header in enumerate(headers_lower):
        if not header:
            continue

        for field, patterns_list in patterns.items():
            if mapping[field] is None:
                for pattern in patterns_list:
                    if pattern.lower() in header:
                        mapping[field] = idx
                        break

    return mapping


def extract_row_data(row: List[Any], col_mapping: Dict[str, Optional[int]]) -> Optional[Dict[str, Any]]:
    """Extract data from an xlrd row."""
    data = {}

    # Extract product name
    if col_mapping['product_name'] is not None:
        data['product_name'] = str(row[col_mapping['product_name']].value).strip()

    # Extract quantity
    if col_mapping['quantity'] is not None:
        try:
            data['quantity'] = float(row[col_mapping['quantity']].value)
        except (ValueError, IndexError):
            data['quantity'] = None

    # Extract weights
    if col_mapping['net_weight'] is not None:
        try:
            data['net_weight'] = float(row[col_mapping['net_weight']].value)
        except (ValueError, IndexError):
            data['net_weight'] = None

    if col_mapping['gross_weight'] is not None:
        try:
            data['gross_weight'] = float(row[col_mapping['gross_weight']].value)
        except (ValueError, IndexError):
            data['gross_weight'] = None

    # Extract boxes
    if col_mapping['boxes'] is not None:
        try:
            data['boxes'] = float(row[col_mapping['boxes']].value)
        except (ValueError, IndexError):
            data['boxes'] = None

    # Extract value
    if col_mapping['value'] is not None:
        try:
            data['value'] = float(row[col_mapping['value']].value)
        except (ValueError, IndexError):
            data['value'] = None

    # Only return rows that have some meaningful data
    if data.get('quantity') or data.get('value'):
        return data

    return None


def extract_row_data_xlsx(row_cells: List[Any], col_mapping: Dict[str, Optional[int]]) -> Optional[Dict[str, Any]]:
    """Extract data from an openpyxl row."""
    data = {}

    # Extract product name
    if col_mapping['product_name'] is not None:
        cell_value = row_cells[col_mapping['product_name']].value
        if cell_value:
            data['product_name'] = str(cell_value).strip()

    # Extract quantity
    if col_mapping['quantity'] is not None:
        cell_value = row_cells[col_mapping['quantity']].value
        try:
            data['quantity'] = float(cell_value) if cell_value else None
        except (ValueError, TypeError):
            data['quantity'] = None

    # Extract weights
    if col_mapping['net_weight'] is not None:
        cell_value = row_cells[col_mapping['net_weight']].value
        try:
            data['net_weight'] = float(cell_value) if cell_value else None
        except (ValueError, TypeError):
            data['net_weight'] = None

    if col_mapping['gross_weight'] is not None:
        cell_value = row_cells[col_mapping['gross_weight']].value
        try:
            data['gross_weight'] = float(cell_value) if cell_value else None
        except (ValueError, TypeError):
            data['gross_weight'] = None

    # Extract boxes
    if col_mapping['boxes'] is not None:
        cell_value = row_cells[col_mapping['boxes']].value
        try:
            data['boxes'] = float(cell_value) if cell_value else None
        except (ValueError, TypeError):
            data['boxes'] = None

    # Extract value
    if col_mapping['value'] is not None:
        cell_value = row_cells[col_mapping['value']].value
        try:
            data['value'] = float(cell_value) if cell_value else None
        except (ValueError, TypeError):
            data['value'] = None

    # Only return rows that have some meaningful data
    if data.get('quantity') or data.get('value'):
        return data

    return None


def calculate_totals(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Calculate totals from inspection rows."""
    totals = {
        'item_count': len(rows),
        'total_qty': None,
        'total_net_weight': None,
        'total_gross_weight': None,
        'total_boxes': None,
        'total_value': None
    }

    # Sum quantities
    quantities = [r.get('quantity') for r in rows if r.get('quantity') is not None]
    if quantities:
        totals['total_qty'] = sum(quantities)

    # Sum net weights
    net_weights = [r.get('net_weight') for r in rows if r.get('net_weight') is not None]
    if net_weights:
        totals['total_net_weight'] = sum(net_weights)

    # Sum gross weights
    gross_weights = [r.get('gross_weight') for r in rows if r.get('gross_weight') is not None]
    if gross_weights:
        totals['total_gross_weight'] = sum(gross_weights)

    # Sum boxes
    boxes = [r.get('boxes') for r in rows if r.get('boxes') is not None]
    if boxes:
        totals['total_boxes'] = sum(boxes)

    # Sum values
    values = [r.get('value') for r in rows if r.get('value') is not None]
    if values:
        totals['total_value'] = sum(values)

    return totals


def extract_declaration_totals(declaration: Dict[str, Any]) -> Dict[str, Any]:
    """Extract totals from declaration data."""
    totals = {
        'item_count': 0,
        'total_qty': None,
        'total_net_weight': None,
        'total_gross_weight': None,
        'total_boxes': None,
        'total_value': None
    }

    # Try to extract from items array
    items = declaration.get('items', [])
    if not items and 'line_items' in declaration:
        items = declaration.get('line_items', [])

    totals['item_count'] = len(items)

    # Sum quantities
    quantities = []
    for item in items:
        qty = item.get('quantity') or item.get('qty')
        if qty:
            try:
                quantities.append(float(qty))
            except (ValueError, TypeError):
                pass
    if quantities:
        totals['total_qty'] = sum(quantities)

    # Sum net weights
    net_weights = []
    for item in items:
        weight = item.get('net_weight') or item.get('net_wt')
        if weight:
            try:
                net_weights.append(float(weight))
            except (ValueError, TypeError):
                pass
    if net_weights:
        totals['total_net_weight'] = sum(net_weights)

    # Sum gross weights
    gross_weights = []
    for item in items:
        weight = item.get('gross_weight') or item.get('gross_wt')
        if weight:
            try:
                gross_weights.append(float(weight))
            except (ValueError, TypeError):
                pass
    if gross_weights:
        totals['total_gross_weight'] = sum(gross_weights)

    # Sum values
    values = []
    for item in items:
        value = item.get('value') or item.get('amount') or item.get('total_amount')
        if value:
            try:
                values.append(float(value))
            except (ValueError, TypeError):
                pass
    if values:
        totals['total_value'] = sum(values)

    return totals


def check_tolerance(expected: Optional[float], actual: Optional[float], tolerance: float) -> bool:
    """Check if actual value is within tolerance of expected value."""
    if expected is None or actual is None:
        return True  # Can't validate if missing

    if expected == 0:
        return actual == 0

    percentage_diff = abs(actual - expected) / expected
    return percentage_diff <= tolerance


def validate(declaration_totals: Dict[str, Any], inspection_totals: Dict[str, Any]) -> Dict[str, Any]:
    """Validate inspection against declaration."""
    mismatches = []

    # 1. Check item count
    if declaration_totals['item_count'] != inspection_totals['item_count']:
        return {
            'result': 'ITEM_COUNT_MISMATCH',
            'details': {
                'declaration_items': declaration_totals['item_count'],
                'inspection_items': inspection_totals['item_count'],
                'declaration_total_qty': declaration_totals['total_qty'],
                'inspection_total_qty': inspection_totals['total_qty'],
                'declaration_total_amount': declaration_totals['total_value'],
                'inspection_total_amount': inspection_totals['total_value'],
                'declaration_gross_weight': declaration_totals['total_gross_weight'],
                'inspection_gross_weight': inspection_totals['total_gross_weight'],
                'mismatches': [f"Item count mismatch: {declaration_totals['item_count']} vs {inspection_totals['item_count']}"]
            }
        }

    # 2. Check values (2% tolerance)
    if not check_tolerance(declaration_totals['total_value'], inspection_totals['total_value'], 0.02):
        mismatches.append(f"Value mismatch: {declaration_totals['total_value']} vs {inspection_totals['total_value']}")
        return {
            'result': 'VALUE_MISMATCH',
            'details': {
                'declaration_items': declaration_totals['item_count'],
                'inspection_items': inspection_totals['item_count'],
                'declaration_total_qty': declaration_totals['total_qty'],
                'inspection_total_qty': inspection_totals['total_qty'],
                'declaration_total_amount': declaration_totals['total_value'],
                'inspection_total_amount': inspection_totals['total_value'],
                'declaration_gross_weight': declaration_totals['total_gross_weight'],
                'inspection_gross_weight': inspection_totals['total_gross_weight'],
                'mismatches': mismatches
            }
        }

    # 3. Check quantities (1% tolerance)
    if not check_tolerance(declaration_totals['total_qty'], inspection_totals['total_qty'], 0.01):
        mismatches.append(f"Quantity mismatch: {declaration_totals['total_qty']} vs {inspection_totals['total_qty']}")

    # 4. Check weights (5% tolerance)
    if not check_tolerance(declaration_totals['total_gross_weight'], inspection_totals['total_gross_weight'], 0.05):
        mismatches.append(f"Gross weight mismatch: {declaration_totals['total_gross_weight']} vs {inspection_totals['total_gross_weight']}")

    if not check_tolerance(declaration_totals['total_net_weight'], inspection_totals['total_net_weight'], 0.05):
        mismatches.append(f"Net weight mismatch: {declaration_totals['total_net_weight']} vs {inspection_totals['total_net_weight']}")

    if mismatches:
        return {
            'result': 'WEIGHT_MISMATCH',
            'details': {
                'declaration_items': declaration_totals['item_count'],
                'inspection_items': inspection_totals['item_count'],
                'declaration_total_qty': declaration_totals['total_qty'],
                'inspection_total_qty': inspection_totals['total_qty'],
                'declaration_total_amount': declaration_totals['total_value'],
                'inspection_total_amount': inspection_totals['total_value'],
                'declaration_gross_weight': declaration_totals['total_gross_weight'],
                'inspection_gross_weight': inspection_totals['total_gross_weight'],
                'mismatches': mismatches
            }
        }

    return {
        'result': 'PASS',
        'details': {
            'declaration_items': declaration_totals['item_count'],
            'inspection_items': inspection_totals['item_count'],
            'declaration_total_qty': declaration_totals['total_qty'],
            'inspection_total_qty': inspection_totals['total_qty'],
            'declaration_total_amount': declaration_totals['total_value'],
            'inspection_total_amount': inspection_totals['total_value'],
            'declaration_gross_weight': declaration_totals['total_gross_weight'],
            'inspection_gross_weight': inspection_totals['total_gross_weight'],
            'mismatches': []
        }
    }


def main():
    parser = argparse.ArgumentParser(
        description='Validate inspection certificate against customs declaration'
    )
    parser.add_argument(
        '--declaration',
        required=True,
        help='Path to purchase_contract.json (declaration data)'
    )
    parser.add_argument(
        '--inspection',
        required=True,
        help='Path to inspection certificate file (.xls or .xlsx)'
    )
    parser.add_argument(
        '--output',
        default='/tmp/validation_result.json',
        help='Path to output validation_result.json (default: /tmp/validation_result.json)'
    )

    args = parser.parse_args()

    try:
        # Load declaration
        declaration = load_declaration(args.declaration)
        declaration_totals = extract_declaration_totals(declaration)

        # Parse inspection certificate
        inspection_file = args.inspection
        if not os.path.exists(inspection_file):
            raise ValueError(f"Inspection file not found: {inspection_file}")

        file_ext = Path(inspection_file).suffix.lower()

        if file_ext == '.xls':
            inspection_rows, inspection_totals = parse_inspection_xls(inspection_file)
        elif file_ext == '.xlsx':
            inspection_rows, inspection_totals = parse_inspection_xlsx(inspection_file)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Expected .xls or .xlsx")

        # Validate
        result = validate(declaration_totals, inspection_totals)

    except ValueError as e:
        result = {
            'result': 'PARSE_ERROR',
            'details': {'error': str(e)}
        }
    except Exception as e:
        result = {
            'result': 'PARSE_ERROR',
            'details': {'error': f"Unexpected error: {str(e)}"}
        }

    # Write output
    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    # Print result to stdout
    print(json.dumps(result, indent=2, ensure_ascii=False))

    # Exit with appropriate code
    if result['result'] == 'PASS':
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
