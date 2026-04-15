#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Export Contract (出货合同) Excel document.
"""

import os
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from helpers import get_lwh, sku_key


# Border sides
_THIN = Side(style='thin')
_MEDIUM = Side(style='medium')
_NONE = Side()


def gen_export_contract(
    items: List[dict],
    contract: dict,
    kb: dict,
    cno: str,
    suffix: str,
    tq: Dict[str, int],
    ta: Dict[str, float],
    ship_alloc: Dict[str, float],
    total_gross: float,
    total_vol: float,
    chargeable: float,
    total_ship: float,
    rate: float,
    ship_rate: float,
    out_dir: str,
) -> str:
    """Generate the export contract Excel file. Returns the output file path."""

    wb = Workbook()
    ws = wb.active
    ws.title = "出口合同"

    # ── Column widths (11 main cols; right-side calc cols start at index 13) ──
    # Mapping (vs old 12-col layout): old col 2 (产品图片) removed; old 3..11 shift left by 1.
    col_widths = {
        0: 5441/256,   # 产品名称
        1: 4864/256,   # FNSKU
        2: 3626/256,   # 规格型号
        3: 2304/256,   # 单位
        4: 2474/256,   # 数量
        5: 2218/256,   # 箱率
        6: 4266/256,   # 含税单价/元
        7: 4736/256,   # 包装尺寸/CM
        8: 3584/256,   # 外箱净重/KG
        9: 3754/256,   # 外箱毛重/KG
        10: 4053/256,  # 总额/元
        12: 6869/256,
        13: 2730/256,  # 实重 (col 14)
        14: 2773/256,  # 体积重 (col 15)
        15: 3285/256,  # 海运费平摊 (col 16)
        16: 2773/256,  # C&F总价 (col 17)
        17: 3242/256,  # C&F单价 (col 18)
        18: 3669/256,
        19: 3114/256,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    # ── Fonts ──
    font10 = Font(name='宋体', size=10)
    font9 = Font(name='宋体', size=9)
    font24 = Font(name='宋体', size=24)

    # ── Alignments ──
    align_cc = Alignment(horizontal='center', vertical='center')
    align_cc_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_lc = Alignment(horizontal='left', vertical='center')
    align_gc = Alignment(vertical='center')  # general horizontal, center vertical

    # ══════════════════════════════════════════════════════════════════
    # HEADER SECTION (rows 1-10, 1-indexed)
    # ══════════════════════════════════════════════════════════════════

    # Row 1: Title (merge across 11 main cols A:K)
    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = '出货合同'
    c.font = font24
    c.alignment = align_cc

    # Row 2: 合同编号 (label G2, value H2)
    ws['G2'] = '合同编号：'
    ws['G2'].font = font10
    ws['H2'] = cno
    ws['H2'].font = font10

    # Row 3: 日期 (value C3)
    ws['A3'] = '日期：'
    ws['A3'].font = font10
    ws['C3'] = contract.get('date', '')
    ws['C3'].font = font10

    supplier = contract.get('supplier', {})
    buyer = contract.get('buyer', {})

    # Row 4: 供方 / 需方
    ws['A4'] = '供方：'
    ws['A4'].font = font10
    ws['C4'] = supplier.get('name', '')
    ws['C4'].font = font10
    ws['G4'] = '需方：'
    ws['G4'].font = font10
    ws['H4'] = buyer.get('name', '')
    ws['H4'].font = font10

    # Row 5: 地址
    ws['A5'] = '地址：'
    ws['A5'].font = font10
    ws['A5'].alignment = Alignment(wrap_text=True)
    ws['C5'] = supplier.get('address', '')
    ws['C5'].font = font10
    ws['G5'] = '地址：'
    ws['G5'].font = font10
    ws['G5'].alignment = Alignment(wrap_text=True)
    ws['H5'] = buyer.get('address', '')
    ws['H5'].font = font10

    # Row 6: 联系人
    ws['A6'] = '联系人：'
    ws['A6'].font = font10
    ws['C6'] = supplier.get('contact', '')
    ws['C6'].font = font10
    ws['G6'] = '联系人：'
    ws['G6'].font = font10
    ws['H6'] = buyer.get('contact', '')
    ws['H6'].font = font10

    # Row 7: 电话
    ws['A7'] = '电话：'
    ws['A7'].font = font10
    ws['C7'] = supplier.get('phone', '')
    ws['C7'].font = font10
    ws['G7'] = '电话：'
    ws['G7'].font = font10
    ws['H7'] = buyer.get('phone', '')
    ws['H7'].font = font10

    # Row 9: Section label
    ws['A9'] = '一、项目名称、规格型号、数量、金额'
    ws['A9'].font = font10

    # ══════════════════════════════════════════════════════════════════
    # TABLE HEADER (row 11, 1-indexed)
    # ══════════════════════════════════════════════════════════════════
    headers_main = ['产品名称', 'FNSKU', '规格型号', '单位', '数量',
                    '箱率', '含税单价/元', '包装尺寸/CM', '外箱净重/KG',
                    '外箱毛重/KG', '总额/元']

    for ci, h in enumerate(headers_main, 1):
        cell = ws.cell(row=11, column=ci, value=h)
        cell.font = font10
        cell.alignment = align_cc_wrap
        # Borders: medium top/bottom; left=MEDIUM for col 1, right=MEDIUM for col 11, others THIN.
        top = _MEDIUM
        bottom = _MEDIUM
        left = _MEDIUM if ci == 1 else _THIN
        right = _MEDIUM if ci == 11 else _THIN
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Right-side calc headers (no borders) — shifted left by 1 (cols 14..18)
    calc_headers = {14: ('实重', font10, align_gc),
                    15: ('体积重', font10, align_gc),
                    16: ('海运费平摊', font10, align_gc),
                    17: ('C&F总价', font9, align_cc),
                    18: ('C&F单价', font9, align_cc)}
    for ci, (h, f, a) in calc_headers.items():
        cell = ws.cell(row=11, column=ci, value=h)
        cell.font = f
        cell.alignment = a

    # ══════════════════════════════════════════════════════════════════
    # DATA ROWS
    # ══════════════════════════════════════════════════════════════════
    row = 12
    sum_qty = 0
    sum_amt = 0.0

    for item in items:
        sku = sku_key(item)
        qty = tq.get(sku, 0)
        if qty == 0:
            continue

        amt = ta.get(sku, 0)
        unit_price = amt / qty if qty > 0 else 0
        l, w, h_ = get_lwh(item)
        def _fmt(x):
            return str(int(x)) if float(x).is_integer() else str(x)
        size_str = f'{_fmt(l)}*{_fmt(w)}*{_fmt(h_)}'

        pr = item.get('packing_rate', 1) or 1
        full_boxes = item['quantity'] / pr
        real_weight = item.get('gross_weight_kg', 0) * full_boxes
        vol_weight = (l * w * h_ / 6000) * full_boxes
        shipping = ship_alloc.get(sku, 0)
        tax_excl = amt / 1.13
        cnf_total = tax_excl + shipping
        cnf_unit = cnf_total / qty if qty > 0 else 0

        name_cn = item.get('name_cn', '')
        kb_info = kb.get(item.get('name_cn', ''), {})
        name_en = kb_info.get('english_name', '')

        # Helper to set data cell with border
        def _set_data_cell(col, value, font_=font10, alignment_=align_cc_wrap, num_fmt=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = font_
            cell.alignment = alignment_
            top = _MEDIUM
            bottom = _THIN
            left = _MEDIUM if col == 1 else _THIN
            right = _MEDIUM if col == 11 else _THIN
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        _set_data_cell(1, name_cn)
        _set_data_cell(2, sku)
        _set_data_cell(3, item.get('spec', ''))
        _set_data_cell(4, item.get('unit', '件'))
        _set_data_cell(5, qty, num_fmt='#,##0')
        _set_data_cell(6, int(pr))
        _set_data_cell(7, round(unit_price, 2), num_fmt='0.00_ ')
        _set_data_cell(8, size_str)
        _set_data_cell(9, item.get('net_weight_kg', 0), num_fmt='0.00_ ')
        _set_data_cell(10, item.get('gross_weight_kg', 0), num_fmt='0.00_ ')
        _set_data_cell(11, round(amt, 2), num_fmt='0.0000_ ')

        # Right-side calculation columns (no borders) — cols 14..18
        c = ws.cell(row=row, column=14, value=round(real_weight, 2))
        c.font = font10
        c.alignment = align_lc
        c = ws.cell(row=row, column=15, value=round(vol_weight, 2))
        c.font = font9
        c.alignment = align_lc
        c = ws.cell(row=row, column=16, value=round(shipping, 2))
        c.font = font9
        c.alignment = align_lc
        c = ws.cell(row=row, column=17, value=round(cnf_total, 2))
        c.font = font9
        c.alignment = align_lc
        c = ws.cell(row=row, column=18, value=round(cnf_unit, 2))
        c.font = font10
        c.alignment = align_cc

        sum_qty += qty
        sum_amt += amt
        row += 1

    # ══════════════════════════════════════════════════════════════════
    # TOTALS ROWS
    # ══════════════════════════════════════════════════════════════════
    first_total_row = row  # 小写合计

    # Helper for total-section cells with borders
    def _total_cell(r, col, value, font_=font10, alignment_=align_cc_wrap, num_fmt=None):
        cell = ws.cell(row=r, column=col, value=value)
        cell.font = font_
        cell.alignment = alignment_
        top = _MEDIUM
        bottom = _THIN
        left = _MEDIUM if col == 1 else _THIN
        right = _MEDIUM if col == 11 else _THIN
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # Row: 小写合计 — merge A:D (cols 1-4), qty in col E (5)
    _total_cell(first_total_row, 1, '小写合计')
    for ci in range(2, 5):
        _total_cell(first_total_row, ci, '')
    _total_cell(first_total_row, 5, sum_qty)
    for ci in range(6, 12):
        _total_cell(first_total_row, ci, '')
    ws.merge_cells(start_row=first_total_row, start_column=1, end_row=first_total_row, end_column=4)

    # Right-side: 总重 (cols 13..15)
    c = ws.cell(row=first_total_row, column=13, value='总重')
    c.font = font9
    c = ws.cell(row=first_total_row, column=14, value=round(total_gross, 1))
    c.font = font9
    c.alignment = align_lc
    c = ws.cell(row=first_total_row, column=15, value=round(total_vol, 4))
    c.font = font9
    c.alignment = align_lc

    # Row: 大写合计 — merge A:D (cols 1-4)
    row2 = first_total_row + 1
    _total_cell(row2, 1, '大写合计')
    for ci in range(2, 5):
        _total_cell(row2, ci, '')
    _total_cell(row2, 5, '')
    for ci in range(6, 12):
        _total_cell(row2, ci, '')
    ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=4)

    # Right-side: 总海运费
    c = ws.cell(row=row2, column=13, value='总海运费')
    c.font = font9
    c = ws.cell(row=row2, column=14, value=round(total_ship, 2))
    c.font = font9
    c.alignment = Alignment(horizontal='left', vertical='bottom')

    # Empty row with borders (cols 1..11)
    row3 = first_total_row + 2
    for ci in range(1, 12):
        cell = ws.cell(row=row3, column=ci, value='')
        cell.font = font10
        top = _MEDIUM
        bottom = _THIN
        left = _MEDIUM if ci == 1 else _THIN
        right = _MEDIUM if ci == 11 else _THIN
        cell.alignment = align_cc_wrap if ci <= 7 else align_gc
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # C&F total (right side)
    c = ws.cell(row=row3, column=14, value=round(sum_amt / 1.13 + total_ship, 2))
    c.font = font10
    c.alignment = Alignment(horizontal='left', vertical='bottom')

    # Row: 共计 — col 1 label, cols 2-10 merged blank, col 11 total
    row4 = first_total_row + 3
    cell = ws.cell(row=row4, column=1, value='共计')
    cell.font = font10
    cell.alignment = align_cc_wrap
    cell.border = Border(top=_THIN, bottom=_MEDIUM, left=_MEDIUM, right=_THIN)

    # Merged cols 2..10 — set cells BEFORE merging
    for ci in range(2, 11):
        cell = ws.cell(row=row4, column=ci, value='')
        cell.font = font10
        cell.alignment = align_cc_wrap
        left = _THIN if ci == 2 else _NONE
        right = _THIN if ci == 10 else _NONE
        cell.border = Border(top=_THIN, bottom=_MEDIUM, left=left, right=right)
    ws.merge_cells(start_row=row4, start_column=2, end_row=row4, end_column=10)

    # Col 11: total amount
    cell = ws.cell(row=row4, column=11, value=round(sum_amt, 2))
    cell.font = font10
    cell.alignment = align_cc_wrap
    cell.border = Border(top=_THIN, bottom=_MEDIUM, left=_THIN, right=_MEDIUM)

    # ══════════════════════════════════════════════════════════════════
    # ROW HEIGHTS
    # ══════════════════════════════════════════════════════════════════
    # Header area (fixed rows 1-10)
    fixed_heights = {
        1: 42, 2: 20, 3: 20, 4: 20, 5: 27, 6: 20, 7: 20, 8: 20, 9: 20, 10: 20, 11: 20,
    }
    for r, h in fixed_heights.items():
        ws.row_dimensions[r].height = h

    # Data rows (row 12 to last data row): 24pt each
    for r in range(12, first_total_row):
        ws.row_dimensions[r].height = 24

    # Total rows
    ws.row_dimensions[first_total_row].height = 13.5      # 小写合计
    ws.row_dimensions[first_total_row + 1].height = 13.5   # 大写合计
    ws.row_dimensions[first_total_row + 2].height = 25     # empty
    ws.row_dimensions[first_total_row + 3].height = 25     # 共计

    # ══════════════════════════════════════════════════════════════════
    # CONTRACT TERMS (below totals)
    # ══════════════════════════════════════════════════════════════════
    terms_start = first_total_row + 5  # leave one blank row

    contract_terms = [
        ('一、交期：自合同签订日起15个自然日内。', 20, 'left', 'bottom', False),
        ('二、账期：下单支付30%订金，验货合格后出货前支付剩余70%尾款', 20, None, 'bottom', False),
        ('三、交货地点：工厂送货至需方指定接收仓库', 20, None, 'bottom', False),
        ('四、保密协议：本协议的各项条款属于双方经营活动内容，任何一方未经对方当事人书面允许不得对外泄露。', 20, None, 'bottom', False),
        ('五、产品及验货标准：', 20, None, 'bottom', False),
        ('（1）供方所供产品应符合国家及行业有关安全、环保规定和检测标准，符合需方对款式、规格、材质、颜色、性能等要求。', 20, None, 'bottom', False),
        ('（2）法定商检产品的，供方必须在无条件提供正确有效的商检单,认证及检验报告.具体根据实际情况双方沟通而定.', 20, None, 'center', False),
        ('（3）需方收到货后会对产品进行检验，倘若有不合格例如产品规格（重量、尺寸）误差太大，脱落，严重色差等外观瑕疵，供方需免费更换合格产品给需方。无论需方是否验货，供方都应对产品质量问题负责。', 35, 'left', 'center', True),
        ('六、产品及包装标准：', 20, None, 'center', False),
        ('（1）出口标准五层或者七层双瓦楞纸箱，不能使用中转箱或废旧纸箱，且纸箱不能打钉，如因纸箱质量问题引起的所有返工费用由供方承担。', 20, None, 'center', False),
        ('（2）必须用透明无字胶带"工"字型打包。', 20, None, 'center', False),
        ('（3）对于特性产品，出口外箱上必须打上特性标志(如易碎，不能倒置，防潮等标志)。', 20, None, 'center', False),
        ('（4）产品必须通过100cm高度、一角三边六面摔箱测试。', 20, None, 'center', False),
        ('（5）包装尺寸：需按照国际包装尺寸，嵌入式泡沫，产品必须固定，整箱不能有任何晃动。', 20, None, 'center', False),
        ('（6）供方负责提供外箱、内包装、铭牌、提示贴、英文说明书等所有包材文件，所有包材供方在下单后一周内提供给需方查阅。 ', 20, None, 'center', False),
        ('（7）需方需要特殊粘贴的文件，由需方提供标签电子文件，供方负责打印及张贴在指定位置。', 20, None, 'center', False),
        ('七、售后保障：', 18, None, 'top', True),
        ('（1）如有交期延误（不可抗拒因素如：恶劣天气、环保检查、疫情等除外），双方先协商解决，如协商无果，延误一天按合同总金额的3\u2030计算供方交货期延迟的罚金。', 18, 'left', 'top', False),
        ('（2）供方所交货物因品质不良，如全部或部分不合格时，造成需方损失，供应商必须负责调换退款。', 18, 'left', 'top', False),
        ('（3）如因产品质量达不到该标准而导致需方客户受到人身意外伤害或财产损害，由供方负责赔偿。', 18, 'left', 'top', False),
        ('八、解决合同纠纷的方式：', 20, None, 'bottom', False),
        ('如因履行本合同产生争议的，双方均应协商解决；协商不成的，则通过原告方所在地法院诉讼解决。', 20, None, 'bottom', False),
        ('九、开票：', 20, None, 'bottom', False),
        ('本合同价格为含税价，供方须为需方开具合法、正式和有效的增值税发票（13个点），发票内容供方与需方沟通确认后方可进行拟定。供方承诺，如需方或任何第三方（包括但不限于政府税务机关、独立审计机构）在任何时候发现供方开具的发票不符合要求， 供方应立即重新为需方开具符合要求的发票。如需方因供方开具的发票不符合要求而受到有权机关处罚，供方需全额赔偿需方因该处罚而受到的全部损失（包括但不限于因票据问题导致需方无法抵扣退税的税款，以及由此产生的须由需方支付的滞纳金、行政罚款等）。', 54, 'left', 'bottom', True),
        ('十、本合同壹式贰份，双方各执壹份', 20, None, 'bottom', False),
    ]

    # Merge row for long wrapped text: rows 28 and 44 in reference are merged A:L
    merge_term_indices = [7, 23]  # indices in contract_terms that need merge (row 28 and row 44)

    for i, (text, height, ha, va, wrap) in enumerate(contract_terms):
        r = terms_start + i
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font10
        alignment_kwargs = {'vertical': va}
        if ha:
            alignment_kwargs['horizontal'] = ha
        if wrap:
            alignment_kwargs['wrap_text'] = True
        cell.alignment = Alignment(**alignment_kwargs)
        ws.row_dimensions[r].height = height

        if i in merge_term_indices:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)

    # ══════════════════════════════════════════════════════════════════
    # SIGNATURES (4 rows after last term)
    # ══════════════════════════════════════════════════════════════════
    sig_start = terms_start + len(contract_terms) + 3  # 3 blank rows

    ws.cell(row=sig_start, column=1, value='供方：').font = font10
    ws.cell(row=sig_start, column=7, value='需方：').font = font10
    ws.cell(row=sig_start, column=8, value=buyer.get('name', '')).font = font10

    ws.cell(row=sig_start + 2, column=1, value='盖章：').font = font10
    ws.cell(row=sig_start + 2, column=7, value='盖章：').font = font10

    # Set row heights for blank/signature rows
    for r in range(terms_start + len(contract_terms), sig_start + 4):
        ws.row_dimensions[r].height = 20

    # ══════════════════════════════════════════════════════════════════
    # GRAY BACKGROUND:
    #   1) columns M (13) and beyond — all rows
    #   2) "盖章" row (sig_start+2) and below — all columns
    # ══════════════════════════════════════════════════════════════════
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    stamp_row = sig_start + 2 + 1  # first row AFTER "盖章："
    max_row = ws.max_row
    max_col = ws.max_column
    if max_col < 18:
        max_col = 18
    for r in range(1, max_row + 1):
        if r >= stamp_row:
            # "盖章" row and below: gray for ALL columns
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = gray_fill
        else:
            # Above "盖章": gray only for columns L (12) and beyond (right of main 11-col table)
            for c in range(12, max_col + 1):
                ws.cell(row=r, column=c).fill = gray_fill

    fn = f'【{cno}】{suffix}出货合同.xlsx'
    fp = os.path.join(out_dir, fn)
    wb.save(fp)
    return fp
